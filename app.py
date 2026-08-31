from __future__ import annotations

import argparse
import base64
import csv
import hashlib
import html
import imaplib
import json
import mimetypes
import os
import re
import secrets
import ssl
import subprocess
import tempfile
import threading
import unicodedata
import webbrowser
from dataclasses import dataclass
from email import policy as email_policy
from email.parser import BytesParser
from email.utils import parseaddr, parsedate_to_datetime
from http.cookies import SimpleCookie
from io import BytesIO
from datetime import date, datetime, timedelta
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import parse_qs, quote, unquote, urlencode, urlparse
from urllib.error import HTTPError
from urllib.request import Request, urlopen

from pdf_extractor import (
    PdfExtractionError,
    build_cropped_pdf,
    build_submission_pdf,
    extract_pdf_certificates,
)
from database import (
    DatabaseConfigError,
    check_database_connection,
    load_database_config,
    redacted_database_status,
)
from inventory import InventoryError, InventoryStore
from cloud_storage import (
    CloudStorageConfigError,
    check_storage_connection,
    delete_file as delete_cloud_file,
    load_storage_config,
    redacted_storage_status,
    upload_file as upload_cloud_file,
)


BASE_DIR = Path(__file__).resolve().parent
STATIC_DIR = BASE_DIR / "static"


def load_environment_files() -> None:
    for name in (".env", ".env.local"):
        path = BASE_DIR / name
        if not path.exists():
            continue
        for raw_line in path.read_text(encoding="utf-8-sig").splitlines():
            line = raw_line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, value = line.split("=", 1)
            key = key.strip()
            value = value.strip().strip('"').strip("'")
            if key and key not in os.environ:
                os.environ[key] = value


load_environment_files()


def configured_data_dir() -> Path:
    configured = os.environ.get("ETC_DATA_DIR", "").strip()
    if configured:
        return Path(configured).expanduser()
    return BASE_DIR / "app_data"


APP_DATA_DIR = configured_data_dir()
INVENTORY = InventoryStore(APP_DATA_DIR, BASE_DIR / "sql" / "cloud_sql_schema.sql")
SETTINGS_FILE = APP_DATA_DIR / "settings.json"
MANAGED_IMPORT_DIR = APP_DATA_DIR / "imports"
VEHICLES_FILE = APP_DATA_DIR / "vehicles.json"
CERTIFICATES_FILE = APP_DATA_DIR / "certificates.json"
USERS_FILE = APP_DATA_DIR / "users.json"
AUDIT_LOG_FILE = APP_DATA_DIR / "audit_log.jsonl"
USER_AVATAR_DIR = APP_DATA_DIR / "user_avatars"
VEHICLE_PHOTO_DIR = APP_DATA_DIR / "vehicle_photos"
VEHICLE_INSPECTION_DIR = APP_DATA_DIR / "vehicle_inspections"
SUBMISSIONS_DIR = APP_DATA_DIR / "submissions"
CERTIFICATE_EXPORTS_DIR = APP_DATA_DIR / "certificate_exports"
SUBMISSIONS_FILE = APP_DATA_DIR / "submissions.json"
LOGISTICS_JOBS_FILE = APP_DATA_DIR / "logistics_jobs.json"
LOGISTICS_RATE_MASTER_FILE = APP_DATA_DIR / "logistics_rate_master.json"
SUBCONTRACTORS_FILE = APP_DATA_DIR / "subcontractors.json"
RETURN_SHIPMENTS_FILE = APP_DATA_DIR / "return_shipments.json"
RETURN_SHIPMENT_EXPORTS_DIR = APP_DATA_DIR / "return_shipments"
MAIL_IMPORTS_FILE = APP_DATA_DIR / "mail_imports.json"
MAIL_ATTACHMENTS_DIR = APP_DATA_DIR / "mail_attachments"
MAIL_SETTINGS_FILE = APP_DATA_DIR / "mail_settings.json"
OUTLOOK_TOKENS_FILE = APP_DATA_DIR / "outlook_tokens.json"
CLOUD_STORAGE_SETTINGS_FILE = APP_DATA_DIR / "cloud_storage_settings.json"
LOGISTICS_EXCEL_EXTRACT_SCRIPT = BASE_DIR / "tools" / "extract_logistics_excel.ps1"
MAX_UPLOAD_SIZE = 100 * 1024 * 1024
MAX_PHOTO_SIZE = 10 * 1024 * 1024
DEFAULT_ADMIN_USER = "yousuke-iida@ithe.co.jp"
DEFAULT_INITIAL_PASSWORD = "change-this-password"
PASSWORD_RESET_KEY_ENV = "ETC_PASSWORD_RESET_KEY"
PASSWORD_ITERATIONS = 240_000
SESSION_COOKIE_NAME = "etc_session"
SESSION_TTL_SECONDS = int(os.environ.get("ETC_SESSION_TTL_SECONDS", str(12 * 60 * 60)))
SESSIONS: dict[str, dict[str, Any]] = {}
SESSIONS_LOCK = threading.Lock()
OUTLOOK_OAUTH_STATES: dict[str, dict[str, Any]] = {}
OUTLOOK_OAUTH_LOCK = threading.Lock()
IMAP_IMPORT_LOCK = threading.Lock()
AUTO_IMPORT_THREAD: threading.Thread | None = None
AUTO_IMPORT_STOP = threading.Event()
AUTO_IMPORT_STATUS: dict[str, Any] = {
    "enabled": False,
    "interval_seconds": 0,
    "running": False,
    "last_started_at": "",
    "last_finished_at": "",
    "next_run_at": "",
    "last_summary": {},
    "last_error": "",
}
AUTO_IMPORT_STATUS_LOCK = threading.Lock()
DEFAULT_IMAP_SUBJECT_KEYWORDS = (
    "申請書",
    "商品交換業務依頼",
    "商品交換作業依頼",
    "商品回収業務依頼",
    "商品手配依頼",
    "商品交換手配",
)
DEFAULT_IMAP_EXCLUDE_SUBJECT_KEYWORDS = (
    "事前確認",
    "完了報告",
    "完了連絡",
    "作業完了",
    "対応完了",
    "キャンセル",
    "中止",
)


def env_int(name: str, default: int) -> int:
    try:
        return int(os.environ.get(name, str(default)))
    except ValueError:
        return default


def env_list(*names: str) -> tuple[str, ...]:
    value = ""
    for name in names:
        configured = os.environ.get(name, "").strip()
        if configured:
            value = configured
            break
    if not value:
        return ()
    return tuple(
        item.strip()
        for item in re.split(r"[,;\r\n]+", value)
        if item.strip()
    )


def merge_keyword_filters(
    defaults: tuple[str, ...],
    *names: str,
) -> tuple[str, ...]:
    merged: list[str] = []
    seen: set[str] = set()
    for item in (*defaults, *env_list(*names)):
        key = item.casefold()
        if key in seen:
            continue
        seen.add(key)
        merged.append(item)
    return tuple(merged)


def auth_credentials() -> tuple[str, str] | None:
    user = (
        os.environ.get("ETC_LOGIN_USER")
        or os.environ.get("ETC_USER")
        or ""
    ).strip()
    password = os.environ.get("ETC_LOGIN_PASSWORD") or os.environ.get("ETC_PASSWORD") or ""
    if user and password:
        return user, password
    return None


def login_is_required() -> bool:
    return os.environ.get("ETC_REQUIRE_LOGIN", "").strip().lower() in {
        "1",
        "true",
        "yes",
        "on",
    }


def basic_auth_is_valid(header_value: str) -> bool:
    return authenticated_user(header_value) is not None


def parse_basic_auth(header_value: str) -> tuple[str, str] | None:
    if not header_value.startswith("Basic "):
        return None
    try:
        decoded = base64.b64decode(header_value.removeprefix("Basic ").strip()).decode(
            "utf-8"
        )
    except (ValueError, UnicodeDecodeError):
        return None
    if ":" not in decoded:
        return None
    supplied_user, supplied_password = decoded.split(":", 1)
    return supplied_user, supplied_password


def hash_password(password: str, salt: str | None = None) -> str:
    salt = salt or secrets.token_hex(16)
    digest = hashlib.pbkdf2_hmac(
        "sha256",
        password.encode("utf-8"),
        salt.encode("ascii"),
        PASSWORD_ITERATIONS,
    ).hex()
    return f"pbkdf2_sha256${PASSWORD_ITERATIONS}${salt}${digest}"


def password_matches(password: str, stored_hash: str) -> bool:
    try:
        algorithm, iterations, salt, expected_digest = stored_hash.split("$", 3)
    except ValueError:
        return False
    if algorithm != "pbkdf2_sha256":
        return False
    try:
        iteration_count = int(iterations)
    except ValueError:
        return False
    actual_digest = hashlib.pbkdf2_hmac(
        "sha256",
        password.encode("utf-8"),
        salt.encode("ascii"),
        iteration_count,
    ).hex()
    return secrets.compare_digest(actual_digest, expected_digest)


def initial_admin_credentials() -> tuple[str, str]:
    user = (
        os.environ.get("ETC_ADMIN_USER")
        or os.environ.get("ETC_LOGIN_USER")
        or os.environ.get("ETC_USER")
        or DEFAULT_ADMIN_USER
    ).strip()
    password = (
        os.environ.get("ETC_ADMIN_PASSWORD")
        or os.environ.get("ETC_LOGIN_PASSWORD")
        or os.environ.get("ETC_PASSWORD")
        or DEFAULT_INITIAL_PASSWORD
    )
    return user, password


def configured_admin_id() -> str:
    return normalize_user_id(initial_admin_credentials()[0] or DEFAULT_ADMIN_USER)


def password_reset_key() -> str:
    return os.environ.get(PASSWORD_RESET_KEY_ENV, "").strip()


def normalize_user_id(user_id: str) -> str:
    return user_id.strip().lower()


def public_user(user: dict[str, Any]) -> dict[str, Any]:
    user_id = str(user.get("id", ""))
    avatar_file = str(user.get("avatar_file", ""))
    contractor_code = str(user.get("contractor_code", "")).strip()
    return {
        "id": user_id,
        "role": str(user.get("role", "user")),
        "contractor_code": contractor_code,
        "company_name": str(user.get("company_name", "")),
        "created_at": str(user.get("created_at", "")),
        "updated_at": str(user.get("updated_at", "")),
        "avatar_url": (
            f"/api/user-avatar?user={quote(normalize_user_id(user_id), safe='')}"
            if avatar_file
            else ""
        ),
    }


def ensure_user_store() -> dict[str, Any]:
    users = load_json_store(USERS_FILE, {})
    if isinstance(users, dict) and users:
        admin_user, admin_password = initial_admin_credentials()
        normalized_admin = normalize_user_id(admin_user)
        changed = False
        now = datetime.now().isoformat(timespec="seconds")
        if normalized_admin and normalized_admin not in users and admin_password:
            users[normalized_admin] = {
                "id": admin_user,
                "role": "admin",
                "password_hash": hash_password(admin_password),
                "created_at": now,
                "updated_at": now,
            }
            append_audit(
                "ensure_admin_account",
                "system",
                normalized_admin,
                {"role": "admin"},
            )
            changed = True

        reset_password = os.environ.get("ETC_RESET_ADMIN_PASSWORD", "")
        if reset_password and normalized_admin in users:
            users[normalized_admin]["password_hash"] = hash_password(reset_password)
            users[normalized_admin]["role"] = "admin"
            users[normalized_admin]["updated_at"] = now
            append_audit(
                "reset_admin_password",
                "system",
                normalized_admin,
                {"source": "environment"},
            )
            changed = True

        if changed:
            save_json_store(USERS_FILE, users)
        return users

    admin_user, admin_password = initial_admin_credentials()
    now = datetime.now().isoformat(timespec="seconds")
    users = {
        normalize_user_id(admin_user): {
            "id": admin_user,
            "role": "admin",
            "password_hash": hash_password(admin_password),
            "created_at": now,
            "updated_at": now,
        }
    }
    save_json_store(USERS_FILE, users)
    append_audit(
        "bootstrap_admin",
        "system",
        normalize_user_id(admin_user),
        {"role": "admin"},
    )
    return users


def load_users() -> dict[str, Any]:
    users = load_json_store(USERS_FILE, {})
    return users if isinstance(users, dict) else {}


def save_users(users: dict[str, Any]) -> None:
    save_json_store(USERS_FILE, users)


def authenticated_user(header_value: str) -> dict[str, Any] | None:
    credentials = auth_credentials()
    supplied = parse_basic_auth(header_value)

    if not credentials and not login_is_required() and not USERS_FILE.exists():
        return {"id": "local", "role": "admin"}

    if supplied is None:
        return None

    supplied_user, supplied_password = supplied
    users = load_users()
    if not users and (login_is_required() or USERS_FILE.exists()):
        users = ensure_user_store()

    if users:
        stored_user = users.get(normalize_user_id(supplied_user))
        if not isinstance(stored_user, dict):
            return None
        if not password_matches(
            supplied_password,
            str(stored_user.get("password_hash", "")),
        ):
            return None
        return public_user(stored_user)

    if credentials:
        expected_user, expected_password = credentials
        if secrets.compare_digest(
            supplied_user,
            expected_user,
        ) and secrets.compare_digest(supplied_password, expected_password):
            return {"id": expected_user, "role": "admin"}

    return None


def authenticate_credentials(user_id: str, password: str) -> dict[str, Any] | None:
    users = load_users()
    if not users and (login_is_required() or USERS_FILE.exists()):
        users = ensure_user_store()

    if users:
        stored_user = users.get(normalize_user_id(user_id))
        if not isinstance(stored_user, dict):
            return None
        if password_matches(password, str(stored_user.get("password_hash", ""))):
            return public_user(stored_user)
        return None

    credentials = auth_credentials()
    if credentials:
        expected_user, expected_password = credentials
        if secrets.compare_digest(user_id, expected_user) and secrets.compare_digest(
            password,
            expected_password,
        ):
            return {"id": expected_user, "role": "admin"}

    if not login_is_required() and not USERS_FILE.exists():
        return {"id": "local", "role": "admin"}

    return None


def create_session(user: dict[str, Any]) -> str:
    token = secrets.token_urlsafe(32)
    expires_at = datetime.now() + timedelta(seconds=SESSION_TTL_SECONDS)
    with SESSIONS_LOCK:
        SESSIONS[token] = {
            "user": user,
            "expires_at": expires_at.timestamp(),
        }
    return token


def session_cookie_header(token: str) -> str:
    max_age = max(60, SESSION_TTL_SECONDS)
    return (
        f"{SESSION_COOKIE_NAME}={token}; Path=/; Max-Age={max_age}; "
        "SameSite=Lax; HttpOnly"
    )


def expired_session_cookie_header() -> str:
    return (
        f"{SESSION_COOKIE_NAME}=; Path=/; Max-Age=0; "
        "SameSite=Lax; HttpOnly"
    )


def session_token_from_cookie(cookie_header: str) -> str:
    cookie = SimpleCookie()
    try:
        cookie.load(cookie_header)
    except Exception:
        return ""
    morsel = cookie.get(SESSION_COOKIE_NAME)
    return morsel.value if morsel else ""


def user_from_session_token(token: str) -> dict[str, Any] | None:
    if not token:
        return None
    with SESSIONS_LOCK:
        session = SESSIONS.get(token)
        if not session:
            return None
        expires_at = float(session.get("expires_at", 0))
        if expires_at < datetime.now().timestamp():
            SESSIONS.pop(token, None)
            return None
        user = session.get("user")
    return user if isinstance(user, dict) else None


def delete_session(token: str) -> None:
    if not token:
        return
    with SESSIONS_LOCK:
        SESSIONS.pop(token, None)


def update_session_user(token: str, user: dict[str, Any]) -> None:
    if not token:
        return
    with SESSIONS_LOCK:
        session = SESSIONS.get(token)
        if session:
            session["user"] = user


def create_user(
    user_id: str,
    password: str,
    role: str = "user",
    actor: str = "system",
    contractor_code: str = "",
    company_name: str = "",
) -> dict[str, Any]:
    normalized = normalize_user_id(user_id)
    role = role if role in {"admin", "user", "contractor"} else "user"
    if not normalized:
        raise DashboardError("ログインIDを入力してください。")
    if len(password) < 8:
        raise DashboardError("パスワードは8文字以上にしてください。")
    if role == "contractor" and not contractor_code.strip():
        contractor_code = user_id

    users = ensure_user_store()
    if normalized in users:
        raise DashboardError("同じログインIDのアカウントが既にあります。")

    now = datetime.now().isoformat(timespec="seconds")
    users[normalized] = {
        "id": user_id.strip(),
        "role": role,
        "contractor_code": contractor_code.strip(),
        "company_name": company_name.strip(),
        "password_hash": hash_password(password),
        "created_at": now,
        "updated_at": now,
    }
    save_users(users)
    append_audit("create_user", actor, normalized, {"role": role})
    return public_user(users[normalized])


def change_user_password(
    user_id: str,
    new_password: str,
    actor: str,
    current_password: str = "",
    admin_override: bool = False,
) -> dict[str, Any]:
    normalized = normalize_user_id(user_id)
    if len(new_password) < 8:
        raise DashboardError("新しいパスワードは8文字以上にしてください。")

    users = ensure_user_store()
    stored_user = users.get(normalized)
    if not isinstance(stored_user, dict):
        raise DashboardError("アカウントが見つかりません。")

    if not admin_override and not password_matches(
        current_password,
        str(stored_user.get("password_hash", "")),
    ):
        raise DashboardError("現在のパスワードが違います。")

    stored_user["password_hash"] = hash_password(new_password)
    stored_user["updated_at"] = datetime.now().isoformat(timespec="seconds")
    save_users(users)
    append_audit("change_password", actor, normalized, {"admin_override": admin_override})
    return public_user(stored_user)


def reset_user_password_with_key(
    user_id: str,
    new_password: str,
    reset_key: str,
    actor: str = "login_screen",
) -> dict[str, Any]:
    expected_key = password_reset_key()
    if not expected_key:
        raise DashboardError(
            "パスワードリセットキーが未設定です。NASのdocker-compose.ymlにETC_PASSWORD_RESET_KEYを設定してください。"
        )
    if not secrets.compare_digest(reset_key.strip(), expected_key):
        raise DashboardError("リセットキーが違います。")
    if len(new_password) < 8:
        raise DashboardError("新しいパスワードは8文字以上にしてください。")

    normalized = normalize_user_id(user_id)
    if not normalized:
        raise DashboardError("ログインIDを入力してください。")

    users = ensure_user_store()
    stored_user = users.get(normalized)
    if not isinstance(stored_user, dict):
        raise DashboardError("アカウントが見つかりません。")

    stored_user["password_hash"] = hash_password(new_password)
    if normalized == configured_admin_id():
        stored_user["role"] = "admin"
    stored_user["updated_at"] = datetime.now().isoformat(timespec="seconds")
    save_users(users)
    append_audit(
        "reset_password",
        actor,
        normalized,
        {
            "source": "login_screen",
            "restored_admin": normalized == configured_admin_id(),
        },
    )
    return public_user(stored_user)


def save_user_avatar(user_id: str, file_name: str, body: bytes) -> dict[str, Any]:
    normalized = normalize_user_id(user_id)
    users = ensure_user_store()
    stored_user = users.get(normalized)
    if not isinstance(stored_user, dict):
        raise DashboardError("アカウントが見つかりません。")

    suffix = Path(unquote(file_name)).suffix.lower()
    if suffix not in {".jpg", ".jpeg", ".png", ".webp"}:
        raise DashboardError("サムネイルはJPG、PNG、WEBPを選択してください。")
    if not body:
        raise DashboardError("選択した画像ファイルが空です。")
    if len(body) > MAX_PHOTO_SIZE:
        raise DashboardError("サムネイルは10MB以下にしてください。")

    safe_user = hashlib.sha256(normalized.encode("utf-8")).hexdigest()[:20]
    USER_AVATAR_DIR.mkdir(parents=True, exist_ok=True)
    for old_path in USER_AVATAR_DIR.glob(f"{safe_user}.*"):
        delete_path_from_cloud(old_path)
        old_path.unlink(missing_ok=True)
    avatar_name = f"{safe_user}{suffix}"
    avatar_path = USER_AVATAR_DIR / avatar_name
    avatar_path.write_bytes(body)
    sync_path_to_cloud(avatar_path)

    stored_user["avatar_file"] = avatar_name
    stored_user["updated_at"] = datetime.now().isoformat(timespec="seconds")
    save_users(users)
    return public_user(stored_user)


def resolve_user_avatar(user_id: str) -> Path:
    normalized = normalize_user_id(user_id)
    users = load_users()
    stored_user = users.get(normalized, {}) if isinstance(users, dict) else {}
    avatar_name = Path(str(stored_user.get("avatar_file", ""))).name
    if not avatar_name:
        raise DashboardError("サムネイルが登録されていません。")
    avatar_path = (USER_AVATAR_DIR / avatar_name).resolve()
    if not avatar_path.is_relative_to(USER_AVATAR_DIR.resolve()) or not avatar_path.is_file():
        raise DashboardError("サムネイルが見つかりません。")
    return avatar_path


def append_audit(
    action: str,
    user_id: str,
    target: str = "",
    details: dict[str, Any] | None = None,
) -> None:
    try:
        APP_DATA_DIR.mkdir(parents=True, exist_ok=True)
        entry = {
            "at": datetime.now().isoformat(timespec="seconds"),
            "user_id": user_id,
            "action": action,
            "target": target,
            "details": details or {},
        }
        with AUDIT_LOG_FILE.open("a", encoding="utf-8") as file:
            file.write(json.dumps(entry, ensure_ascii=False) + "\n")
    except OSError:
        return


def read_audit_log(limit: int = 100) -> list[dict[str, Any]]:
    if not AUDIT_LOG_FILE.exists():
        return []
    try:
        lines = AUDIT_LOG_FILE.read_text(encoding="utf-8").splitlines()
    except OSError:
        return []
    entries: list[dict[str, Any]] = []
    for line in lines[-limit:]:
        try:
            entry = json.loads(line)
        except json.JSONDecodeError:
            continue
        if isinstance(entry, dict):
            entries.append(entry)
    return entries

CSV_HEADERS = {
    "date_start": "利用年月日（自）",
    "time_start": "時分（自）",
    "date_end": "利用年月日（至）",
    "time_end": "時分（至）",
    "entry_ic": "利用ＩＣ（自）",
    "exit_ic": "利用ＩＣ（至）",
    "pre_discount_fee": "割引前料金",
    "discount": "ＥＴＣ割引額",
    "toll_fee": "通行料金",
    "reduction_target_fee": "還元額適用料金",
    "postpaid_fee": "後納料金",
    "vehicle_type": "車種",
    "vehicle_number": "車両番号",
    "card_number": "ＥＴＣカード番号",
    "status": "備考",
}

MONEY_FIELDS = {
    "pre_discount_fee",
    "discount",
    "toll_fee",
    "reduction_target_fee",
    "postpaid_fee",
}


class DashboardError(Exception):
    """An error that can be safely shown in the dashboard."""


def corrupted_text_paths(value: Any, prefix: str = "") -> list[str]:
    """Return field paths containing replacement characters or invalid Unicode."""
    paths: list[str] = []
    if isinstance(value, dict):
        for key, item in value.items():
            path = f"{prefix}.{key}" if prefix else str(key)
            paths.extend(corrupted_text_paths(item, path))
    elif isinstance(value, (list, tuple)):
        for index, item in enumerate(value):
            path = f"{prefix}[{index}]" if prefix else f"[{index}]"
            paths.extend(corrupted_text_paths(item, path))
    elif isinstance(value, str):
        has_invalid_control = any(
            ord(character) < 32 and character not in "\n\r\t"
            for character in value
        )
        has_surrogate = any(0xD800 <= ord(character) <= 0xDFFF for character in value)
        if "\ufffd" in value or "\x00" in value or has_invalid_control or has_surrogate:
            paths.append(prefix or "value")
    return paths


def corrupted_text_message(value: Any, context: str) -> str:
    paths = corrupted_text_paths(value)
    if not paths:
        return ""
    visible = ", ".join(paths[:5])
    suffix = " ほか" if len(paths) > 5 else ""
    return f"{context}に文字化けまたは不正な文字を検出しました: {visible}{suffix}"


def raise_for_corrupted_text(value: Any, context: str) -> None:
    message = corrupted_text_message(value, context)
    if message:
        raise DashboardError(message)


def load_settings() -> dict[str, str]:
    env_folder = os.environ.get("ETC_IMPORT_FOLDER", "").strip()
    if env_folder:
        return {"import_folder": env_folder}

    if not SETTINGS_FILE.exists():
        return {"import_folder": ""}

    try:
        data = json.loads(SETTINGS_FILE.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {"import_folder": ""}

    return {"import_folder": str(data.get("import_folder", "")).strip()}


def save_settings(import_folder: str) -> dict[str, str]:
    folder = Path(import_folder).expanduser()
    if not folder.exists():
        raise DashboardError("指定したフォルダーが見つかりません。")
    if not folder.is_dir():
        raise DashboardError("フォルダーを指定してください。")

    resolved = str(folder.resolve())
    APP_DATA_DIR.mkdir(parents=True, exist_ok=True)
    SETTINGS_FILE.write_text(
        json.dumps({"import_folder": resolved}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return {"import_folder": resolved}


def load_json_store(path: Path, default: Any) -> Any:
    if not path.exists():
        return default
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return default


def save_json_store(path: Path, data: Any) -> None:
    APP_DATA_DIR.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(data, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def load_saved_cloud_storage_settings() -> dict[str, Any]:
    data = load_json_store(CLOUD_STORAGE_SETTINGS_FILE, {})
    return data if isinstance(data, dict) else {}


def load_cloud_storage_config():
    return load_storage_config(settings=load_saved_cloud_storage_settings())


def cloud_storage_status_payload(check: bool = False) -> dict[str, Any]:
    config = load_cloud_storage_config()
    payload: dict[str, Any] = {"storage": redacted_storage_status(config)}
    if check:
        payload["check"] = check_storage_connection(config)
    return payload


def save_cloud_storage_settings(payload: dict[str, Any], actor: str = "") -> dict[str, Any]:
    backend = str(payload.get("backend", "")).strip().lower() or "local"
    bucket = str(payload.get("bucket", "")).strip()
    prefix = str(payload.get("prefix", "")).strip()
    config = load_storage_config(
        settings={
            "backend": backend,
            "bucket": bucket,
            "prefix": prefix,
        }
    )
    settings = {
        "backend": config.backend,
        "bucket": config.bucket,
        "prefix": config.prefix,
        "updated_at": datetime.now().isoformat(timespec="seconds"),
        "updated_by": actor,
    }
    save_json_store(CLOUD_STORAGE_SETTINGS_FILE, settings)
    append_audit(
        "save_storage_settings",
        actor,
        "gcs",
        {
            "backend": settings["backend"],
            "bucket": settings["bucket"],
            "prefix": settings["prefix"],
        },
    )
    return redacted_storage_status(load_cloud_storage_config())


def sync_path_to_cloud(path: Path) -> dict[str, Any]:
    try:
        config = load_cloud_storage_config()
        if not config.enabled:
            return {"enabled": False, "uploaded": False}
        return upload_cloud_file(path, APP_DATA_DIR, config)
    except Exception as exc:
        print(f"[ETC] Google Cloud Storage upload skipped: {exc}")
        return {"enabled": True, "uploaded": False, "error": str(exc)}


def delete_path_from_cloud(path: Path) -> dict[str, Any]:
    try:
        config = load_cloud_storage_config()
        if not config.enabled:
            return {"enabled": False, "deleted": False}
        return delete_cloud_file(path, APP_DATA_DIR, config)
    except Exception as exc:
        print(f"[ETC] Google Cloud Storage delete skipped: {exc}")
        return {"enabled": True, "deleted": False, "error": str(exc)}


@dataclass(frozen=True)
class OutlookMailConfig:
    tenant_id: str
    client_id: str
    client_secret: str
    redirect_uri: str
    mailbox: str
    processed_category: str
    error_category: str
    max_messages: int
    use_me_endpoint: bool

    @property
    def missing_keys(self) -> list[str]:
        missing: list[str] = []
        if not self.client_id:
            missing.append("OUTLOOK_CLIENT_ID")
        if not self.client_secret:
            missing.append("OUTLOOK_CLIENT_SECRET")
        if not self.redirect_uri:
            missing.append("OUTLOOK_REDIRECT_URI")
        return missing

    @property
    def is_configured(self) -> bool:
        return not self.missing_keys

    @property
    def token_url(self) -> str:
        return (
            "https://login.microsoftonline.com/"
            f"{quote(self.tenant_id, safe='')}/oauth2/v2.0/token"
        )

    @property
    def authorize_url(self) -> str:
        return (
            "https://login.microsoftonline.com/"
            f"{quote(self.tenant_id, safe='')}/oauth2/v2.0/authorize"
        )


def load_outlook_mail_config(default_redirect_uri: str = "") -> OutlookMailConfig:
    return OutlookMailConfig(
        tenant_id=(
            os.environ.get("OUTLOOK_TENANT_ID")
            or os.environ.get("MS_TENANT_ID")
            or "organizations"
        ).strip(),
        client_id=(
            os.environ.get("OUTLOOK_CLIENT_ID")
            or os.environ.get("MS_CLIENT_ID")
            or ""
        ).strip(),
        client_secret=(
            os.environ.get("OUTLOOK_CLIENT_SECRET")
            or os.environ.get("MS_CLIENT_SECRET")
            or ""
        ).strip(),
        redirect_uri=(
            os.environ.get("OUTLOOK_REDIRECT_URI") or default_redirect_uri
        ).strip(),
        mailbox=(os.environ.get("OUTLOOK_MAILBOX") or "info_order@ithe.co.jp").strip(),
        processed_category=(
            os.environ.get("OUTLOOK_PROCESSED_CATEGORY") or "処理済み"
        ).strip(),
        error_category=(os.environ.get("OUTLOOK_ERROR_CATEGORY") or "取込エラー").strip(),
        max_messages=max(1, min(50, env_int("OUTLOOK_MAX_MESSAGES", 10))),
        use_me_endpoint=os.environ.get("OUTLOOK_USE_ME", "1").strip().lower()
        not in {"0", "false", "no", "off"},
    )


def outlook_scope() -> str:
    configured = os.environ.get("OUTLOOK_SCOPES", "").strip()
    return configured or "offline_access https://graph.microsoft.com/Mail.ReadWrite"


def outlook_base_path(config: OutlookMailConfig) -> str:
    if config.use_me_endpoint:
        return "/me"
    return f"/users/{quote(config.mailbox, safe='')}"


def outlook_authorization_url(config: OutlookMailConfig, state: str) -> str:
    params = {
        "client_id": config.client_id,
        "response_type": "code",
        "redirect_uri": config.redirect_uri,
        "response_mode": "query",
        "scope": outlook_scope(),
        "state": state,
        "prompt": "select_account",
    }
    return f"{config.authorize_url}?{urlencode(params)}"


def outlook_redacted_status(config: OutlookMailConfig) -> dict[str, Any]:
    tokens = load_json_store(OUTLOOK_TOKENS_FILE, {})
    expires_at = float(tokens.get("expires_at", 0) or 0) if isinstance(tokens, dict) else 0
    return {
        "configured": config.is_configured,
        "missing": config.missing_keys,
        "tenant_id": config.tenant_id,
        "client_id_set": bool(config.client_id),
        "client_secret_set": bool(config.client_secret),
        "redirect_uri": config.redirect_uri,
        "mailbox": config.mailbox,
        "scope": outlook_scope(),
        "processed_category": config.processed_category,
        "error_category": config.error_category,
        "use_me_endpoint": config.use_me_endpoint,
        "authenticated": bool(isinstance(tokens, dict) and tokens.get("refresh_token")),
        "access_token_valid": expires_at > datetime.now().timestamp() + 60,
        "token_expires_at": (
            datetime.fromtimestamp(expires_at).isoformat(timespec="seconds")
            if expires_at
            else ""
        ),
    }


def create_outlook_oauth_state(actor: str) -> str:
    state = secrets.token_urlsafe(24)
    with OUTLOOK_OAUTH_LOCK:
        OUTLOOK_OAUTH_STATES[state] = {
            "actor": actor,
            "expires_at": (datetime.now() + timedelta(minutes=10)).timestamp(),
        }
    return state


def consume_outlook_oauth_state(state: str) -> dict[str, Any]:
    with OUTLOOK_OAUTH_LOCK:
        payload = OUTLOOK_OAUTH_STATES.pop(state, None)
    if not payload or float(payload.get("expires_at", 0)) < datetime.now().timestamp():
        raise DashboardError("Outlook連携の確認キーが期限切れです。もう一度連携してください。")
    return payload


def request_json(
    url: str,
    *,
    method: str = "GET",
    headers: dict[str, str] | None = None,
    body: bytes | None = None,
    timeout: int = 30,
    response_metadata: dict[str, Any] | None = None,
) -> dict[str, Any]:
    request = Request(url, data=body, method=method, headers=headers or {})
    try:
        with urlopen(request, timeout=timeout) as response:
            raw = response.read()
            http_status = int(getattr(response, "status", 0) or response.getcode() or 0)
    except HTTPError as exc:
        raw = exc.read(4096)
        summary = raw.decode("utf-8", errors="replace").strip()[:1000]
        if response_metadata is not None:
            response_metadata.update(
                http_status=int(exc.code or 0),
                response_summary=summary,
            )
        detail = summary
        try:
            parsed = json.loads(summary) if summary else {}
            if isinstance(parsed, dict):
                detail = str(parsed.get("detail") or parsed.get("error") or summary)
        except json.JSONDecodeError:
            pass
        error = DashboardError(
            f"外部サービスがHTTP {int(exc.code or 0)}を返しました"
            + (f": {detail}" if detail else "。")
        )
        error.http_status = int(exc.code or 0)  # type: ignore[attr-defined]
        error.response_summary = summary  # type: ignore[attr-defined]
        raise error from exc
    except OSError as exc:
        if response_metadata is not None:
            response_metadata.update(http_status=0, response_summary=str(exc)[:1000])
        error = DashboardError(f"外部サービスへ接続できませんでした: {exc}")
        error.http_status = 0  # type: ignore[attr-defined]
        error.response_summary = str(exc)[:1000]  # type: ignore[attr-defined]
        raise error from exc
    summary = raw.decode("utf-8", errors="replace").strip()[:1000]
    if response_metadata is not None:
        response_metadata.update(
            http_status=http_status,
            response_summary=summary,
        )
    try:
        payload = json.loads(raw.decode("utf-8")) if raw else {}
    except (UnicodeDecodeError, json.JSONDecodeError) as exc:
        error = DashboardError("外部サービスからUTF-8の正しい応答が返りませんでした。")
        error.http_status = http_status  # type: ignore[attr-defined]
        error.response_summary = summary  # type: ignore[attr-defined]
        raise error from exc
    if not isinstance(payload, dict):
        error = DashboardError("外部サービスからJSONオブジェクトが返りませんでした。")
        error.http_status = http_status  # type: ignore[attr-defined]
        error.response_summary = summary  # type: ignore[attr-defined]
        raise error
    return payload


def exchange_outlook_code(config: OutlookMailConfig, code: str) -> dict[str, Any]:
    if not config.is_configured:
        raise DashboardError("Outlook連携の環境変数が未設定です。")
    form = urlencode(
        {
            "client_id": config.client_id,
            "client_secret": config.client_secret,
            "grant_type": "authorization_code",
            "code": code,
            "redirect_uri": config.redirect_uri,
            "scope": outlook_scope(),
        }
    ).encode("utf-8")
    payload = request_json(
        config.token_url,
        method="POST",
        headers={"Content-Type": "application/x-www-form-urlencoded"},
        body=form,
    )
    if "error" in payload:
        raise DashboardError(str(payload.get("error_description") or payload["error"]))
    expires_in = int(payload.get("expires_in", 3600))
    payload["expires_at"] = datetime.now().timestamp() + max(60, expires_in - 60)
    save_json_store(OUTLOOK_TOKENS_FILE, payload)
    return payload


def refresh_outlook_token(config: OutlookMailConfig) -> dict[str, Any]:
    tokens = load_json_store(OUTLOOK_TOKENS_FILE, {})
    if not isinstance(tokens, dict) or not tokens.get("refresh_token"):
        raise DashboardError("Outlook連携がまだ完了していません。")
    if float(tokens.get("expires_at", 0) or 0) > datetime.now().timestamp() + 60:
        return tokens

    form = urlencode(
        {
            "client_id": config.client_id,
            "client_secret": config.client_secret,
            "grant_type": "refresh_token",
            "refresh_token": str(tokens.get("refresh_token", "")),
            "redirect_uri": config.redirect_uri,
            "scope": outlook_scope(),
        }
    ).encode("utf-8")
    payload = request_json(
        config.token_url,
        method="POST",
        headers={"Content-Type": "application/x-www-form-urlencoded"},
        body=form,
    )
    if "error" in payload:
        raise DashboardError(str(payload.get("error_description") or payload["error"]))
    if not payload.get("refresh_token"):
        payload["refresh_token"] = tokens.get("refresh_token")
    expires_in = int(payload.get("expires_in", 3600))
    payload["expires_at"] = datetime.now().timestamp() + max(60, expires_in - 60)
    save_json_store(OUTLOOK_TOKENS_FILE, payload)
    return payload


def outlook_graph_json(
    config: OutlookMailConfig,
    path: str,
    *,
    method: str = "GET",
    payload: dict[str, Any] | None = None,
) -> dict[str, Any]:
    if not config.is_configured:
        raise DashboardError("Outlook連携の環境変数が未設定です。")
    tokens = refresh_outlook_token(config)
    headers = {
        "Authorization": f"Bearer {tokens.get('access_token', '')}",
        "Accept": "application/json",
    }
    body = None
    if payload is not None:
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        headers["Content-Type"] = "application/json; charset=utf-8"
    return request_json(
        f"https://graph.microsoft.com/v1.0{path}",
        method=method,
        headers=headers,
        body=body,
    )


def mail_excel_suffix(name: str) -> bool:
    return Path(name).suffix.lower() in {".xlsx", ".xlsm", ".xls"}


def outlook_excel_suffix(name: str) -> bool:
    return mail_excel_suffix(name)


@dataclass(frozen=True)
class ImapMailConfig:
    host: str
    port: int
    username: str
    password: str
    inbox: str
    processed_folder: str
    error_folder: str
    max_messages: int
    use_ssl: bool
    allowed_senders: tuple[str, ...]
    subject_keywords: tuple[str, ...]
    attachment_keywords: tuple[str, ...]
    exclude_subject_keywords: tuple[str, ...]
    exclude_attachment_keywords: tuple[str, ...]

    @property
    def missing_keys(self) -> list[str]:
        missing: list[str] = []
        if not self.host:
            missing.append("IMAP_HOST")
        if not self.username:
            missing.append("IMAP_USER")
        if not self.password:
            missing.append("IMAP_PASSWORD")
        return missing

    @property
    def is_configured(self) -> bool:
        return not self.missing_keys


def load_saved_imap_mail_settings() -> dict[str, Any]:
    data = load_json_store(MAIL_SETTINGS_FILE, {})
    return data if isinstance(data, dict) else {}


def saved_or_env(
    settings: dict[str, Any],
    key: str,
    *env_names: str,
    default: str = "",
) -> str:
    value = settings.get(key)
    if value is not None and str(value).strip():
        return str(value).strip()
    for env_name in env_names:
        env_value = os.environ.get(env_name, "").strip()
        if env_value:
            return env_value
    return default


def saved_or_env_int(
    settings: dict[str, Any],
    key: str,
    env_name: str,
    default: int,
    minimum: int,
    maximum: int,
) -> int:
    raw_value = settings.get(key)
    if raw_value is not None and str(raw_value).strip():
        try:
            value = int(raw_value)
        except (TypeError, ValueError):
            value = default
    else:
        value = env_int(env_name, default)
    return max(minimum, min(maximum, value))


def saved_or_env_bool(
    settings: dict[str, Any],
    key: str,
    env_name: str,
    default: bool,
) -> bool:
    raw_value = settings.get(key)
    if raw_value is None or str(raw_value).strip() == "":
        raw_value = os.environ.get(env_name, "1" if default else "0")
    return str(raw_value).strip().lower() not in {"0", "false", "no", "off"}


def load_imap_mail_config() -> ImapMailConfig:
    saved_settings = load_saved_imap_mail_settings()
    return ImapMailConfig(
        host=saved_or_env(saved_settings, "host", "IMAP_HOST", "MAIL_IMAP_HOST"),
        port=saved_or_env_int(saved_settings, "port", "IMAP_PORT", 993, 1, 65535),
        username=saved_or_env(
            saved_settings,
            "username",
            "IMAP_USER",
            "MAIL_IMAP_USER",
            default="info_order@ithe.co.jp",
        ),
        password=saved_or_env(
            saved_settings,
            "password",
            "IMAP_PASSWORD",
            "MAIL_IMAP_PASSWORD",
        ),
        inbox=saved_or_env(saved_settings, "inbox", "IMAP_INBOX", default="INBOX"),
        processed_folder=saved_or_env(
            saved_settings,
            "processed_folder",
            "IMAP_PROCESSED_FOLDER",
            default="Processed",
        ),
        error_folder=saved_or_env(
            saved_settings,
            "error_folder",
            "IMAP_ERROR_FOLDER",
            default="ImportError",
        ),
        max_messages=saved_or_env_int(
            saved_settings,
            "max_messages",
            "IMAP_MAX_MESSAGES",
            20,
            1,
            100,
        ),
        use_ssl=saved_or_env_bool(saved_settings, "use_ssl", "IMAP_USE_SSL", True),
        allowed_senders=env_list("IMAP_ALLOWED_SENDERS", "MAIL_ALLOWED_SENDERS"),
        subject_keywords=merge_keyword_filters(
            DEFAULT_IMAP_SUBJECT_KEYWORDS,
            "IMAP_SUBJECT_KEYWORDS",
            "MAIL_SUBJECT_KEYWORDS",
        ),
        attachment_keywords=env_list(
            "IMAP_ATTACHMENT_KEYWORDS",
            "MAIL_ATTACHMENT_KEYWORDS",
        ),
        exclude_subject_keywords=merge_keyword_filters(
            DEFAULT_IMAP_EXCLUDE_SUBJECT_KEYWORDS,
            "IMAP_EXCLUDE_SUBJECT_KEYWORDS",
            "MAIL_EXCLUDE_SUBJECT_KEYWORDS",
        ),
        exclude_attachment_keywords=env_list(
            "IMAP_EXCLUDE_ATTACHMENT_KEYWORDS",
            "MAIL_EXCLUDE_ATTACHMENT_KEYWORDS",
        ),
    )


def imap_redacted_status(config: ImapMailConfig) -> dict[str, Any]:
    return {
        "configured": config.is_configured,
        "missing": config.missing_keys,
        "host": config.host,
        "port": config.port,
        "user": config.username,
        "password_set": bool(config.password),
        "inbox": config.inbox,
        "processed_folder": config.processed_folder,
        "error_folder": config.error_folder,
        "max_messages": config.max_messages,
        "ssl": config.use_ssl,
        "filters": {
            "allowed_senders": list(config.allowed_senders),
            "subject_keywords": list(config.subject_keywords),
            "attachment_keywords": list(config.attachment_keywords),
            "exclude_subject_keywords": list(config.exclude_subject_keywords),
            "exclude_attachment_keywords": list(config.exclude_attachment_keywords),
        },
    }


def save_imap_mail_settings(payload: dict[str, Any], actor: str = "") -> dict[str, Any]:
    existing = load_saved_imap_mail_settings()
    password = str(payload.get("password") or "").strip()
    settings = {
        "host": str(payload.get("host", existing.get("host", ""))).strip(),
        "port": saved_or_env_int(payload, "port", "IMAP_PORT", 993, 1, 65535),
        "username": str(payload.get("username", existing.get("username", ""))).strip(),
        "password": password or str(existing.get("password", "")).strip(),
        "inbox": str(payload.get("inbox", existing.get("inbox", "INBOX"))).strip() or "INBOX",
        "processed_folder": str(
            payload.get(
                "processed_folder",
                existing.get("processed_folder", "Processed"),
            )
        ).strip()
        or "Processed",
        "error_folder": str(
            payload.get("error_folder", existing.get("error_folder", "ImportError"))
        ).strip()
        or "ImportError",
        "max_messages": saved_or_env_int(
            payload,
            "max_messages",
            "IMAP_MAX_MESSAGES",
            20,
            1,
            100,
        ),
        "use_ssl": bool(payload.get("use_ssl", existing.get("use_ssl", True))),
        "updated_at": datetime.now().isoformat(timespec="seconds"),
        "updated_by": actor,
    }
    save_json_store(MAIL_SETTINGS_FILE, settings)
    append_audit(
        "save_mail_settings",
        actor,
        "imap",
        {
            "host": settings["host"],
            "username": settings["username"],
            "password_changed": bool(password),
        },
    )
    return imap_redacted_status(load_imap_mail_config())


def imap_ok(status: Any) -> bool:
    return str(status).upper() == "OK"


def open_imap_connection(config: ImapMailConfig) -> imaplib.IMAP4:
    if not config.is_configured:
        raise DashboardError(
            "IMAP設定が未設定です。メール取込み画面でメール設定を保存してください。"
        )
    try:
        connection: imaplib.IMAP4
        if config.use_ssl:
            connection = imaplib.IMAP4_SSL(config.host, config.port)
        else:
            connection = imaplib.IMAP4(config.host, config.port)
        connection.login(config.username, config.password)
        return connection
    except imaplib.IMAP4.error as exc:
        raise DashboardError(f"メールサーバーへログインできませんでした: {exc}") from exc
    except OSError as exc:
        raise DashboardError(f"メールサーバーへ接続できませんでした: {exc}") from exc


def close_imap_connection(connection: imaplib.IMAP4) -> None:
    try:
        connection.logout()
    except imaplib.IMAP4.error:
        pass


def select_imap_mailbox(
    connection: imaplib.IMAP4,
    mailbox: str,
    *,
    readonly: bool,
) -> None:
    try:
        status, _ = connection.select(mailbox, readonly=readonly)
    except imaplib.IMAP4.error as exc:
        raise DashboardError(f"メールボックス `{mailbox}` を開けませんでした。") from exc
    if not imap_ok(status):
        raise DashboardError(f"メールボックス `{mailbox}` を開けませんでした。")


def ensure_imap_folder(connection: imaplib.IMAP4, folder: str) -> None:
    if not folder:
        return
    try:
        status, _ = connection.create(folder)
        if imap_ok(status):
            return
    except imaplib.IMAP4.error:
        return


def move_imap_message(connection: imaplib.IMAP4, uid: str, folder: str) -> None:
    if not uid or not folder:
        return
    ensure_imap_folder(connection, folder)
    try:
        status, _ = connection.uid("MOVE", uid, folder)
        if imap_ok(status):
            return
    except imaplib.IMAP4.error:
        pass

    try:
        status, _ = connection.uid("COPY", uid, folder)
        if not imap_ok(status):
            raise DashboardError(f"メールを `{folder}` へ移動できませんでした。")
        connection.uid("STORE", uid, "+FLAGS", r"(\Deleted)")
        connection.expunge()
    except imaplib.IMAP4.error as exc:
        raise DashboardError(f"メールを `{folder}` へ移動できませんでした。") from exc


def imap_received_at(message: Any) -> str:
    raw = str(message.get("date", ""))
    if not raw:
        return ""
    try:
        parsed = parsedate_to_datetime(raw)
    except (TypeError, ValueError, IndexError, OverflowError):
        return ""
    return parsed.isoformat(timespec="seconds") if parsed else ""


def imap_sender(message: Any) -> str:
    raw = str(message.get("from", ""))
    _, address = parseaddr(raw)
    return address or raw


def imap_sender_parts(message: Any) -> tuple[str, str]:
    raw = str(message.get("from", ""))
    name, address = parseaddr(raw)
    return name.strip(), (address or raw).strip()


def imap_message_body(message: Any) -> str:
    try:
        body = message.get_body(preferencelist=("plain", "html"))
        value = body.get_content() if body is not None else ""
        if body is not None and body.get_content_type() == "text/html":
            value = re.sub(r"<br\s*/?>", "\n", str(value), flags=re.IGNORECASE)
            value = re.sub(r"</p\s*>", "\n\n", value, flags=re.IGNORECASE)
            value = re.sub(r"<[^>]+>", " ", value)
            value = html.unescape(value)
        return str(value or "").replace("\r\n", "\n").strip()
    except (AttributeError, KeyError, TypeError, ValueError):
        return ""


def normalized_excel_sheet_name(value: Any) -> str:
    return re.sub(
        r"\s+",
        "",
        unicodedata.normalize("NFKC", str(value or "")),
    ).casefold()


AIZA_SHEET_NAME_PRIORITY = ("アイザ様", "アイザ")
AIZA_SHEET_SIGNATURE = {
    "C9": "発注元名",
    "C14": "お客様カナ名",
    "C19": "品名",
    "B26": "弊社問合番号",
}


def preferred_aiza_sheet_name(sheet_names: Iterable[str]) -> str:
    names = [str(name) for name in sheet_names]
    for expected_name in AIZA_SHEET_NAME_PRIORITY:
        expected = normalized_excel_sheet_name(expected_name)
        for name in names:
            if normalized_excel_sheet_name(name) == expected:
                return name
    return ""


def excel_a1_position(address: str) -> tuple[int, int]:
    match = re.fullmatch(r"([A-Z]+)([1-9]\d*)", address.upper())
    if not match:
        raise ValueError(f"Invalid Excel address: {address}")
    column = 0
    for character in match.group(1):
        column = column * 26 + (ord(character) - ord("A") + 1)
    return int(match.group(2)) - 1, column - 1


def aiza_sheet_signature_diagnostics(cell_value: Any) -> dict[str, Any]:
    missing: list[dict[str, str]] = []
    for address, expected in AIZA_SHEET_SIGNATURE.items():
        actual = str(cell_value(address) or "").strip()
        if normalized_excel_sheet_name(expected) not in normalized_excel_sheet_name(actual):
            missing.append(
                {
                    "address": address,
                    "expected": expected,
                    "actual": actual,
                }
            )
    total = len(AIZA_SHEET_SIGNATURE)
    return {
        "matched_labels": total - len(missing),
        "total_labels": total,
        "missing_labels": missing,
        "valid": not missing,
    }


def excel_workbook_preview_result(content: bytes, file_name: str) -> dict[str, Any]:
    diagnostics: dict[str, Any] = {
        "status": "opening",
        "file_name": Path(file_name).name,
        "file_type": Path(file_name).suffix.lower(),
        "selected_sheet_name": "",
        "available_sheet_names": [],
        "matched_labels": 0,
        "total_labels": len(AIZA_SHEET_SIGNATURE),
        "missing_labels": [],
        "error": "",
    }
    if Path(file_name).suffix.lower() == ".xls":
        try:
            import xlrd
        except ImportError as exc:
            diagnostics.update(
                status="dependency_error",
                error=f"Excel読取りライブラリを読み込めませんでした: {exc}",
            )
            return {"sheets": [], "diagnostics": diagnostics}
        try:
            workbook = xlrd.open_workbook(file_contents=content)
        except Exception as exc:
            diagnostics.update(
                status="open_error",
                error=f"Excelファイルを開けませんでした: {type(exc).__name__}: {exc}",
            )
            return {"sheets": [], "diagnostics": diagnostics}
        diagnostics["available_sheet_names"] = workbook.sheet_names()
        sheet_name = preferred_aiza_sheet_name(workbook.sheet_names())
        if not sheet_name:
            diagnostics.update(
                status="target_sheet_missing",
                error="「アイザ様」または「アイザ」シートが見つかりませんでした。",
            )
            return {"sheets": [], "diagnostics": diagnostics}
        sheet = workbook.sheet_by_name(sheet_name)
        diagnostics["selected_sheet_name"] = sheet_name

        def xls_cell_value(address: str) -> Any:
            row_index, column_index = excel_a1_position(address)
            if row_index >= sheet.nrows or column_index >= sheet.ncols:
                return ""
            return sheet.cell_value(row_index, column_index)

        signature = aiza_sheet_signature_diagnostics(xls_cell_value)
        diagnostics.update(signature)
        if not signature["valid"]:
            diagnostics.update(
                status="schema_mismatch",
                error="対象シートの固定項目を確認できませんでした。",
            )
            return {"sheets": [], "diagnostics": diagnostics}
        rows: list[dict[str, Any]] = []
        for row_index in range(sheet.nrows):
            cells = []
            for column_index in range(sheet.ncols):
                cell = sheet.cell(row_index, column_index)
                if cell.value is None or str(cell.value).strip() == "":
                    continue
                if cell.ctype == xlrd.XL_CELL_DATE:
                    if not isinstance(cell.value, (int, float)) or cell.value <= 0:
                        continue
                    value = xlrd.xldate_as_datetime(cell.value, workbook.datemode).isoformat()
                else:
                    value = str(cell.value).strip()
                cells.append(
                    {
                        "address": f"{xlrd.formula.colname(column_index)}{row_index + 1}",
                        "column": column_index + 1,
                        "value": value,
                    }
                )
            if cells:
                rows.append({"row": row_index + 1, "cells": cells})
        diagnostics["status"] = "ok"
        return {
            "sheets": [
                {
                    "file_name": Path(file_name).name,
                    "sheet_name": sheet.name,
                    "range": f"A1:{xlrd.formula.colname(max(sheet.ncols - 1, 0))}{max(sheet.nrows, 1)}",
                    "rows": rows,
                }
            ],
            "diagnostics": diagnostics,
        }
    try:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        diagnostics.update(
            status="dependency_error",
            error=f"Excel読取りライブラリを読み込めませんでした: {exc}",
        )
        return {"sheets": [], "diagnostics": diagnostics}
    try:
        workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:
        diagnostics.update(
            status="open_error",
            error=f"Excelファイルを開けませんでした: {type(exc).__name__}: {exc}",
        )
        return {"sheets": [], "diagnostics": diagnostics}
    try:
        diagnostics["available_sheet_names"] = list(workbook.sheetnames)
        sheet_name = preferred_aiza_sheet_name(workbook.sheetnames)
        if not sheet_name:
            diagnostics.update(
                status="target_sheet_missing",
                error="「アイザ様」または「アイザ」シートが見つかりませんでした。",
            )
            return {"sheets": [], "diagnostics": diagnostics}
        sheet = workbook[sheet_name]
        diagnostics["selected_sheet_name"] = sheet_name
        signature = aiza_sheet_signature_diagnostics(
            lambda address: sheet[address].value
        )
        diagnostics.update(signature)
        if not signature["valid"]:
            diagnostics.update(
                status="schema_mismatch",
                error="対象シートの固定項目を確認できませんでした。",
            )
            return {"sheets": [], "diagnostics": diagnostics}
        rows: list[dict[str, Any]] = []
        for row_number, values in enumerate(
            sheet.iter_rows(
                min_row=1,
                max_row=sheet.max_row,
                max_col=sheet.max_column,
                values_only=True,
            ),
            start=1,
        ):
            cells = []
            for column_number, value in enumerate(values, start=1):
                if value is None or str(value).strip() == "":
                    continue
                if isinstance(value, (date, datetime)):
                    if value.year <= 1900:
                        continue
                    display = value.isoformat()
                else:
                    display = str(value).strip()
                cells.append(
                    {
                        "address": f"{get_column_letter(column_number)}{row_number}",
                        "column": column_number,
                        "value": display,
                    }
                )
            if cells:
                rows.append({"row": row_number, "cells": cells})
        diagnostics["status"] = "ok"
        return {
            "sheets": [
                {
                    "file_name": Path(file_name).name,
                    "sheet_name": sheet.title,
                    "range": sheet.calculate_dimension(),
                    "rows": rows,
                }
            ],
            "diagnostics": diagnostics,
        }
    finally:
        workbook.close()


def excel_workbook_preview(content: bytes, file_name: str) -> list[dict[str, Any]]:
    return excel_workbook_preview_result(content, file_name)["sheets"]


def imap_source_message_id(config: ImapMailConfig, uid: str) -> str:
    seed = f"{config.username}:{config.inbox}:{uid}"
    digest = hashlib.sha256(seed.encode("utf-8")).hexdigest()[:16]
    return f"imap:{digest}:{uid}"


def parse_imap_message(
    config: ImapMailConfig,
    uid: str,
    raw_message: bytes,
    *,
    include_content: bool,
) -> dict[str, Any]:
    parsed = BytesParser(policy=email_policy.default).parsebytes(raw_message)
    sender_name, sender_address = imap_sender_parts(parsed)
    attachments: list[dict[str, Any]] = []
    excel_sheets: list[dict[str, Any]] = []
    for index, part in enumerate(parsed.iter_attachments(), start=1):
        original_name = part.get_filename() or f"attachment-{index}.xlsx"
        if not mail_excel_suffix(original_name):
            continue
        content = part.get_payload(decode=True)
        if content is None:
            payload = part.get_content()
            if isinstance(payload, bytes):
                content = payload
            else:
                content = str(payload or "").encode("utf-8")
        attachment: dict[str, Any] = {
            "id": hashlib.sha256(
                f"{uid}:{index}:{original_name}".encode("utf-8")
            ).hexdigest()[:16],
            "name": Path(str(original_name)).name,
            "size": len(content),
            "content_type": part.get_content_type(),
        }
        if include_content:
            attachment["content"] = content
            preview_result = excel_workbook_preview_result(content, str(original_name))
            attachment_sheets = preview_result["sheets"]
            attachment["excel_sheets"] = attachment_sheets
            attachment["excel_diagnostics"] = preview_result["diagnostics"]
            if preview_result["diagnostics"]["status"] != "ok":
                attachment["excel_error"] = preview_result["diagnostics"]["error"]
            excel_sheets.extend(attachment_sheets)
        attachments.append(attachment)

    return {
        "id": imap_source_message_id(config, uid),
        "uid": uid,
        "source": "imap",
        "mailbox": config.inbox,
        "subject": str(parsed.get("subject", "")),
        "sender": sender_address,
        "sender_name": sender_name,
        "sender_address": sender_address,
        "to": str(parsed.get("to", "")),
        "cc": str(parsed.get("cc", "")),
        "body": imap_message_body(parsed),
        "received_at": imap_received_at(parsed),
        "attachments": attachments,
        "excel_sheets": excel_sheets,
    }


def sanitize_mail_message(message: dict[str, Any]) -> dict[str, Any]:
    sanitized = dict(message)
    sanitized["attachments"] = [
        {key: value for key, value in attachment.items() if key != "content"}
        for attachment in message.get("attachments", [])
        if isinstance(attachment, dict)
    ]
    return sanitized


def imap_fetched_message_parts(fetched: Any) -> tuple[bytes, bool]:
    raw_message = b""
    metadata_parts: list[bytes] = []
    for item in fetched or []:
        if isinstance(item, tuple):
            if item and isinstance(item[0], bytes):
                metadata_parts.append(item[0])
            if len(item) > 1 and isinstance(item[1], bytes):
                raw_message = item[1]
        elif isinstance(item, bytes):
            metadata_parts.append(item)
    metadata = b" ".join(metadata_parts).upper()
    return raw_message, b"\\SEEN" not in metadata


def fetch_imap_message_with_flags(
    connection: imaplib.IMAP4,
    uid: str,
) -> tuple[bytes, bool]:
    try:
        status, fetched = connection.uid("FETCH", uid, "(BODY.PEEK[] FLAGS)")
    except imaplib.IMAP4.error:
        return b"", False
    if not imap_ok(status):
        return b"", False
    return imap_fetched_message_parts(fetched)


def mail_attachment_signature(attachment: dict[str, Any]) -> str:
    name = str(attachment.get("name", "")).strip().casefold()
    size = str(attachment.get("size", "")).strip()
    return f"{name}\0{size}" if name and size else ""


def imported_attachment_hashes(imports: list[dict[str, Any]]) -> set[str]:
    hashes: set[str] = set()
    for item in imports:
        if not isinstance(item, dict):
            continue
        for attachment in item.get("attachments", []):
            if not isinstance(attachment, dict):
                continue
            digest = str(attachment.get("sha256", "")).strip()
            if digest:
                hashes.add(digest)
    return hashes


def imported_attachment_signatures(imports: list[dict[str, Any]]) -> set[str]:
    signatures: set[str] = set()
    for item in imports:
        if not isinstance(item, dict):
            continue
        for attachment in item.get("attachments", []):
            if not isinstance(attachment, dict):
                continue
            signature = mail_attachment_signature(attachment)
            if signature:
                signatures.add(signature)
    return signatures


def text_contains_any(value: str, keywords: tuple[str, ...]) -> bool:
    lower = value.lower()
    return any(keyword.lower() in lower for keyword in keywords)


def sender_matches_filters(sender: str, filters: tuple[str, ...]) -> bool:
    lower = sender.lower()
    for raw_filter in filters:
        item = raw_filter.strip().lower()
        if not item:
            continue
        if item.startswith("@") and lower.endswith(item):
            return True
        if "@" in item and lower == item:
            return True
        if "@" not in item and (lower == item or lower.endswith(f"@{item}")):
            return True
    return False


def imap_message_matches_filters(
    config: ImapMailConfig,
    message: dict[str, Any],
) -> bool:
    subject = str(message.get("subject", ""))
    sender = str(message.get("sender", ""))
    attachment_names = " ".join(
        str(attachment.get("name", ""))
        for attachment in message.get("attachments", [])
        if isinstance(attachment, dict)
    )

    if config.allowed_senders and not sender_matches_filters(
        sender,
        config.allowed_senders,
    ):
        return False
    if config.subject_keywords and not text_contains_any(
        subject,
        config.subject_keywords,
    ):
        return False
    if config.attachment_keywords and not text_contains_any(
        attachment_names,
        config.attachment_keywords,
    ):
        return False
    if config.exclude_subject_keywords and text_contains_any(
        subject,
        config.exclude_subject_keywords,
    ):
        return False
    if config.exclude_attachment_keywords and text_contains_any(
        attachment_names,
        config.exclude_attachment_keywords,
    ):
        return False
    return True


def fetch_imap_candidate_messages(
    connection: imaplib.IMAP4,
    config: ImapMailConfig,
    *,
    include_content: bool,
    readonly: bool,
    include_all: bool = False,
) -> list[dict[str, Any]]:
    select_imap_mailbox(connection, config.inbox, readonly=readonly)
    try:
        status, data = connection.uid("SEARCH", None, "ALL")
    except imaplib.IMAP4.error as exc:
        raise DashboardError("メール一覧を取得できませんでした。") from exc
    if not imap_ok(status) or not data:
        return []

    uids = data[0].split() if isinstance(data[0], bytes) else str(data[0]).split()
    selected_uids = list(reversed(uids[-config.max_messages * 3 :]))
    imports = load_mail_imports()
    existing_keys = {
        str(item.get("message_id", ""))
        for item in imports
        if isinstance(item, dict)
    }
    existing_attachment_signatures = imported_attachment_signatures(imports)
    candidates: list[dict[str, Any]] = []
    for raw_uid in selected_uids:
        uid = raw_uid.decode("ascii", errors="ignore") if isinstance(raw_uid, bytes) else str(raw_uid)
        message_id = imap_source_message_id(config, uid)
        if not uid or (not include_all and message_id in existing_keys):
            continue
        raw_message, is_unread = fetch_imap_message_with_flags(connection, uid)
        if not raw_message:
            continue
        message = parse_imap_message(
            config,
            uid,
            raw_message,
            include_content=include_content,
        )
        message["is_unread"] = is_unread
        attachments = [
            attachment
            for attachment in message.get("attachments", [])
            if isinstance(attachment, dict)
        ]
        if not include_all and attachments and all(
            mail_attachment_signature(attachment) in existing_attachment_signatures
            for attachment in attachments
        ):
            continue
        if include_all:
            candidates.append(message if include_content else sanitize_mail_message(message))
        elif attachments and imap_message_matches_filters(config, message):
            candidates.append(message if include_content else sanitize_mail_message(message))
        if len(candidates) >= config.max_messages:
            break
    return candidates


def imap_candidate_messages(config: ImapMailConfig) -> list[dict[str, Any]]:
    connection = open_imap_connection(config)
    try:
        return fetch_imap_candidate_messages(
            connection,
            config,
            include_content=False,
            readonly=True,
        )
    finally:
        close_imap_connection(connection)


def imap_mailbox_messages(
    config: ImapMailConfig,
    *,
    page: int = 1,
    page_size: int = 20,
) -> dict[str, Any]:
    connection = open_imap_connection(config)
    try:
        select_imap_mailbox(connection, config.inbox, readonly=True)
        try:
            status, data = connection.uid("SEARCH", None, "ALL")
            unread_status, unread_data = connection.uid("SEARCH", None, "UNSEEN")
        except imaplib.IMAP4.error as exc:
            raise DashboardError("メール一覧を取得できませんでした。") from exc
        if not imap_ok(status) or not data:
            uids: list[Any] = []
        else:
            uids = data[0].split() if isinstance(data[0], bytes) else str(data[0]).split()
        unread_uids = (
            unread_data[0].split()
            if imap_ok(unread_status) and unread_data and isinstance(unread_data[0], bytes)
            else []
        )
        total_count = len(uids)
        safe_page_size = max(1, min(int(page_size), 100))
        total_pages = max(1, (total_count + safe_page_size - 1) // safe_page_size)
        safe_page = max(1, min(int(page), total_pages))
        end = total_count - ((safe_page - 1) * safe_page_size)
        start = max(0, end - safe_page_size)
        selected_uids = list(reversed(uids[start:end]))
        messages: list[dict[str, Any]] = []
        for raw_uid in selected_uids:
            uid = raw_uid.decode("ascii", errors="ignore") if isinstance(raw_uid, bytes) else str(raw_uid)
            raw_message, is_unread = fetch_imap_message_with_flags(connection, uid)
            if not raw_message:
                continue
            message = parse_imap_message(
                config,
                uid,
                raw_message,
                include_content=False,
            )
            message["is_unread"] = is_unread
            messages.append(sanitize_mail_message(message))
        return {
            "messages": messages,
            "page": safe_page,
            "page_size": safe_page_size,
            "total_count": total_count,
            "total_pages": total_pages,
            "unread_count": len(unread_uids),
        }
    finally:
        close_imap_connection(connection)


def imap_mailbox_message_detail(config: ImapMailConfig, uid: str) -> dict[str, Any]:
    if not uid.isdigit():
        raise DashboardError("メールIDが正しくありません。")
    connection = open_imap_connection(config)
    try:
        select_imap_mailbox(connection, config.inbox, readonly=True)
        raw_message, is_unread = fetch_imap_message_with_flags(connection, uid)
        if not raw_message:
            raise DashboardError("メール本文を取得できませんでした。")
        message = parse_imap_message(
            config,
            uid,
            raw_message,
            include_content=True,
        )
        message["is_unread"] = is_unread
        return sanitize_mail_message(message)
    except imaplib.IMAP4.error as exc:
        raise DashboardError("メール本文を取得できませんでした。") from exc
    finally:
        close_imap_connection(connection)


def set_imap_message_read_state(
    config: ImapMailConfig,
    uid: str,
    *,
    unread: bool,
) -> dict[str, Any]:
    if not uid.isdigit():
        raise DashboardError("メールIDが正しくありません。")
    connection = open_imap_connection(config)
    try:
        select_imap_mailbox(connection, config.inbox, readonly=False)
        operation = "-FLAGS.SILENT" if unread else "+FLAGS.SILENT"
        status, _ = connection.uid("STORE", uid, operation, r"(\Seen)")
        if not imap_ok(status):
            raise DashboardError("メールの既読状態を変更できませんでした。")
        return {"uid": uid, "is_unread": unread}
    except imaplib.IMAP4.error as exc:
        raise DashboardError("メールの既読状態を変更できませんでした。") from exc
    finally:
        close_imap_connection(connection)


def save_imap_attachment(
    message: dict[str, Any],
    attachment: dict[str, Any],
) -> dict[str, Any]:
    content = attachment.get("content")
    if not isinstance(content, bytes) or not content:
        raise DashboardError("添付ファイルの内容を取得できませんでした。")
    received = safe_filename_part(str(message.get("received_at", ""))[:10], "mail")
    message_hash = hashlib.sha256(str(message.get("id", "")).encode("utf-8")).hexdigest()[:12]
    original_name = Path(str(attachment.get("name", "attachment.xlsx"))).name
    safe_name = safe_submission_filename(original_name)
    target_dir = MAIL_ATTACHMENTS_DIR / received / message_hash
    target_dir.mkdir(parents=True, exist_ok=True)
    target_path = target_dir / safe_name
    target_path.write_bytes(content)
    sync_path_to_cloud(target_path)
    return {
        "id": attachment.get("id", ""),
        "name": original_name,
        "stored_name": safe_name,
        "relative_path": str(target_path.relative_to(MAIL_ATTACHMENTS_DIR)).replace("\\", "/"),
        "size": len(content),
        "content_type": attachment.get("content_type", ""),
        "sha256": hashlib.sha256(content).hexdigest(),
    }


def import_imap_messages(
    actor: str = "",
    selected_message_ids: set[str] | None = None,
) -> dict[str, Any]:
    if not IMAP_IMPORT_LOCK.acquire(blocking=False):
        raise DashboardError("メール取込が実行中です。少し待ってから再度実行してください。")
    try:
        return import_imap_messages_locked(
            actor=actor,
            selected_message_ids=selected_message_ids,
        )
    finally:
        IMAP_IMPORT_LOCK.release()


def import_imap_messages_locked(
    actor: str = "",
    selected_message_ids: set[str] | None = None,
) -> dict[str, Any]:
    if selected_message_ids is not None and not selected_message_ids:
        raise DashboardError("取り込むメールを選択してください。")

    config = load_imap_mail_config()
    history = load_mail_imports()
    imported: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []
    duplicates: list[dict[str, Any]] = []
    existing_keys = {
        str(item.get("message_id", ""))
        for item in history
        if isinstance(item, dict)
    }
    existing_hashes = imported_attachment_hashes(history)

    connection = open_imap_connection(config)
    try:
        candidates = fetch_imap_candidate_messages(
            connection,
            config,
            include_content=True,
            readonly=False,
        )
        for message in candidates:
            message_id = str(message.get("id", ""))
            uid = str(message.get("uid", ""))
            if selected_message_ids is not None and message_id not in selected_message_ids:
                continue
            if not message_id or message_id in existing_keys:
                continue
            try:
                saved_attachments: list[dict[str, Any]] = []
                duplicate_attachments: list[dict[str, Any]] = []
                for attachment in message.get("attachments", []):
                    if not isinstance(attachment, dict):
                        continue
                    content = attachment.get("content")
                    digest = (
                        hashlib.sha256(content).hexdigest()
                        if isinstance(content, bytes) and content
                        else ""
                    )
                    if digest and digest in existing_hashes:
                        duplicate_attachments.append(
                            {
                                "name": attachment.get("name", ""),
                                "size": attachment.get("size", 0),
                                "sha256": digest,
                            }
                        )
                        continue
                    saved = save_imap_attachment(message, attachment)
                    existing_hashes.add(str(saved.get("sha256", "")))
                    saved_attachments.append(saved)
                if not saved_attachments:
                    duplicate_entry = {
                        "id": secrets.token_hex(12),
                        "source": "imap",
                        "message_id": message_id,
                        "uid": uid,
                        "mailbox": message.get("mailbox", config.inbox),
                        "subject": message.get("subject", ""),
                        "sender": message.get("sender", ""),
                        "received_at": message.get("received_at", ""),
                        "attachments": duplicate_attachments,
                        "status": "duplicate",
                        "created_jobs": 0,
                        "created_at": datetime.now().isoformat(timespec="seconds"),
                        "imported_by": actor,
                    }
                    history.append(duplicate_entry)
                    existing_keys.add(message_id)
                    duplicates.append(duplicate_entry)
                    move_imap_message(connection, uid, config.processed_folder)
                    continue
                entry = {
                    "id": secrets.token_hex(12),
                    "source": "imap",
                    "message_id": message_id,
                    "uid": uid,
                    "mailbox": message.get("mailbox", config.inbox),
                    "subject": message.get("subject", ""),
                    "sender": message.get("sender", ""),
                    "received_at": message.get("received_at", ""),
                    "attachments": saved_attachments,
                    "status": "saved",
                    "created_jobs": 0,
                    "created_at": datetime.now().isoformat(timespec="seconds"),
                    "imported_by": actor,
                }
                materialize_mail_import_entry_jobs(entry, actor=actor)
                history.append(entry)
                existing_keys.add(message_id)
                imported.append(entry)
                move_imap_message(connection, uid, config.processed_folder)
            except DashboardError as exc:
                errors.append({"message_id": message_id, "error": str(exc)})
                try:
                    move_imap_message(connection, uid, config.error_folder)
                except DashboardError:
                    pass
    finally:
        close_imap_connection(connection)

    save_mail_imports(history)
    return {
        "imported": imported,
        "errors": errors,
        "summary": {
            "candidate_count": len(candidates),
            "imported_count": len(imported),
            "duplicate_count": len(duplicates),
            "created_jobs": sum(int(item.get("created_jobs", 0)) for item in imported),
            "error_count": len(errors),
        },
    }


def load_mail_imports() -> list[dict[str, Any]]:
    data = load_json_store(MAIL_IMPORTS_FILE, [])
    return data if isinstance(data, list) else []


def save_mail_imports(imports: list[dict[str, Any]]) -> None:
    save_json_store(MAIL_IMPORTS_FILE, imports)


def auto_import_interval_seconds() -> int:
    raw = env_int("IMAP_AUTO_IMPORT_INTERVAL_SECONDS", 0)
    if raw <= 0:
        raw = env_int("MAIL_AUTO_IMPORT_INTERVAL_SECONDS", 0)
    if raw <= 0:
        return 0
    return max(60, min(24 * 60 * 60, raw))


def auto_import_status_payload() -> dict[str, Any]:
    interval = auto_import_interval_seconds()
    with AUTO_IMPORT_STATUS_LOCK:
        payload = dict(AUTO_IMPORT_STATUS)
    payload["enabled"] = interval > 0
    payload["interval_seconds"] = interval
    return payload


def set_auto_import_status(**updates: Any) -> None:
    with AUTO_IMPORT_STATUS_LOCK:
        AUTO_IMPORT_STATUS.update(updates)


def run_auto_imap_import_once() -> dict[str, Any]:
    started_at = datetime.now()
    set_auto_import_status(
        enabled=True,
        running=True,
        last_started_at=started_at.isoformat(timespec="seconds"),
        last_error="",
    )
    try:
        result = import_imap_messages(actor="auto-import")
        summary = result.get("summary", {})
        append_audit("auto_import_imap_mail", "auto-import", "info_order", summary)
        set_auto_import_status(
            running=False,
            last_finished_at=datetime.now().isoformat(timespec="seconds"),
            last_summary=summary,
            last_error="",
        )
        return result
    except DashboardError as exc:
        set_auto_import_status(
            running=False,
            last_finished_at=datetime.now().isoformat(timespec="seconds"),
            last_error=str(exc),
        )
        return {"errors": [{"error": str(exc)}], "summary": {"error_count": 1}}


def auto_import_loop(interval: int) -> None:
    set_auto_import_status(
        enabled=True,
        interval_seconds=interval,
        next_run_at=(datetime.now() + timedelta(seconds=interval)).isoformat(timespec="seconds"),
    )
    while not AUTO_IMPORT_STOP.wait(interval):
        run_auto_imap_import_once()
        set_auto_import_status(
            next_run_at=(datetime.now() + timedelta(seconds=interval)).isoformat(timespec="seconds")
        )


def start_auto_import_worker() -> None:
    global AUTO_IMPORT_THREAD
    interval = auto_import_interval_seconds()
    if interval <= 0:
        set_auto_import_status(enabled=False, interval_seconds=0, next_run_at="")
        return
    if AUTO_IMPORT_THREAD and AUTO_IMPORT_THREAD.is_alive():
        return
    AUTO_IMPORT_STOP.clear()
    AUTO_IMPORT_THREAD = threading.Thread(
        target=auto_import_loop,
        args=(interval,),
        name="imap-auto-import",
        daemon=True,
    )
    AUTO_IMPORT_THREAD.start()


def stop_auto_import_worker() -> None:
    AUTO_IMPORT_STOP.set()


def mail_imports_payload() -> dict[str, Any]:
    imports = sorted(
        load_mail_imports(),
        key=lambda item: str(item.get("created_at", "")),
        reverse=True,
    )
    jobs_by_id = {
        str(job.get("id", "")): job
        for job in load_logistics_jobs()
        if str(job.get("id", ""))
    }
    visible_imports = []
    for item in imports[:100]:
        visible = dict(item)
        visible["jobs"] = [
            {
                "id": job["id"],
                "work_order_number": job.get("work_order_number", ""),
                "customer_name": job.get("customer_name", ""),
                "scheduled_date": job.get("scheduled_date", ""),
                "sagyou_sync_status": job.get("sagyou_sync_status", ""),
                "sagyou_job_id": job.get("sagyou_job_id", ""),
                "sagyou_synced_at": job.get("sagyou_synced_at", ""),
                "sagyou_last_error": job.get("sagyou_last_error", ""),
                "sagyou_last_attempt_at": job.get("sagyou_last_attempt_at", ""),
                "sagyou_last_http_status": job.get("sagyou_last_http_status", 0),
                "sagyou_last_response": job.get("sagyou_last_response", ""),
            }
            for job_id in item.get("job_ids", [])
            if (job := jobs_by_id.get(str(job_id))) is not None
        ]
        visible_imports.append(visible)
    return {
        "imports": visible_imports,
        "summary": {
            "count": len(imports),
            "saved_attachments": sum(
                len(item.get("attachments", []))
                for item in imports
                if isinstance(item, dict)
            ),
            "created_jobs": sum(
                int(item.get("created_jobs", 0) or 0)
                for item in imports
                if isinstance(item, dict)
            ),
        },
        "auto_import": auto_import_status_payload(),
    }


def outlook_message_list_path(config: OutlookMailConfig) -> str:
    params = urlencode(
        {
            "$top": str(config.max_messages),
            "$select": "id,subject,from,receivedDateTime,hasAttachments,categories,internetMessageId",
            "$orderby": "receivedDateTime desc",
        }
    )
    return f"{outlook_base_path(config)}/mailFolders/inbox/messages?{params}"


def outlook_message_attachments_path(config: OutlookMailConfig, message_id: str) -> str:
    params = urlencode(
        {
            "$select": "id,name,contentType,size,isInline",
        }
    )
    return (
        f"{outlook_base_path(config)}/messages/{quote(message_id, safe='')}"
        f"/attachments?{params}"
    )


def outlook_attachment_path(
    config: OutlookMailConfig,
    message_id: str,
    attachment_id: str,
) -> str:
    return (
        f"{outlook_base_path(config)}/messages/{quote(message_id, safe='')}"
        f"/attachments/{quote(attachment_id, safe='')}"
    )


def outlook_patch_message_path(config: OutlookMailConfig, message_id: str) -> str:
    return f"{outlook_base_path(config)}/messages/{quote(message_id, safe='')}"


def message_sender(message: dict[str, Any]) -> str:
    return str(
        message.get("from", {})
        .get("emailAddress", {})
        .get("address", "")
    )


def message_categories(message: dict[str, Any]) -> list[str]:
    raw = message.get("categories", [])
    return [str(item) for item in raw if str(item).strip()] if isinstance(raw, list) else []


def message_is_unprocessed(config: OutlookMailConfig, message: dict[str, Any]) -> bool:
    categories = set(message_categories(message))
    return (
        bool(message.get("hasAttachments"))
        and config.processed_category not in categories
        and config.error_category not in categories
    )


def outlook_candidate_messages(config: OutlookMailConfig) -> list[dict[str, Any]]:
    payload = outlook_graph_json(config, outlook_message_list_path(config))
    messages = payload.get("value", [])
    if not isinstance(messages, list):
        return []
    candidates: list[dict[str, Any]] = []
    for message in messages:
        if not isinstance(message, dict) or not message_is_unprocessed(config, message):
            continue
        attachments = outlook_graph_json(
            config,
            outlook_message_attachments_path(config, str(message.get("id", ""))),
        ).get("value", [])
        excel_attachments = [
            attachment
            for attachment in attachments
            if isinstance(attachment, dict)
            and not attachment.get("isInline")
            and outlook_excel_suffix(str(attachment.get("name", "")))
        ]
        if not excel_attachments:
            continue
        candidates.append(
            {
                "id": message.get("id", ""),
                "subject": message.get("subject", ""),
                "sender": message_sender(message),
                "received_at": message.get("receivedDateTime", ""),
                "categories": message_categories(message),
                "attachments": [
                    {
                        "id": attachment.get("id", ""),
                        "name": attachment.get("name", ""),
                        "size": attachment.get("size", 0),
                        "content_type": attachment.get("contentType", ""),
                    }
                    for attachment in excel_attachments
                ],
            }
        )
    return candidates


def append_outlook_category(
    config: OutlookMailConfig,
    message_id: str,
    category: str,
    existing_categories: list[str],
) -> None:
    categories = [item for item in existing_categories if item and item != config.error_category]
    if category not in categories:
        categories.append(category)
    outlook_graph_json(
        config,
        outlook_patch_message_path(config, message_id),
        method="PATCH",
        payload={"categories": categories},
    )


def save_outlook_attachment(
    config: OutlookMailConfig,
    message: dict[str, Any],
    attachment: dict[str, Any],
) -> dict[str, Any]:
    detail = outlook_graph_json(
        config,
        outlook_attachment_path(
            config,
            str(message.get("id", "")),
            str(attachment.get("id", "")),
        ),
    )
    content = str(detail.get("contentBytes", ""))
    if not content:
        raise DashboardError("添付ファイルの内容を取得できませんでした。")
    try:
        body = base64.b64decode(content)
    except ValueError as exc:
        raise DashboardError("添付ファイルを読み取れませんでした。") from exc

    received = safe_filename_part(str(message.get("received_at", ""))[:10], "mail")
    message_hash = hashlib.sha256(str(message.get("id", "")).encode("utf-8")).hexdigest()[:12]
    original_name = Path(str(attachment.get("name", "attachment.xlsx"))).name
    safe_name = safe_submission_filename(original_name)
    target_dir = MAIL_ATTACHMENTS_DIR / received / message_hash
    target_dir.mkdir(parents=True, exist_ok=True)
    target_path = target_dir / safe_name
    target_path.write_bytes(body)
    sync_path_to_cloud(target_path)
    return {
        "name": original_name,
        "stored_name": safe_name,
        "relative_path": str(target_path.relative_to(MAIL_ATTACHMENTS_DIR)).replace("\\", "/"),
        "size": len(body),
        "content_type": detail.get("contentType", attachment.get("content_type", "")),
    }


def import_outlook_messages(actor: str = "") -> dict[str, Any]:
    config = load_outlook_mail_config()
    candidates = outlook_candidate_messages(config)
    history = load_mail_imports()
    imported: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []
    existing_keys = {
        str(item.get("message_id", ""))
        for item in history
        if isinstance(item, dict)
    }

    for message in candidates:
        message_id = str(message.get("id", ""))
        if not message_id or message_id in existing_keys:
            continue
        try:
            saved_attachments = [
                save_outlook_attachment(config, message, attachment)
                for attachment in message.get("attachments", [])
                if isinstance(attachment, dict)
            ]
            if not saved_attachments:
                raise DashboardError("Excel添付がありませんでした。")
            entry = {
                "id": secrets.token_hex(12),
                "message_id": message_id,
                "subject": message.get("subject", ""),
                "sender": message.get("sender", ""),
                "received_at": message.get("received_at", ""),
                "attachments": saved_attachments,
                "status": "saved",
                "created_jobs": 0,
                "created_at": datetime.now().isoformat(timespec="seconds"),
                "imported_by": actor,
            }
            history.append(entry)
            imported.append(entry)
            append_outlook_category(
                config,
                message_id,
                config.processed_category,
                message_categories(message),
            )
        except DashboardError as exc:
            errors.append({"message_id": message_id, "error": str(exc)})
            try:
                append_outlook_category(
                    config,
                    message_id,
                    config.error_category,
                    message_categories(message),
                )
            except DashboardError:
                pass
    save_mail_imports(history)
    return {
        "imported": imported,
        "errors": errors,
        "summary": {
            "candidate_count": len(candidates),
            "imported_count": len(imported),
            "error_count": len(errors),
        },
    }


JOB_STATUSES = {
    "unprocessed": "新規依頼",
    "scheduled": "作業報告待ち",
    "completed": "作業完了",
    "reported": "請求確定待ち",
    "billed": "請求確定",
    "needs_review": "要確認",
}

JOB_MONEY_FIELDS = {
    "installation_fee_yen",
    "distance_fee_yen",
    "toll_fee_yen",
    "parking_fee_yen",
    "other_fee_yen",
    "subcontract_fee_yen",
    "purchase_amount_yen",
}


SUBCONTRACTOR_FIELDS = (
    "contractor_code",
    "customer_number",
    "company_name",
    "postal_code",
    "address",
    "contact_name",
    "contact_phone",
    "warehouse_address",
    "memo",
)


def normalize_contractor_code(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "").strip())


def public_subcontractor(item: dict[str, Any]) -> dict[str, Any]:
    return {
        "id": str(item.get("id", "")),
        "contractor_code": str(item.get("contractor_code", "")),
        "customer_number": str(item.get("customer_number", "")),
        "company_name": str(item.get("company_name", "")),
        "postal_code": str(item.get("postal_code", "")),
        "address": str(item.get("address", "")),
        "contact_name": str(item.get("contact_name", "")),
        "contact_phone": str(item.get("contact_phone", "")),
        "warehouse_address": str(item.get("warehouse_address", "")),
        "memo": str(item.get("memo", "")),
        "login_user_id": str(item.get("login_user_id", "")),
        "is_active": bool(item.get("is_active", True)),
        "created_at": str(item.get("created_at", "")),
        "updated_at": str(item.get("updated_at", "")),
    }


def load_subcontractors() -> list[dict[str, Any]]:
    data = load_json_store(SUBCONTRACTORS_FILE, [])
    if isinstance(data, dict):
        data = data.get("subcontractors", [])
    if not isinstance(data, list):
        return []
    items = [
        public_subcontractor(item)
        for item in data
        if isinstance(item, dict) and str(item.get("contractor_code", "")).strip()
    ]
    items.sort(key=lambda item: (item.get("contractor_code", ""), item.get("company_name", "")))
    return items


def save_subcontractors(items: list[dict[str, Any]]) -> None:
    save_json_store(SUBCONTRACTORS_FILE, items)


def subcontractor_by_code(code: str) -> dict[str, Any]:
    normalized = normalize_contractor_code(code)
    for item in load_subcontractors():
        if normalize_contractor_code(item.get("contractor_code")) == normalized:
            return item
    return {}


def ensure_contractor_login_account(
    contractor: dict[str, Any],
    password: str,
    actor: str,
) -> dict[str, Any] | None:
    password = str(password or "")
    if not password:
        return None
    if len(password) < 8:
        raise DashboardError("業者ログイン用パスワードは8文字以上にしてください。")

    user_id = normalize_contractor_code(contractor.get("contractor_code"))
    if not user_id:
        raise DashboardError("業者コードを入力してください。")

    users = ensure_user_store()
    normalized = normalize_user_id(user_id)
    now = datetime.now().isoformat(timespec="seconds")
    users[normalized] = {
        **users.get(normalized, {}),
        "id": user_id,
        "role": "contractor",
        "contractor_code": user_id,
        "company_name": str(contractor.get("company_name", "")),
        "password_hash": hash_password(password),
        "created_at": str(users.get(normalized, {}).get("created_at") or now),
        "updated_at": now,
    }
    save_users(users)
    append_audit(
        "upsert_contractor_account",
        actor,
        user_id,
        {"company_name": contractor.get("company_name", "")},
    )
    return public_user(users[normalized])


def save_subcontractor(payload: dict[str, Any], actor: str = "") -> dict[str, Any]:
    now = datetime.now().isoformat(timespec="seconds")
    contractor_code = normalize_contractor_code(
        payload.get("contractor_code") or payload.get("customer_number")
    )
    if not contractor_code:
        raise DashboardError("顧客番号（業者コード）を入力してください。")
    company_name = str(payload.get("company_name", "")).strip()
    if not company_name:
        raise DashboardError("会社名を入力してください。")

    contractors = load_subcontractors()
    contractor_id = str(payload.get("id", "")).strip()
    existing_index = next(
        (
            index
            for index, item in enumerate(contractors)
            if str(item.get("id", "")) == contractor_id
            or normalize_contractor_code(item.get("contractor_code")) == contractor_code
        ),
        None,
    )
    existing = contractors[existing_index] if existing_index is not None else {}
    contractor = {
        "id": str(existing.get("id") or contractor_id or secrets.token_hex(12)),
        "contractor_code": contractor_code,
        "customer_number": contractor_code,
        "company_name": company_name,
        "postal_code": str(payload.get("postal_code", existing.get("postal_code", ""))).strip(),
        "address": str(payload.get("address", existing.get("address", ""))).strip(),
        "contact_name": str(payload.get("contact_name", existing.get("contact_name", ""))).strip(),
        "contact_phone": str(payload.get("contact_phone", existing.get("contact_phone", ""))).strip(),
        "warehouse_address": str(
            payload.get("warehouse_address", existing.get("warehouse_address", ""))
        ).strip(),
        "memo": str(payload.get("memo", existing.get("memo", ""))).strip(),
        "login_user_id": contractor_code,
        "is_active": bool(payload.get("is_active", existing.get("is_active", True))),
        "created_at": str(existing.get("created_at") or now),
        "updated_at": now,
    }

    ensure_contractor_login_account(
        contractor,
        str(payload.get("login_password", "")),
        actor,
    )

    if existing_index is None:
        contractors.append(contractor)
        action = "create_subcontractor"
    else:
        contractors[existing_index] = contractor
        action = "update_subcontractor"
    save_subcontractors(contractors)
    append_audit(action, actor, contractor_code, {"company_name": company_name})
    return contractor


def subcontractors_payload() -> dict[str, Any]:
    return {
        "subcontractors": load_subcontractors(),
        "updated_at": datetime.now().isoformat(timespec="seconds"),
    }


def coerce_int(value: Any) -> int:
    if isinstance(value, bool):
        return 0
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    cleaned = re.sub(r"[^0-9.-]", "", str(value or ""))
    if not cleaned or cleaned in {"-", ".", "-."}:
        return 0
    try:
        return int(float(cleaned))
    except ValueError:
        return 0


RATE_WORK_ITEM_KEYWORDS = (
    "階段上げ",
    "階段揚げ",
    "階段下ろし",
    "梱包作業費",
    "室外機",
    "マッサージチェア階段",
    "4人作業費",
    "４人作業費",
)


def normalize_rate_key(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("　", " ")
    text = re.sub(r"\s+", "", text)
    text = text.replace("kg", "ｋｇ")
    return text


def normalize_model_key(value: Any) -> str:
    text = str(value or "").strip().upper()
    text = text.replace("　", " ")
    text = re.sub(r"\s+", "", text)
    return text


def model_without_color(value: Any) -> str:
    return re.sub(r"\([^)]*\)$", "", normalize_model_key(value))


def model_family_key(value: Any) -> str:
    model = model_without_color(value)
    match = re.match(r"^([A-Z]+(?:-[A-Z]+)*-?[A-Z]*\d+)", model)
    return match.group(1) if match else model


def model_revisionless_key(value: Any) -> str:
    model = model_without_color(value)
    match = re.match(r"^(.+?\d+)[A-Z]+$", model)
    return match.group(1) if match else model


def refrigerator_capacity_from_model(model: str) -> int:
    normalized = normalize_model_key(model)
    if not normalized.startswith("AQR"):
        return 0
    match = re.search(r"(?:AQR-[A-Z]+|AQR-)(\d{2})", normalized)
    if not match:
        match = re.search(r"(\d{2})", normalized)
    if not match:
        return 0
    return int(match.group(1)) * 10


def refrigerator_category_from_capacity(capacity_liters: int) -> str:
    if capacity_liters >= 600:
        return "冷蔵庫600L以上"
    if capacity_liters >= 500:
        return "冷蔵庫500-599L"
    if capacity_liters >= 400:
        return "冷蔵庫400-499L"
    if capacity_liters >= 300:
        return "冷蔵庫300-399L"
    if capacity_liters >= 200:
        return "冷蔵庫200-299L"
    if capacity_liters >= 100:
        return "冷蔵庫100-199L"
    return ""


def washer_category_from_model(model: str, product_summary: str = "") -> str:
    normalized = normalize_model_key(model)
    summary = str(product_summary or "")
    if not normalized.startswith("AQW") and "洗濯" not in summary:
        return ""
    if "ドラム" in summary or re.search(r"AQW-DX|AQW-D", normalized):
        return "ドラム式洗濯機75ｋｇ以上"
    match = re.search(r"(\d{1,2})", normalized)
    capacity = int(match.group(1)) if match else 0
    if capacity >= 7:
        return "洗濯機縦型7ｋｇ以上"
    if capacity > 0:
        return "洗濯機縦型7ｋｇ未満"
    return ""


def inferred_model_rate(model: str, product_summary: str = "") -> dict[str, Any]:
    capacity = refrigerator_capacity_from_model(model)
    category = refrigerator_category_from_capacity(capacity)
    if category:
        return {
            "model": model,
            "install_category": category,
            "removal_category": f"{category} 搬出のみ",
            "match_type": "inferred_capacity",
            "inferred_capacity_liters": capacity,
        }
    category = washer_category_from_model(model, product_summary)
    if category:
        return {
            "model": model,
            "install_category": category,
            "removal_category": f"{category} 搬出のみ",
            "match_type": "inferred_washer",
        }
    return {}


def load_logistics_rate_master() -> dict[str, Any]:
    data = load_json_store(LOGISTICS_RATE_MASTER_FILE, {})
    return data if isinstance(data, dict) else {}


def rate_index(items: Any) -> dict[str, dict[str, Any]]:
    index: dict[str, dict[str, Any]] = {}
    if not isinstance(items, list):
        return index
    for item in items:
        if not isinstance(item, dict):
            continue
        label = str(item.get("label", "")).strip()
        if not label:
            continue
        index.setdefault(normalize_rate_key(label), item)
    return index


def remember_model_rate_if_needed(
    master: dict[str, Any],
    requested_model: str,
    model_rate: dict[str, Any],
) -> None:
    requested_model = str(requested_model or "").strip()
    if not requested_model or not model_rate:
        return
    if not model_rate.get("install_category") and not model_rate.get("removal_category"):
        return
    model_rates = master.setdefault("model_rates", [])
    if not isinstance(model_rates, list):
        return
    if any(
        normalize_model_key(item.get("model", "")) == normalize_model_key(requested_model)
        for item in model_rates
        if isinstance(item, dict)
    ):
        return
    model_rates.append(
        {
            "model": requested_model,
            "install_category": str(model_rate.get("install_category", "")),
            "removal_category": str(model_rate.get("removal_category", "")),
            "sheet": "自動補完",
            "row": "",
            "inferred_from_model": str(model_rate.get("model", "")),
            "match_type": str(model_rate.get("match_type", "auto")),
            "created_at": datetime.now().isoformat(timespec="seconds"),
        }
    )
    try:
        save_json_store(LOGISTICS_RATE_MASTER_FILE, master)
    except OSError:
        pass


def find_model_rate(
    master: dict[str, Any],
    model: str,
    product_summary: str = "",
) -> dict[str, Any]:
    target = normalize_model_key(model)
    if not target:
        return {}
    items = [item for item in master.get("model_rates", []) if isinstance(item, dict)]
    for item in items:
        if normalize_model_key(item.get("model", "")) == target:
            return {**item, "match_type": "exact"}

    no_color = model_without_color(model)
    for item in items:
        if model_without_color(item.get("model", "")) == no_color:
            return {**item, "match_type": "model_without_color"}

    revisionless = model_revisionless_key(model)
    family = model_family_key(model)
    for key_name, target_key in (
        ("revisionless", revisionless),
        ("family", family),
    ):
        if not target_key:
            continue
        candidates: list[dict[str, Any]] = []
        for item in items:
            item_model = item.get("model", "")
            item_key = (
                model_revisionless_key(item_model)
                if key_name == "revisionless"
                else model_family_key(item_model)
            )
            if item_key == target_key:
                candidates.append(item)
        if candidates:
            category_counts: dict[tuple[str, str], int] = {}
            for item in candidates:
                category = (
                    str(item.get("install_category", "")),
                    str(item.get("removal_category", "")),
                )
                category_counts[category] = category_counts.get(category, 0) + 1
            best_category = max(category_counts, key=category_counts.get)
            for item in candidates:
                category = (
                    str(item.get("install_category", "")),
                    str(item.get("removal_category", "")),
                )
                if category == best_category:
                    return {
                        **item,
                        "match_type": key_name,
                        "matched_candidates": len(candidates),
                    }

    for item in items:
        if not isinstance(item, dict):
            continue
        item_model = normalize_model_key(item.get("model", ""))
        if item_model and (item_model in target or target in item_model):
            return {**item, "match_type": "partial"}
    return inferred_model_rate(model, product_summary)


def lookup_rate_fee(index: dict[str, dict[str, Any]], label: str) -> int:
    item = index.get(normalize_rate_key(label))
    return coerce_int(item.get("fee_yen")) if item else 0


def is_removal_only_job(job: dict[str, Any]) -> bool:
    text = normalize_rate_key(
        " ".join(
            str(job.get(key, ""))
            for key in (
                "work_summary",
                "product_summary",
                "memo",
            )
        )
    )
    if any(keyword in text for keyword in ("商品交換", "交換", "入替", "設置")):
        return False
    return any(
        keyword in text
        for keyword in (
            "搬出のみ",
            "回収のみ",
            "引取のみ",
            "引取りのみ",
            "商品買取",
            "買取",
            "回収",
        )
    )


def job_product_kind(model: str, product_summary: str, install_category: str) -> str:
    normalized_model = normalize_model_key(model)
    combined = f"{product_summary} {install_category}"
    if (
        normalized_model.startswith(("JAA", "AQA"))
        or "エアコン" in combined
        or "kw" in normalize_rate_key(combined)
    ):
        return "aircon"
    if (
        normalized_model.startswith("AQW")
        or "洗濯" in combined
        or "ドラム" in combined
    ):
        return "washer"
    if normalized_model.startswith("AQR") or "冷蔵" in combined:
        return "refrigerator"
    return "general"


def refrigerator_size_token(category: str) -> str:
    normalized = normalize_rate_key(category)
    match = re.search(r"冷蔵庫(\d{3}[-－~～]\d{3}|600l以上|600l～|100[-－~～]199l)", normalized)
    if not match:
        return ""
    return match.group(1).replace("－", "-").replace("~", "～")


def additional_rate_priority(
    label: str,
    product_kind: str,
    install_category: str,
    removal_only: bool = False,
) -> int:
    normalized = normalize_rate_key(label)
    category = normalize_rate_key(install_category)
    if not normalized or normalized.startswith("【大阪第2】"):
        return 999

    if normalized == normalize_rate_key("搬出品配送費(搬出のみ）"):
        return 45 if removal_only else 140
    if normalized == normalize_rate_key("搬出品配送費(入替）"):
        return 46 if not removal_only else 140

    common_priority = {
        "階段上げ": 40,
        "階段揚げ": 40,
        "階段下ろし": 41,
        "重量物手当": 50,
        "時間指定": 51,
        "梱包作業費": 52,
        "人員増員費": 53,
        "見積費": 54,
        "商品引取り・配送費": 55,
    }
    for keyword, priority in common_priority.items():
        if normalize_rate_key(keyword) == normalized:
            return priority

    if product_kind == "refrigerator":
        size_token = refrigerator_size_token(install_category)
        if size_token and "冷蔵庫" in normalized and "搬出" in normalized:
            if normalize_rate_key(size_token) in normalized:
                return 5
        if "検品費" in normalized and "冷蔵庫" in normalized:
            return 20
        if "リサイクル費" in normalized and "冷蔵庫" in normalized:
            return 22
        if "冷蔵庫" in normalized and "搬出" in normalized:
            return 80

    if product_kind == "washer":
        if "ドラム式洗濯機75" in category and "ドラム式洗濯機75" in normalized:
            return 5
        if "7ｋｇ以上" in category and "縦型洗濯機7kg以上" in normalized:
            return 5
        if "7ｋｇ未満" in category and "縦型洗濯機7kg未満" in normalized:
            return 5
        if "洗濯機" in normalized and "搬出" in normalized:
            return 12
        if "検品費" in normalized and "洗濯機" in normalized:
            return 20
        if "リサイクル費" in normalized and "洗濯機" in normalized:
            return 22
        if any(normalize_rate_key(keyword) in normalized for keyword in ("排水溝清掃", "乾燥機移動", "乾燥機設置搬出")):
            return 24

    if product_kind == "aircon":
        if label.startswith("◇") or "エアコン" in normalized or "室外機" in normalized:
            return 10
        if any(
            keyword in normalized
            for keyword in (
                "配管",
                "コンセント",
                "分電盤",
                "はしご作業",
                "高所作業",
                "窒素",
                "フロン",
            )
        ):
            return 12

    if product_kind == "general":
        return 100
    return 999


def likely_worker_check_item(label: str) -> bool:
    normalized = normalize_rate_key(label)
    if not normalized or normalized.startswith("【大阪第2】"):
        return False
    excluded_prefixes = ("◎",)
    if str(label).startswith(excluded_prefixes):
        return False
    return True


def build_rate_suggestion(job: dict[str, Any]) -> dict[str, Any]:
    master = load_logistics_rate_master()
    if not master:
        return {}

    model = str(job.get("new_product_model") or job.get("old_product_model") or "").strip()
    product_summary = str(job.get("product_summary", ""))
    model_rate = find_model_rate(master, model, product_summary)
    remember_model_rate_if_needed(master, model, model_rate)
    billable_install = rate_index(master.get("billable_install_rates"))
    billable_work = rate_index(master.get("billable_work_item_rates"))
    subcontract_install = rate_index(master.get("subcontract_install_rates"))
    subcontract_work = rate_index(master.get("subcontract_work_item_rates"))

    install_category = str(model_rate.get("install_category", "")).strip()
    removal_category = str(model_rate.get("removal_category", "")).strip()
    product_kind = job_product_kind(model, product_summary, install_category)
    removal_only = is_removal_only_job(job)
    selected = {
        normalize_rate_key(label)
        for label in job.get("selected_rate_items", [])
        if str(label or "").strip()
    }

    seen_items: set[str] = set()
    work_items: list[dict[str, Any]] = []
    for index, item in enumerate(master.get("billable_work_item_rates", [])):
        if not isinstance(item, dict):
            continue
        label = str(item.get("label", "")).strip()
        key = normalize_rate_key(label)
        if key in seen_items or not likely_worker_check_item(label):
            continue
        seen_items.add(key)
        priority = additional_rate_priority(
            label,
            product_kind,
            install_category,
            removal_only,
        )
        selected_item = key in selected
        recommended = selected_item or priority < 100
        work_items.append(
            {
                "label": label,
                "billable_fee_yen": coerce_int(item.get("fee_yen")),
                "subcontract_fee_yen": lookup_rate_fee(subcontract_work, label),
                "selected": selected_item,
                "recommended": recommended,
                "category": "よく使う" if recommended else "その他",
                "priority": priority,
                "order": index,
                "source_cell": item.get("cell", ""),
            }
        )
    work_items.sort(
        key=lambda item: (
            0 if item.get("recommended") else 1,
            coerce_int(item.get("priority")),
            coerce_int(item.get("order")),
            str(item.get("label", "")),
        )
    )

    removal_fee_yen = 0
    subcontract_removal_fee_yen = 0
    if removal_only:
        removal_fee_yen = lookup_rate_fee(billable_work, removal_category) or lookup_rate_fee(
            billable_install,
            removal_category,
        )
        subcontract_removal_fee_yen = lookup_rate_fee(
            subcontract_work,
            removal_category,
        ) or lookup_rate_fee(subcontract_install, removal_category)

    return {
        "source": master.get("source", ""),
        "model": model,
        "matched_model": model_rate.get("model", ""),
        "match_type": model_rate.get("match_type", ""),
        "rate_note": (
            "引張AB列に完全一致がないため、自動補完した料金候補です。"
            if model_rate.get("match_type") not in {"", "exact"}
            else ""
        ),
        "install_category": install_category,
        "installation_fee_yen": lookup_rate_fee(billable_install, install_category),
        "subcontract_installation_fee_yen": lookup_rate_fee(
            subcontract_install,
            install_category,
        ),
        "removal_category": removal_category,
        "removal_fee_yen": removal_fee_yen,
        "subcontract_removal_fee_yen": subcontract_removal_fee_yen,
        "removal_mode": "removal_only" if removal_only else "additional_item",
        "removal_note": (
            "搬出のみの作業はこの候補、設置と搬出を含む作業は追加料金候補を使います。"
            if removal_only
            else "設置と搬出を含む作業はG/H列の追加料金候補を使います。"
        ),
        "work_items": work_items,
    }


def today_iso() -> str:
    return date.today().isoformat()


def sagyou_integration_config() -> dict[str, Any]:
    base_url = os.environ.get("SAGYOU_API_BASE_URL", "").strip().rstrip("/")
    api_key = os.environ.get("SAGYOU_INTEGRATION_API_KEY", "").strip()
    default_service = os.environ.get("SAGYOU_DEFAULT_SERVICE", "haier").strip()
    if default_service not in {"haier", "toshiba", "air_conditioner", "low_voltage"}:
        default_service = "haier"
    return {
        "configured": bool(base_url and api_key),
        "base_url": base_url,
        "api_key": api_key,
        "default_service": default_service,
        "default_start": os.environ.get("SAGYOU_DEFAULT_START", "09:00").strip(),
        "default_end": os.environ.get("SAGYOU_DEFAULT_END", "10:00").strip(),
    }


def sagyou_integration_key_is_valid(value: str) -> bool:
    expected = sagyou_integration_config().get("api_key", "")
    return bool(expected and value and secrets.compare_digest(str(expected), value))


def valid_sagyou_time(value: str) -> bool:
    return bool(re.fullmatch(r"(?:[01]\d|2[0-3]):[0-5]\d", value))


def build_sagyou_case_payload(
    job: dict[str, Any],
    config: dict[str, Any] | None = None,
) -> tuple[dict[str, Any] | None, str]:
    if str(job.get("source", "")) != "mail_import":
        return None, "メール取込案件ではありません。"

    corruption_error = corrupted_text_message(job, "連携対象の案件データ")
    if corruption_error:
        return None, corruption_error

    required = {
        "work_order_number": "作業番号",
        "customer_name": "お客様名",
        "customer_address": "住所",
        "scheduled_date": "作業日",
    }
    missing = [label for key, label in required.items() if not str(job.get(key, "")).strip()]
    if missing:
        return None, f"{', '.join(missing)}が未設定です。"

    config = config or sagyou_integration_config()
    scheduled_start = str(config.get("default_start", "09:00"))
    scheduled_end = str(config.get("default_end", "10:00"))
    if (
        not valid_sagyou_time(scheduled_start)
        or not valid_sagyou_time(scheduled_end)
        or scheduled_end <= scheduled_start
    ):
        return None, "連携用の標準訪問時間が正しくありません。"

    service = str(config.get("default_service", "haier"))
    category = "delivery" if service in {"haier", "toshiba"} else "construction"
    work_order_number = str(job.get("work_order_number", "")).strip()
    notes = "\n".join(
        value
        for value in (
            str(job.get("work_summary", "")).strip(),
            str(job.get("memo", "")).strip(),
        )
        if value
    )
    source_payload = {
        "source_system": "kanri-app",
        "kanri_job_id": str(job.get("id", "")),
        "work_order_number": work_order_number,
        "mail_import_id": str(job.get("source_mail_import_id", "")),
        "message_id": str(job.get("source_message_id", "")),
        "attachment_name": str(job.get("source_attachment_name", "")),
        "attachment_sha256": str(job.get("source_attachment_sha256", "")),
        "branch": str(job.get("branch", "")),
        "area": str(job.get("area", "")),
    }
    return {
        "external_id": f"KANRI-{work_order_number}"[:100],
        "work_order_number": work_order_number,
        "customer_name": str(job.get("customer_name", "")).strip(),
        "customer_phone": str(job.get("customer_phone", "")).strip() or None,
        "address": str(job.get("customer_address", "")).strip(),
        "category": category,
        "service": service,
        "scheduled_date": str(job.get("scheduled_date", "")).strip(),
        "scheduled_start": scheduled_start,
        "scheduled_end": scheduled_end,
        "product_name": str(job.get("product_summary", "")).strip() or None,
        "product_model": (
            str(job.get("new_product_model", "")).strip()
            or str(job.get("old_product_model", "")).strip()
            or None
        ),
        "retailer": (
            str(job.get("store_name", "")).strip()
            or str(job.get("branch", "")).strip()
            or None
        ),
        "notes": notes or None,
        "source_payload": source_payload,
    }, ""


def sync_logistics_job_to_sagyou(
    job: dict[str, Any],
    *,
    actor: str = "",
    force: bool = False,
) -> bool:
    if str(job.get("source", "")) != "mail_import":
        return False

    config = sagyou_integration_config()
    if not config["configured"]:
        next_status = "error" if force else "disabled"
        changed = job.get("sagyou_sync_status") != next_status
        job["sagyou_sync_status"] = next_status
        job["sagyou_last_error"] = "作業管理システム連携の環境変数が未設定です。"
        if force:
            attempted_at = datetime.now().isoformat(timespec="seconds")
            job["sagyou_last_attempt_at"] = attempted_at
            job["sagyou_last_http_status"] = 0
            job["sagyou_last_response"] = ""
            append_audit(
                "sync_sagyou_job_error",
                actor,
                str(job.get("work_order_number", "")),
                {
                    "job_id": job.get("id", ""),
                    "sent_at": attempted_at,
                    "http_status": 0,
                    "response_summary": "",
                    "error": job["sagyou_last_error"],
                },
            )
        return changed

    payload, reason = build_sagyou_case_payload(job, config)
    if payload is None:
        next_status = "error" if force else "pending"
        changed = (
            job.get("sagyou_sync_status") != next_status
            or job.get("sagyou_last_error") != reason
        )
        job["sagyou_sync_status"] = next_status
        job["sagyou_last_error"] = reason
        if force:
            attempted_at = datetime.now().isoformat(timespec="seconds")
            job["sagyou_last_attempt_at"] = attempted_at
            job["sagyou_last_http_status"] = 0
            job["sagyou_last_response"] = ""
            append_audit(
                "sync_sagyou_job_error",
                actor,
                str(job.get("work_order_number", "")),
                {
                    "job_id": job.get("id", ""),
                    "sent_at": attempted_at,
                    "http_status": 0,
                    "response_summary": "",
                    "error": reason,
                },
            )
        return changed

    inventory_model = str(job.get("new_product_model", "")).strip()
    if inventory_model:
        try:
            INVENTORY.validate_job_reservation(job)
        except InventoryError as exc:
            next_status = "error" if force else "pending"
            changed = (
                job.get("sagyou_sync_status") != next_status
                or job.get("sagyou_last_error") != str(exc)
            )
            job["sagyou_sync_status"] = next_status
            job["sagyou_last_error"] = str(exc)
            job["inventory_reservation_status"] = "error"
            if force:
                attempted_at = datetime.now().isoformat(timespec="seconds")
                job["sagyou_last_attempt_at"] = attempted_at
                append_audit(
                    "reserve_inventory_error",
                    actor,
                    str(job.get("work_order_number", "")),
                    {"job_id": job.get("id", ""), "error": str(exc)},
                )
            return changed

    payload_bytes = json.dumps(
        payload,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    ).encode("utf-8")
    payload_sha256 = hashlib.sha256(payload_bytes).hexdigest()
    if (
        not force
        and job.get("sagyou_sync_status") == "synced"
        and job.get("sagyou_payload_sha256") == payload_sha256
    ):
        return False

    attempted_at = datetime.now().isoformat(timespec="seconds")
    response_metadata: dict[str, Any] = {}
    job["sagyou_last_attempt_at"] = attempted_at
    try:
        response = request_json(
            f"{config['base_url']}/cases/intake",
            method="POST",
            headers={
                "Content-Type": "application/json; charset=utf-8",
                "X-Integration-Key": str(config["api_key"]),
                "User-Agent": "kanri-app/1.0",
            },
            body=payload_bytes,
            timeout=20,
            response_metadata=response_metadata,
        )
        remote_job_id = str(response.get("id", "")).strip()
        if not remote_job_id:
            raise DashboardError("作業管理システムから案件IDが返りませんでした。")
        response_summary = str(response_metadata.get("response_summary", ""))
        if not response_summary:
            response_summary = json.dumps(response, ensure_ascii=False, separators=(",", ":"))[:1000]
        http_status = int(response_metadata.get("http_status", 200) or 200)
        job["sagyou_sync_status"] = "synced"
        job["sagyou_job_id"] = remote_job_id
        job["sagyou_synced_at"] = datetime.now().isoformat(timespec="seconds")
        job["sagyou_payload_sha256"] = payload_sha256
        job["sagyou_last_error"] = ""
        job["sagyou_last_http_status"] = http_status
        job["sagyou_last_response"] = response_summary
        if inventory_model:
            try:
                reservation = INVENTORY.reserve_for_job(job, actor=actor)
                job["inventory_reservation_id"] = str(reservation.get("id", ""))
                job["inventory_reservation_status"] = str(
                    reservation.get("status", "reserved")
                )
            except InventoryError as exc:
                # The remote case already exists. Keep that fact visible and leave a
                # repairable status instead of retrying the external intake blindly.
                job["inventory_reservation_status"] = "error"
                job["sagyou_last_error"] = (
                    f"sagyou-appへの連携は成功しましたが、出庫予定を登録できませんでした: {exc}"
                )
                append_audit(
                    "reserve_inventory_after_sync_error",
                    actor,
                    str(job.get("work_order_number", "")),
                    {
                        "job_id": job.get("id", ""),
                        "sagyou_job_id": remote_job_id,
                        "error": str(exc),
                    },
                )
        append_audit(
            "sync_sagyou_job",
            actor,
            str(job.get("work_order_number", "")),
            {
                "job_id": job.get("id", ""),
                "sagyou_job_id": remote_job_id,
                "external_id": payload["external_id"],
                "sent_at": attempted_at,
                "http_status": http_status,
                "response_summary": response_summary,
                "payload_sha256": payload_sha256,
            },
        )
    except DashboardError as exc:
        http_status = int(
            getattr(exc, "http_status", response_metadata.get("http_status", 0)) or 0
        )
        response_summary = str(
            getattr(
                exc,
                "response_summary",
                response_metadata.get("response_summary", ""),
            )
            or ""
        )[:1000]
        job["sagyou_sync_status"] = "error"
        job["sagyou_last_error"] = str(exc)
        job["sagyou_last_http_status"] = http_status
        job["sagyou_last_response"] = response_summary
        append_audit(
            "sync_sagyou_job_error",
            actor,
            str(job.get("work_order_number", "")),
            {
                "job_id": job.get("id", ""),
                "external_id": payload["external_id"],
                "sent_at": attempted_at,
                "http_status": http_status,
                "response_summary": response_summary,
                "payload_sha256": payload_sha256,
                "error": str(exc),
            },
        )
    return True


def normalize_job_status(value: Any) -> str:
    status = str(value or "").strip()
    return status if status in JOB_STATUSES else "unprocessed"


def normalize_job_date(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    try:
        return date.fromisoformat(text).isoformat()
    except ValueError as exc:
        raise DashboardError("日付は yyyy-mm-dd 形式で入力してください。") from exc


def normalize_job(payload: dict[str, Any], existing: dict[str, Any] | None = None) -> dict[str, Any]:
    existing = existing or {}
    now = datetime.now().isoformat(timespec="seconds")
    work_order_number = str(
        payload.get("work_order_number", existing.get("work_order_number", ""))
    ).strip()
    if not work_order_number:
        raise DashboardError("作業番号を入力してください。")

    status = normalize_job_status(payload.get("status", existing.get("status")))
    scheduled_date = normalize_job_date(
        payload.get("scheduled_date", existing.get("scheduled_date", ""))
    )
    if status == "unprocessed" and scheduled_date:
        status = "scheduled"
    selected_rate_items = payload.get(
        "selected_rate_items",
        existing.get("selected_rate_items", []),
    )
    if not isinstance(selected_rate_items, list):
        selected_rate_items = []
    subcontractor_code = normalize_contractor_code(
        payload.get("subcontractor_code", existing.get("subcontractor_code", ""))
    )
    subcontractor = subcontractor_by_code(subcontractor_code) if subcontractor_code else {}
    subcontractor_issued_at = str(existing.get("subcontractor_issued_at", ""))
    if payload.get("subcontractor_issue") and subcontractor_code:
        subcontractor_issued_at = now

    job = {
        "id": str(existing.get("id") or payload.get("id") or secrets.token_hex(12)),
        "work_order_number": work_order_number,
        "status": status,
        "scheduled_date": scheduled_date,
        "customer_name": str(
            payload.get("customer_name", existing.get("customer_name", ""))
        ).strip(),
        "customer_phone": str(
            payload.get("customer_phone", existing.get("customer_phone", ""))
        ).strip(),
        "customer_address": str(
            payload.get("customer_address", existing.get("customer_address", ""))
        ).strip(),
        "area": str(payload.get("area", existing.get("area", ""))).strip(),
        "branch": str(payload.get("branch", existing.get("branch", ""))).strip(),
        "store_name": str(payload.get("store_name", existing.get("store_name", ""))).strip(),
        "staff_name": str(payload.get("staff_name", existing.get("staff_name", ""))).strip(),
        "vehicle_number": str(
            payload.get("vehicle_number", existing.get("vehicle_number", ""))
        ).strip(),
        "old_product_model": str(
            payload.get("old_product_model", existing.get("old_product_model", ""))
        ).strip(),
        "new_product_model": str(
            payload.get("new_product_model", existing.get("new_product_model", ""))
        ).strip(),
        "product_summary": str(
            payload.get("product_summary", existing.get("product_summary", ""))
        ).strip(),
        "work_summary": str(
            payload.get("work_summary", existing.get("work_summary", ""))
        ).strip(),
        "memo": str(payload.get("memo", existing.get("memo", ""))).strip(),
        "source": str(payload.get("source", existing.get("source", ""))).strip(),
        "source_mail_import_id": str(
            payload.get("source_mail_import_id", existing.get("source_mail_import_id", ""))
        ).strip(),
        "source_message_id": str(
            payload.get("source_message_id", existing.get("source_message_id", ""))
        ).strip(),
        "source_attachment_name": str(
            payload.get("source_attachment_name", existing.get("source_attachment_name", ""))
        ).strip(),
        "source_attachment_path": str(
            payload.get("source_attachment_path", existing.get("source_attachment_path", ""))
        ).strip(),
        "source_attachment_sha256": str(
            payload.get("source_attachment_sha256", existing.get("source_attachment_sha256", ""))
        ).strip(),
        "sagyou_sync_status": str(
            payload.get("sagyou_sync_status", existing.get("sagyou_sync_status", ""))
        ).strip(),
        "sagyou_job_id": str(
            payload.get("sagyou_job_id", existing.get("sagyou_job_id", ""))
        ).strip(),
        "sagyou_synced_at": str(
            payload.get("sagyou_synced_at", existing.get("sagyou_synced_at", ""))
        ).strip(),
        "sagyou_payload_sha256": str(
            payload.get("sagyou_payload_sha256", existing.get("sagyou_payload_sha256", ""))
        ).strip(),
        "sagyou_last_error": str(
            payload.get("sagyou_last_error", existing.get("sagyou_last_error", ""))
        ).strip(),
        "sagyou_last_attempt_at": str(
            payload.get(
                "sagyou_last_attempt_at",
                existing.get("sagyou_last_attempt_at", ""),
            )
        ).strip(),
        "sagyou_last_http_status": int(
            payload.get(
                "sagyou_last_http_status",
                existing.get("sagyou_last_http_status", 0),
            )
            or 0
        ),
        "sagyou_last_response": str(
            payload.get(
                "sagyou_last_response",
                existing.get("sagyou_last_response", ""),
            )
        ).strip()[:1000],
        "inventory_reservation_id": str(
            payload.get(
                "inventory_reservation_id",
                existing.get("inventory_reservation_id", ""),
            )
        ).strip(),
        "inventory_reservation_status": str(
            payload.get(
                "inventory_reservation_status",
                existing.get("inventory_reservation_status", ""),
            )
        ).strip(),
        "subcontractor_code": subcontractor_code,
        "subcontractor_name": str(
            subcontractor.get("company_name")
            or payload.get("subcontractor_name", existing.get("subcontractor_name", ""))
        ).strip(),
        "subcontractor_issued_at": subcontractor_issued_at,
        "return_shipment_manifest_id": str(
            payload.get(
                "return_shipment_manifest_id",
                existing.get("return_shipment_manifest_id", ""),
            )
        ).strip(),
        "return_shipment_exported_at": str(
            payload.get(
                "return_shipment_exported_at",
                existing.get("return_shipment_exported_at", ""),
            )
        ).strip(),
        "return_shipment_destination_key": str(
            payload.get(
                "return_shipment_destination_key",
                existing.get("return_shipment_destination_key", ""),
            )
        ).strip(),
        "raw_payload": payload.get("raw_payload", existing.get("raw_payload", {}))
        if isinstance(payload.get("raw_payload", existing.get("raw_payload", {})), dict)
        else {},
        "selected_rate_items": [
            str(item).strip()
            for item in selected_rate_items
            if str(item or "").strip()
        ],
        "created_at": str(existing.get("created_at") or now),
        "updated_at": now,
    }
    for field in JOB_MONEY_FIELDS:
        job[field] = max(0, coerce_int(payload.get(field, existing.get(field, 0))))
    job["sales_total_yen"] = (
        job["installation_fee_yen"]
        + job["distance_fee_yen"]
        + job["toll_fee_yen"]
        + job["parking_fee_yen"]
        + job["other_fee_yen"]
    )
    job["gross_profit_yen"] = job["sales_total_yen"] - job["subcontract_fee_yen"]
    job["etc_link_status"] = "linked" if job["toll_fee_yen"] > 0 else "unlinked"
    job["rate_suggestion"] = build_rate_suggestion(job)
    return job


def load_logistics_jobs() -> list[dict[str, Any]]:
    data = load_json_store(LOGISTICS_JOBS_FILE, [])
    if isinstance(data, dict):
        data = data.get("jobs", [])
    if not isinstance(data, list):
        return []
    jobs: list[dict[str, Any]] = []
    for item in data:
        if not isinstance(item, dict):
            continue
        try:
            jobs.append(normalize_job(item, item))
        except DashboardError:
            continue
    jobs.sort(
        key=lambda job: (
            job.get("scheduled_date") or "9999-12-31",
            job.get("work_order_number", ""),
        )
    )
    return jobs


def save_logistics_jobs(jobs: list[dict[str, Any]]) -> None:
    save_json_store(LOGISTICS_JOBS_FILE, jobs)


def logistics_jobs_payload(
    query: dict[str, list[str]],
    current_user: dict[str, Any] | None = None,
) -> dict[str, Any]:
    jobs = load_logistics_jobs()
    current_user = current_user or {}
    if current_user.get("role") == "contractor":
        contractor_code = normalize_contractor_code(current_user.get("contractor_code"))
        jobs = [
            job
            for job in jobs
            if normalize_contractor_code(job.get("subcontractor_code")) == contractor_code
            and str(job.get("subcontractor_issued_at", "")).strip()
        ]
    status = (query.get("status") or [""])[0].strip()
    month = (query.get("month") or [""])[0].strip()
    keyword = (query.get("keyword") or [""])[0].strip().lower()
    area = (query.get("area") or [""])[0].strip()
    vehicle = (query.get("vehicle") or [""])[0].strip()

    filtered = jobs
    if status:
        filtered = [job for job in filtered if job.get("status") == status]
    if month:
        filtered = [
            job
            for job in filtered
            if str(job.get("scheduled_date", "")).startswith(f"{month}-")
        ]
    if area:
        filtered = [job for job in filtered if job.get("area") == area]
    if vehicle:
        vehicle_numbers = vehicle_group_numbers(vehicle)
        filtered = [
            job
            for job in filtered
            if str(job.get("vehicle_number", "")).strip() in vehicle_numbers
        ]
    if keyword:
        searchable_fields = (
            "work_order_number",
            "customer_name",
            "area",
            "branch",
            "store_name",
            "staff_name",
            "vehicle_number",
            "product_summary",
            "work_summary",
        )
        filtered = [
            job
            for job in filtered
            if any(keyword in str(job.get(field, "")).lower() for field in searchable_fields)
        ]

    summary = {
        "count": len(filtered),
        "sales_total_yen": sum(job.get("sales_total_yen", 0) for job in filtered),
        "toll_total_yen": sum(job.get("toll_fee_yen", 0) for job in filtered),
        "parking_total_yen": sum(job.get("parking_fee_yen", 0) for job in filtered),
        "gross_profit_yen": sum(job.get("gross_profit_yen", 0) for job in filtered),
        "needs_review_count": sum(1 for job in filtered if job.get("status") == "needs_review"),
        "etc_unlinked_count": sum(1 for job in filtered if job.get("toll_fee_yen", 0) == 0),
    }
    return {
        "jobs": filtered,
        "summary": summary,
        "statuses": [{"value": key, "label": value} for key, value in JOB_STATUSES.items()],
        "areas": sorted({str(job.get("area", "")).strip() for job in jobs if job.get("area")}),
        "vehicles": sorted(
            {str(job.get("vehicle_number", "")).strip() for job in jobs if job.get("vehicle_number")}
        ),
        "updated_at": datetime.now().isoformat(timespec="seconds"),
    }


def save_logistics_job(payload: dict[str, Any], actor: str = "") -> dict[str, Any]:
    raise_for_corrupted_text(payload, "保存対象の案件データ")
    jobs = load_logistics_jobs()
    job_id = str(payload.get("id", "")).strip()
    work_order_number = str(payload.get("work_order_number", "")).strip()
    existing_index = next(
        (index for index, job in enumerate(jobs) if str(job.get("id")) == job_id),
        None,
    )
    if existing_index is None and work_order_number:
        existing_index = next(
            (
                index
                for index, job in enumerate(jobs)
                if str(job.get("work_order_number", "")).strip() == work_order_number
            ),
            None,
        )
    existing_job = jobs[existing_index] if existing_index is not None else None
    normalized = normalize_job(payload, existing_job)
    if existing_index is None:
        jobs.append(normalized)
        action = "create_logistics_job"
    else:
        jobs[existing_index] = normalized
        action = "update_logistics_job"
    save_logistics_jobs(jobs)
    append_audit(action, actor, normalized["work_order_number"], {"job_id": normalized["id"]})
    if sync_logistics_job_to_sagyou(normalized, actor=actor):
        save_logistics_jobs(jobs)
    return normalized


def sync_mail_jobs_to_sagyou(actor: str = "", force: bool = False) -> dict[str, int]:
    jobs = load_logistics_jobs()
    targeted = 0
    changed = 0
    for job in jobs:
        if str(job.get("source", "")) != "mail_import":
            continue
        targeted += 1
        if sync_logistics_job_to_sagyou(job, actor=actor, force=force):
            changed += 1
    if changed:
        save_logistics_jobs(jobs)
    return {
        "targeted": targeted,
        "changed": changed,
        "synced": sum(1 for job in jobs if job.get("sagyou_sync_status") == "synced"),
        "pending": sum(1 for job in jobs if job.get("sagyou_sync_status") == "pending"),
        "errors": sum(1 for job in jobs if job.get("sagyou_sync_status") == "error"),
        "disabled": sum(1 for job in jobs if job.get("sagyou_sync_status") == "disabled"),
    }


def sync_mail_job_to_sagyou(
    job_id: str = "",
    actor: str = "",
    scheduled_date: str = "",
    work_order_number: str = "",
) -> dict[str, Any]:
    requested_id = str(job_id or "").strip()
    requested_work_order = str(work_order_number or "").strip()
    if not requested_id and not requested_work_order:
        raise DashboardError("連携する案件を指定してください。")
    jobs = load_logistics_jobs()
    if requested_id:
        job = next((item for item in jobs if str(item.get("id", "")) == requested_id), None)
    else:
        job = next(
            (
                item
                for item in jobs
                if str(item.get("work_order_number", "")).strip() == requested_work_order
            ),
            None,
        )
    if job is None:
        raise DashboardError("指定された案件が見つかりません。")
    if str(job.get("source", "")) != "mail_import":
        raise DashboardError("メールから取り込んだ案件だけ連携できます。")

    requested_date = str(scheduled_date or "").strip()
    if requested_date:
        job["scheduled_date"] = normalize_job_date(requested_date)
        if job.get("status") == "unprocessed":
            job["status"] = "scheduled"
        job["updated_at"] = datetime.now().isoformat(timespec="seconds")

    sync_logistics_job_to_sagyou(job, actor=actor, force=True)
    save_logistics_jobs(jobs)
    return {
        "id": job.get("id", ""),
        "work_order_number": job.get("work_order_number", ""),
        "sagyou_sync_status": job.get("sagyou_sync_status", ""),
        "sagyou_job_id": job.get("sagyou_job_id", ""),
        "sagyou_synced_at": job.get("sagyou_synced_at", ""),
        "sagyou_last_error": job.get("sagyou_last_error", ""),
        "sagyou_last_attempt_at": job.get("sagyou_last_attempt_at", ""),
        "sagyou_last_http_status": job.get("sagyou_last_http_status", 0),
        "sagyou_last_response": job.get("sagyou_last_response", ""),
        "inventory_reservation_id": job.get("inventory_reservation_id", ""),
        "inventory_reservation_status": job.get("inventory_reservation_status", ""),
    }


RETURN_DESTINATIONS = {
    "maizuru": {
        "key": "maizuru",
        "name": "舞鶴倉庫",
        "title": "AQUA洗濯機 舞鶴倉庫配送分",
        "address": "京都府向日市鶏冠井町馬司25",
        "detail": "舞鶴倉庫 株式会社京都事業所 構内",
        "company": "アクア株式会社 京都検品センター",
        "phone": "075-924-1318",
    },
    "gunma": {
        "key": "gunma",
        "name": "群馬製品受け入れセンター",
        "title": "AQUA冷蔵庫 群馬製品受け入れセンター配送分",
        "address": "群馬県邑楽郡千代田福島611-1",
        "detail": "",
        "company": "群馬製品受け入れセンター",
        "phone": "0276-91-5877・5878",
    },
}


def logistics_job_excel_fields(job: dict[str, Any]) -> dict[str, Any]:
    raw_payload = job.get("raw_payload")
    if not isinstance(raw_payload, dict):
        return {}
    excel_fields = raw_payload.get("excel_fields")
    return excel_fields if isinstance(excel_fields, dict) else {}


def labeled_memo_value(job: dict[str, Any], label: str) -> str:
    memo = str(job.get("memo", ""))
    pattern = rf"^{re.escape(label)}:\s*(.+)$"
    for line in memo.splitlines():
        match = re.match(pattern, line.strip())
        if match:
            return match.group(1).strip()
    return ""


def logistics_job_detail(job: dict[str, Any], key: str, memo_label: str = "") -> str:
    fields = logistics_job_excel_fields(job)
    value = str(fields.get(key, "")).strip()
    if value:
        return value
    return labeled_memo_value(job, memo_label or key)


def return_product_model(job: dict[str, Any]) -> str:
    return str(
        job.get("old_product_model")
        or job.get("new_product_model")
        or ""
    ).strip()


def return_destination_for_job(job: dict[str, Any]) -> dict[str, str]:
    model = return_product_model(job).upper()
    summary = str(job.get("product_summary", ""))
    if model.startswith("AQW") or "洗濯" in summary:
        return RETURN_DESTINATIONS["maizuru"]
    if model.startswith("AQR") or "冷蔵" in summary:
        return RETURN_DESTINATIONS["gunma"]
    return RETURN_DESTINATIONS["gunma"]


def return_shipment_row(job: dict[str, Any]) -> dict[str, Any]:
    destination = return_destination_for_job(job)
    return {
        "id": str(job.get("id", "")),
        "work_order_number": str(job.get("work_order_number", "")),
        "product_summary": str(job.get("product_summary", "")),
        "sto_slip": "回収のみ",
        "branch": str(job.get("branch") or job.get("area") or ""),
        "product_model": return_product_model(job),
        "product_serial": logistics_job_detail(job, "product_serial", "製造番号"),
        "approval_number": logistics_job_detail(job, "approval_number", "承認番号"),
        "customer_name": str(job.get("customer_name", "")),
        "application_type": str(job.get("work_summary", "")),
        "symptom": logistics_job_detail(job, "symptom", "症状"),
        "destination_key": destination["key"],
        "destination_name": destination["name"],
        "scheduled_date": str(job.get("scheduled_date", "")),
    }


def return_shipment_candidates() -> list[dict[str, Any]]:
    return [
        return_shipment_row(job)
        for job in load_logistics_jobs()
        if job.get("status") == "completed"
        and not str(job.get("return_shipment_manifest_id", "")).strip()
    ]


def load_return_shipments() -> list[dict[str, Any]]:
    data = load_json_store(RETURN_SHIPMENTS_FILE, [])
    if isinstance(data, dict):
        data = data.get("shipments", [])
    return data if isinstance(data, list) else []


def save_return_shipments(shipments: list[dict[str, Any]]) -> None:
    save_json_store(RETURN_SHIPMENTS_FILE, shipments)


def return_shipments_payload() -> dict[str, Any]:
    candidates = return_shipment_candidates()
    by_destination: dict[str, int] = {}
    for row in candidates:
        key = str(row.get("destination_key", ""))
        by_destination[key] = by_destination.get(key, 0) + 1
    return {
        "candidates": candidates,
        "destinations": list(RETURN_DESTINATIONS.values()),
        "summary": {
            "candidate_count": len(candidates),
            "by_destination": by_destination,
            "manifest_count": len(load_return_shipments()),
        },
        "manifests": load_return_shipments()[-20:],
        "updated_at": datetime.now().isoformat(timespec="seconds"),
    }


def build_return_shipment_workbook(rows: list[dict[str, Any]]) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise DashboardError("Excel発行に必要な openpyxl が見つかりません。") from exc

    workbook = Workbook()
    workbook.remove(workbook.active)
    grouped: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        grouped.setdefault(str(row.get("destination_key", "")), []).append(row)

    headers = [
        "作業指示番号",
        "製品名",
        "STO伝票",
        "依頼部署",
        "品番",
        "製造番号",
        "承認番号",
        "お客様名",
        "申請区分",
        "症状",
    ]
    widths = [5, 14, 16, 14, 18, 20, 14, 18, 18, 26, 36]
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for destination_key, destination_rows in grouped.items():
        destination = RETURN_DESTINATIONS.get(destination_key, RETURN_DESTINATIONS["gunma"])
        sheet = workbook.create_sheet(destination["name"][:31])
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
        sheet.sheet_view.showGridLines = False
        sheet.merge_cells("A1:K1")
        sheet["A1"] = destination["title"]
        sheet["A1"].font = Font(size=14, bold=True)
        sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
        sheet.merge_cells("A2:K2")
        sheet["A2"] = f"配送先: {destination['name']} / {destination['address']}"
        sheet.merge_cells("A3:K3")
        detail = " / ".join(
            part
            for part in (destination.get("detail"), destination.get("company"), destination.get("phone"))
            if part
        )
        sheet["A3"] = detail
        for row_index in range(1, 4):
            sheet.row_dimensions[row_index].height = 22

        header_row = 5
        sheet.cell(header_row, 1).value = ""
        for index, header in enumerate(headers, start=2):
            cell = sheet.cell(header_row, index)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill("solid", fgColor="F2F7F8")
            cell.border = border

        keys = [
            "work_order_number",
            "product_summary",
            "sto_slip",
            "branch",
            "product_model",
            "product_serial",
            "approval_number",
            "customer_name",
            "application_type",
            "symptom",
        ]
        for number_index, row in enumerate(destination_rows, start=1):
            sheet_row = header_row + number_index
            sheet.cell(sheet_row, 1).value = number_index
            sheet.cell(sheet_row, 1).alignment = Alignment(horizontal="center")
            for column_index, key in enumerate(keys, start=2):
                cell = sheet.cell(sheet_row, column_index)
                cell.value = row.get(key, "")
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border
            sheet.row_dimensions[sheet_row].height = 22

        for column_index, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(column_index)].width = width

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def unique_return_shipment_path(file_name: str) -> Path:
    RETURN_SHIPMENT_EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    stem = safe_submission_filename(Path(file_name).stem or "お帰り便配送表")
    suffix = ".xlsx"
    candidate = RETURN_SHIPMENT_EXPORTS_DIR / f"{stem}{suffix}"
    counter = 2
    while candidate.exists():
        candidate = RETURN_SHIPMENT_EXPORTS_DIR / f"{stem}_{counter}{suffix}"
        counter += 1
    return candidate


def create_return_shipment_export(payload: dict[str, Any], actor: str = "") -> dict[str, Any]:
    selected_ids = {
        str(item)
        for item in payload.get("job_ids", [])
        if str(item).strip()
    }
    candidates = return_shipment_candidates()
    selected_rows = [
        row
        for row in candidates
        if not selected_ids or str(row.get("id", "")) in selected_ids
    ]
    if not selected_rows:
        raise DashboardError("配送表に入れる作業完了案件がありません。")

    workbook_bytes = build_return_shipment_workbook(selected_rows)
    manifest_id = secrets.token_hex(12)
    exported_at = datetime.now().isoformat(timespec="seconds")
    destination_names = sorted({str(row.get("destination_name", "")) for row in selected_rows})
    file_name = f"お帰り便配送表_{date.today().isoformat()}_{manifest_id[:6]}.xlsx"
    output_path = unique_return_shipment_path(file_name)
    output_path.write_bytes(workbook_bytes)
    sync_path_to_cloud(output_path)
    relative_path = str(output_path.relative_to(RETURN_SHIPMENT_EXPORTS_DIR)).replace("\\", "/")

    jobs = load_logistics_jobs()
    selected_job_ids = {str(row.get("id", "")) for row in selected_rows}
    for job in jobs:
        if str(job.get("id", "")) in selected_job_ids:
            destination = return_destination_for_job(job)
            job["return_shipment_manifest_id"] = manifest_id
            job["return_shipment_exported_at"] = exported_at
            job["return_shipment_destination_key"] = destination["key"]
    save_logistics_jobs(jobs)

    manifest = {
        "id": manifest_id,
        "created_at": exported_at,
        "created_by": actor,
        "job_ids": list(selected_job_ids),
        "job_count": len(selected_rows),
        "destinations": destination_names,
        "file_name": output_path.name,
        "relative_path": relative_path,
        "url": f"/api/return-shipments/download?file={quote(relative_path, safe='')}",
    }
    shipments = load_return_shipments()
    shipments.append(manifest)
    save_return_shipments(shipments)
    append_audit(
        "create_return_shipment",
        actor,
        manifest_id,
        {"job_count": len(selected_rows), "destinations": destination_names},
    )
    return manifest


def resolve_return_shipment_export(relative_path: str) -> Path:
    candidate = (RETURN_SHIPMENT_EXPORTS_DIR / relative_path).resolve()
    root = RETURN_SHIPMENT_EXPORTS_DIR.resolve()
    if not str(candidate).startswith(str(root)) or not candidate.exists():
        raise DashboardError("配送表ファイルが見つかりません。")
    return candidate


LOGISTICS_EXCEL_CELL_MAP = {
    "入力シート": (
        "D10",
        "D11",
        "D12",
        "D13",
        "D22",
        "D23",
        "D24",
        "D25",
        "D28",
        "D29",
        "D30",
        "D31",
        "D32",
        "D33",
        "D34",
        "D46",
        "D47",
        "D48",
        "D49",
        "D50",
        "D52",
        "D54",
        "D55",
        "D56",
        "D57",
        "D58",
        "D59",
        "D60",
        "D61",
        "D62",
        "D63",
        "D64",
    ),
    "データ抜取シート": ("P3", "R3", "S3", "U3", "V3", "W3", "H7"),
    "申請書": ("E54", "F54", "E55", "F55", "E56", "F56", "E57", "F57"),
}


def clean_excel_text(value: Any) -> str:
    text = str(value or "").replace("\r", "\n").strip()
    text = re.sub(r"[ \t\u3000]+", " ", text)
    text = re.sub(r"\n+", "\n", text)
    return "" if text in {"-", "0", "入力忘れ", "未入力", "なし"} else text


def join_excel_text(*values: Any) -> str:
    return " ".join(part for part in (clean_excel_text(value) for value in values) if part)


def read_logistics_excel_cells_openpyxl(path: Path) -> dict[str, dict[str, str]]:
    if path.suffix.lower() == ".xls":
        raise DashboardError("古いExcel形式です。")
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise DashboardError("openpyxl が見つからないためExcelを読み取れません。") from exc

    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        result: dict[str, dict[str, str]] = {}
        for sheet_name, addresses in LOGISTICS_EXCEL_CELL_MAP.items():
            if sheet_name not in workbook.sheetnames:
                continue
            sheet = workbook[sheet_name]
            result[sheet_name] = {
                address: clean_excel_text(sheet[address].value)
                for address in addresses
            }
        return result
    finally:
        workbook.close()


def read_logistics_excel_cells_xlrd(path: Path) -> dict[str, dict[str, str]]:
    if path.suffix.lower() != ".xls":
        raise DashboardError("xlrdの対象ではないExcel形式です。")
    try:
        import xlrd
    except ImportError as exc:
        raise DashboardError("xlrd が見つからないため旧Excelを読み取れません。") from exc
    try:
        workbook = xlrd.open_workbook(str(path), on_demand=True)
    except Exception as exc:
        raise DashboardError(f"旧Excel添付を開けませんでした: {type(exc).__name__}") from exc
    try:
        result: dict[str, dict[str, str]] = {}
        for sheet_name, addresses in LOGISTICS_EXCEL_CELL_MAP.items():
            if sheet_name not in workbook.sheet_names():
                continue
            sheet = workbook.sheet_by_name(sheet_name)
            values: dict[str, str] = {}
            for address in addresses:
                row_index, column_index = excel_a1_position(address)
                if row_index >= sheet.nrows or column_index >= sheet.ncols:
                    values[address] = ""
                    continue
                cell = sheet.cell(row_index, column_index)
                value: Any = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE:
                    value = xlrd.xldate_as_datetime(cell.value, workbook.datemode).isoformat()
                elif cell.ctype == xlrd.XL_CELL_NUMBER and float(cell.value).is_integer():
                    value = str(int(cell.value))
                elif cell.ctype in {xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK, xlrd.XL_CELL_ERROR}:
                    value = ""
                values[address] = clean_excel_text(value)
            result[sheet_name] = values
        return result
    finally:
        workbook.release_resources()


def read_logistics_excel_cells_with_excel(path: Path) -> dict[str, dict[str, str]]:
    if not LOGISTICS_EXCEL_EXTRACT_SCRIPT.exists():
        raise DashboardError("Excel読取スクリプトが見つかりません。")
    output_fd, output_name = tempfile.mkstemp(prefix="logistics_excel_", suffix=".json")
    os.close(output_fd)
    output_path = Path(output_name)
    try:
        completed = subprocess.run(
            [
                "powershell.exe",
                "-NoProfile",
                "-ExecutionPolicy",
                "Bypass",
                "-File",
                str(LOGISTICS_EXCEL_EXTRACT_SCRIPT),
                "-InputPath",
                str(path),
                "-OutputPath",
                str(output_path),
            ],
            check=False,
            capture_output=True,
            text=True,
            timeout=90,
        )
        if completed.returncode != 0:
            detail = (completed.stderr or completed.stdout or "").strip()
            raise DashboardError(f"Excel添付を読み取れませんでした。{detail}")
        data = json.loads(output_path.read_text(encoding="utf-8-sig"))
        if not isinstance(data, dict):
            raise DashboardError("Excel添付の読取結果が不正です。")
        return {
            str(sheet): {
                str(address): clean_excel_text(value)
                for address, value in cells.items()
            }
            for sheet, cells in data.items()
            if isinstance(cells, dict)
        }
    finally:
        try:
            output_path.unlink(missing_ok=True)
        except OSError:
            pass


def read_logistics_excel_cells(path: Path) -> dict[str, dict[str, str]]:
    if path.suffix.lower() == ".xls":
        try:
            return read_logistics_excel_cells_xlrd(path)
        except DashboardError as xlrd_error:
            try:
                return read_logistics_excel_cells_with_excel(path)
            except DashboardError as excel_error:
                raise DashboardError(f"{xlrd_error} {excel_error}") from excel_error
    try:
        return read_logistics_excel_cells_openpyxl(path)
    except DashboardError as openpyxl_error:
        try:
            return read_logistics_excel_cells_with_excel(path)
        except DashboardError as excel_error:
            raise DashboardError(f"{openpyxl_error} {excel_error}") from excel_error


def infer_work_order_number(*values: str) -> str:
    for value in values:
        match = re.search(r"\b(\d{8})\b", value)
        if match:
            return match.group(1)
    return ""


def infer_area(branch: str, address: str) -> str:
    if branch:
        return branch
    match = re.match(r"(.+?[都道府県])", address)
    return match.group(1) if match else ""


def build_import_job_memo(fields: dict[str, str]) -> str:
    notes = []
    for label, key in (
        ("住所", "customer_address"),
        ("連絡先", "customer_phone"),
        ("製造番号", "product_serial"),
        ("症状", "symptom"),
        ("申請理由", "request_reason"),
        ("買取金額", "purchase_amount_yen"),
        ("補償金額", "compensation_amount_yen"),
        ("修理代負担金", "repair_burden_fee_yen"),
        ("お客さま負担金", "customer_burden_fee_yen"),
        ("交換上の注意", "work_note"),
        ("配送先", "delivery_summary"),
        ("承認番号", "approval_number"),
        ("承認コメント", "approval_comment"),
    ):
        value = clean_excel_text(fields.get(key, ""))
        if value:
            notes.append(f"{label}: {value}")
    return "\n".join(notes)


def extract_logistics_job_payload_from_excel(path: Path) -> dict[str, Any]:
    cells = read_logistics_excel_cells(path)
    input_sheet = cells.get("入力シート", {})
    extract_sheet = cells.get("データ抜取シート", {})
    application_sheet = cells.get("申請書", {})
    customer_address = join_excel_text(
        input_sheet.get("D29", ""),
        input_sheet.get("D30", ""),
        input_sheet.get("D31", ""),
    )
    delivery_summary = join_excel_text(
        input_sheet.get("D59", ""),
        input_sheet.get("D60", ""),
        input_sheet.get("D62", ""),
        input_sheet.get("D63", ""),
        input_sheet.get("D64", ""),
    )
    branch = clean_excel_text(input_sheet.get("D11", ""))
    old_model = clean_excel_text(input_sheet.get("D47", ""))
    new_model = clean_excel_text(extract_sheet.get("W3", "")) or clean_excel_text(
        extract_sheet.get("H7", "")
    )
    fields = {
        "work_order_number": clean_excel_text(input_sheet.get("D23", ""))
        or infer_work_order_number(path.name),
        "branch": branch,
        "staff_name": clean_excel_text(input_sheet.get("D12", "")),
        "customer_name": clean_excel_text(input_sheet.get("D24", "")),
        "customer_phone": clean_excel_text(input_sheet.get("D32", "")),
        "customer_address": customer_address,
        "store_name": clean_excel_text(input_sheet.get("D34", "")),
        "product_summary": clean_excel_text(input_sheet.get("D46", "")),
        "old_product_model": old_model,
        "new_product_model": new_model or old_model,
        "product_serial": clean_excel_text(input_sheet.get("D48", "")),
        "symptom": clean_excel_text(input_sheet.get("D49", "")),
        "request_reason": clean_excel_text(input_sheet.get("D50", "")),
        "application_type": clean_excel_text(input_sheet.get("D52", "")),
        "work_note": clean_excel_text(input_sheet.get("D58", "")),
        "delivery_summary": delivery_summary,
        "approval_number": clean_excel_text(extract_sheet.get("U3", "")),
        "approval_comment": clean_excel_text(extract_sheet.get("P3", "")),
        "purchase_amount_yen": clean_excel_text(application_sheet.get("F54", "")),
        "compensation_amount_yen": clean_excel_text(application_sheet.get("F55", "")),
        "repair_burden_fee_yen": clean_excel_text(application_sheet.get("F56", "")),
        "customer_burden_fee_yen": clean_excel_text(application_sheet.get("F57", "")),
    }
    raise_for_corrupted_text(fields, "Excel取込データ")
    if not fields["work_order_number"]:
        raise DashboardError("Excelから作業番号を読み取れませんでした。")
    purchase_amount_yen = coerce_int(fields.get("purchase_amount_yen", ""))
    other_fee_yen = sum(
        coerce_int(fields.get(key, ""))
        for key in (
            "compensation_amount_yen",
            "repair_burden_fee_yen",
            "customer_burden_fee_yen",
        )
    )
    payload = {
        "work_order_number": fields["work_order_number"],
        "status": "unprocessed",
        "scheduled_date": "",
        "customer_name": fields["customer_name"],
        "customer_phone": fields["customer_phone"],
        "customer_address": fields["customer_address"],
        "area": infer_area(branch, customer_address),
        "branch": branch,
        "store_name": fields["store_name"],
        "staff_name": fields["staff_name"],
        "old_product_model": fields["old_product_model"],
        "new_product_model": fields["new_product_model"],
        "product_summary": fields["product_summary"],
        "work_summary": fields["application_type"] or fields["symptom"],
        "purchase_amount_yen": purchase_amount_yen,
        "other_fee_yen": other_fee_yen,
        "memo": build_import_job_memo(fields),
        "raw_payload": {
            "source_file": str(path.name),
            "excel_fields": fields,
        },
    }
    raise_for_corrupted_text(payload, "保存対象の案件データ")
    return payload


def resolve_mail_attachment(relative_path: str) -> Path:
    candidate = (MAIL_ATTACHMENTS_DIR / relative_path).resolve()
    root = MAIL_ATTACHMENTS_DIR.resolve()
    if not candidate.is_relative_to(root) or not candidate.is_file():
        raise DashboardError("保存済みExcel添付を確認できませんでした。")
    return candidate


def materialize_mail_import_entry_jobs(
    entry: dict[str, Any],
    *,
    actor: str = "",
) -> dict[str, Any]:
    created_jobs: list[dict[str, Any]] = []
    errors: list[dict[str, str]] = []
    for attachment in entry.get("attachments", []):
        if not isinstance(attachment, dict):
            continue
        try:
            relative_path = str(attachment.get("relative_path", "")).strip()
            if not relative_path:
                raise DashboardError("添付ファイルの保存先がありません。")
            attachment_path = resolve_mail_attachment(relative_path)
            payload = extract_logistics_job_payload_from_excel(attachment_path)
            aiza_sheets = excel_workbook_preview(
                attachment_path.read_bytes(),
                attachment_path.name,
            )
            if aiza_sheets:
                raw_payload = payload.setdefault("raw_payload", {})
                if isinstance(raw_payload, dict):
                    raw_payload["aiza_sheet"] = aiza_sheets[0]
            payload.update(
                {
                    "source": "mail_import",
                    "source_mail_import_id": entry.get("id", ""),
                    "source_message_id": entry.get("message_id", ""),
                    "source_attachment_name": attachment.get("name", ""),
                    "source_attachment_path": relative_path,
                    "source_attachment_sha256": attachment.get("sha256", ""),
                }
            )
            job = save_logistics_job(payload, actor=actor)
            created_jobs.append(job)
        except DashboardError as exc:
            errors.append(
                {
                    "attachment": str(attachment.get("name", "")),
                    "error": str(exc),
                }
            )

    entry["created_jobs"] = len(created_jobs)
    entry["job_ids"] = [job.get("id", "") for job in created_jobs]
    if errors:
        entry["job_errors"] = errors
        entry["status"] = "job_error" if not created_jobs else "partial"
    elif created_jobs:
        entry["status"] = "job_created"
        entry.pop("job_errors", None)
    return {
        "created_jobs": created_jobs,
        "errors": errors,
    }


def mail_import_entry_matches_current_filters(entry: dict[str, Any]) -> bool:
    config = load_imap_mail_config()
    message = {
        "subject": entry.get("subject", ""),
        "sender": entry.get("sender", ""),
        "attachments": entry.get("attachments", []),
    }
    return imap_message_matches_filters(config, message)


def remove_logistics_jobs_for_mail_import(entry: dict[str, Any], actor: str = "") -> int:
    import_id = str(entry.get("id", ""))
    message_id = str(entry.get("message_id", ""))
    if not import_id and not message_id:
        return 0
    jobs = load_logistics_jobs()
    kept: list[dict[str, Any]] = []
    removed: list[dict[str, Any]] = []
    for job in jobs:
        if (
            (import_id and str(job.get("source_mail_import_id", "")) == import_id)
            or (message_id and str(job.get("source_message_id", "")) == message_id)
        ):
            removed.append(job)
        else:
            kept.append(job)
    if removed:
        save_logistics_jobs(kept)
        for job in removed:
            append_audit(
                "remove_filtered_mail_job",
                actor,
                str(job.get("work_order_number", "")),
                {"job_id": job.get("id", ""), "mail_import_id": import_id},
            )
    return len(removed)


def sync_mail_import_jobs(actor: str = "") -> dict[str, Any]:
    history = load_mail_imports()
    created = 0
    errors = 0
    filtered = 0
    removed = 0
    for entry in history:
        if not isinstance(entry, dict):
            continue
        if entry.get("status") == "duplicate":
            continue
        if not mail_import_entry_matches_current_filters(entry):
            removed += remove_logistics_jobs_for_mail_import(entry, actor=actor)
            entry["status"] = "filtered_out"
            entry["created_jobs"] = 0
            entry.pop("job_errors", None)
            filtered += 1
            save_mail_imports(history)
            continue
        if int(entry.get("created_jobs", 0) or 0) > 0:
            continue
        result = materialize_mail_import_entry_jobs(entry, actor=actor)
        created += len(result["created_jobs"])
        errors += len(result["errors"])
        save_mail_imports(history)
    return {
        "created_jobs": created,
        "errors": errors,
        "filtered_out": filtered,
        "removed_jobs": removed,
    }


def save_uploaded_file(
    file_name: str,
    body: bytes,
    destination_folder: Path | None = None,
) -> dict[str, Any]:
    decoded_name = unquote(file_name).strip()
    safe_name = Path(decoded_name).name
    if not safe_name or safe_name != decoded_name:
        raise DashboardError("ファイル名を確認できませんでした。")

    suffix = Path(safe_name).suffix.lower()
    if suffix not in {".csv", ".pdf"}:
        raise DashboardError("CSVまたはPDFファイルを選択してください。")
    if not body:
        raise DashboardError("選択したファイルが空です。")
    if len(body) > MAX_UPLOAD_SIZE:
        raise DashboardError("ファイルサイズは100MB以下にしてください。")

    folder = destination_folder or MANAGED_IMPORT_DIR
    folder.mkdir(parents=True, exist_ok=True)
    target = folder / safe_name
    target.write_bytes(body)
    sync_path_to_cloud(target)

    if destination_folder is None:
        save_settings(str(folder))

    return {
        "name": safe_name,
        "relative_path": safe_name,
        "type": suffix.removeprefix("."),
        "size": len(body),
    }


def compact_date_for_filename(value: str) -> str:
    normalized = normalize_date(str(value))
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", normalized):
        return normalized.replace("-", "")
    digits = "".join(character for character in normalized if character.isdigit())
    return digits or "unknown"


def safe_filename_part(value: str, fallback: str = "unknown") -> str:
    cleaned = re.sub(r"[^0-9A-Za-z_-]+", "_", str(value).strip())
    cleaned = cleaned.strip("._-")
    return cleaned or fallback


def safe_submission_filename(value: str) -> str:
    cleaned = re.sub(r'[\\/:*?"<>|\r\n]+', "_", value.strip())
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" ._")
    return cleaned or "ETC提出用領収書"


def content_disposition_header(disposition: str, file_name: str) -> str:
    ascii_name = re.sub(r"[^A-Za-z0-9_.-]+", "_", file_name).strip("._")
    ascii_name = ascii_name or "download.pdf"
    return (
        f'{disposition}; filename="{ascii_name}"; '
        f"filename*=UTF-8''{quote(file_name)}"
    )


def vehicle_period_pdf_name(
    vehicle_number: str,
    certificates: list[dict[str, Any]],
) -> str:
    dates = sorted(
        {
            str(certificate.get("date", "")).strip()
            for certificate in certificates
            if str(certificate.get("date", "")).strip()
        }
    )
    start_date = compact_date_for_filename(dates[0] if dates else "")
    end_date = compact_date_for_filename(dates[-1] if dates else "")
    safe_vehicle = safe_filename_part(vehicle_number, "vehicle")
    return f"{safe_vehicle}_{start_date}_{end_date}.pdf"


def relative_import_path(root: Path, path: Path) -> str:
    return str(path.resolve().relative_to(root.resolve())).replace("\\", "/")


def decode_csv(path: Path) -> tuple[str, str]:
    raw = path.read_bytes()
    for encoding in ("utf-8-sig", "cp932"):
        try:
            return raw.decode(encoding), encoding
        except UnicodeDecodeError:
            continue
    raise DashboardError(f"{path.name} の文字コードを判定できませんでした。")


def normalize_date(value: str) -> str:
    value = value.strip()
    if not value:
        return ""

    for pattern in ("%y/%m/%d", "%Y/%m/%d", "%Y-%m-%d"):
        try:
            parsed = datetime.strptime(value, pattern)
            return parsed.strftime("%Y-%m-%d")
        except ValueError:
            pass
    return value


def parse_japanese_date(value: str) -> str:
    compact = re.sub(r"\s+", "", value.strip())
    if not compact:
        return ""

    era_bases = {
        "令和": 2018,
        "平成": 1988,
        "昭和": 1925,
    }
    era_match = re.search(
        r"(令和|平成|昭和)(元|\d{1,2})年(\d{1,2})月(\d{1,2})日",
        compact,
    )
    if era_match:
        era, year_value, month_value, day_value = era_match.groups()
        era_year = 1 if year_value == "元" else int(year_value)
        try:
            return date(
                era_bases[era] + era_year,
                int(month_value),
                int(day_value),
            ).isoformat()
        except ValueError:
            return ""

    western_match = re.search(
        r"(\d{4})[年/\-.](\d{1,2})[月/\-.](\d{1,2})日?",
        compact,
    )
    if western_match:
        year_value, month_value, day_value = western_match.groups()
        try:
            return date(
                int(year_value),
                int(month_value),
                int(day_value),
            ).isoformat()
        except ValueError:
            return ""

    return normalize_date(compact)


def extract_inspection_details_from_text(text: str) -> dict[str, str]:
    compact = re.sub(r"\s+", "", text)
    search_area = compact
    label_index = compact.find("有効期間の満了する日")
    if label_index >= 0:
        search_area = compact[label_index : label_index + 80]

    expiration_date = parse_japanese_date(search_area)
    registration_match = re.search(
        r"(札幌|函館|旭川|室蘭|苫小牧|青森|岩手|宮城|秋田|山形|福島|水戸|土浦|つくば|宇都宮|那須|群馬|高崎|埼玉|熊谷|所沢|川口|春日部|越谷|千葉|成田|習志野|袖ヶ浦|野田|柏|品川|世田谷|練馬|杉並|板橋|足立|江東|葛飾|八王子|多摩|横浜|川崎|相模|湘南|山梨|新潟|長岡|富山|石川|福井|長野|松本|諏訪|岐阜|飛騨|静岡|浜松|沼津|伊豆|名古屋|豊橋|三河|岡崎|豊田|尾張小牧|一宮|春日井|三重|鈴鹿|滋賀|京都|大阪|なにわ|和泉|堺|神戸|姫路|奈良|和歌山|鳥取|島根|岡山|倉敷|広島|福山|山口|徳島|香川|愛媛|高知|福岡|北九州|久留米|筑豊|佐賀|長崎|佐世保|熊本|大分|宮崎|鹿児島|奄美|沖縄)\d{3}[ぁ-ん]\d{1,4}",
        compact,
    )
    vehicle_number = ""
    if registration_match:
        digits = re.findall(r"\d+", registration_match.group(0))
        vehicle_number = digits[-1] if digits else ""

    return {
        "expiration_date": expiration_date,
        "detected_vehicle_number": vehicle_number,
    }


def extract_inspection_details_from_pdf(path: Path) -> dict[str, str]:
    try:
        import pdfplumber
    except ImportError:
        return {"expiration_date": "", "detected_vehicle_number": ""}

    try:
        with pdfplumber.open(path) as document:
            text = "\n".join(page.extract_text() or "" for page in document.pages)
    except Exception:
        return {"expiration_date": "", "detected_vehicle_number": ""}

    return extract_inspection_details_from_text(text)


def parse_money(value: str) -> int:
    cleaned = value.replace(",", "").replace("￥", "").replace("¥", "").strip()
    if not cleaned:
        return 0
    try:
        return int(float(cleaned))
    except ValueError:
        return 0


def mask_card_number(value: str) -> str:
    compact = value.replace(" ", "").replace("-", "")
    visible = compact[-6:] if len(compact) >= 6 else ""
    return f"************{visible}" if visible else ""


def card_number_suffix(value: str) -> str:
    digits = "".join(character for character in value if character.isdigit())
    return digits[-6:] if len(digits) >= 6 else digits


def pdf_record_id(transaction_number: str) -> str:
    digest = hashlib.sha256(transaction_number.encode("utf-8")).hexdigest()[:16]
    return f"pdf-{digest}"


def build_vehicle_alias_map(
    registry: dict[str, dict[str, Any]] | None = None,
) -> dict[str, str]:
    registry = registry if registry is not None else load_json_store(VEHICLES_FILE, {})
    if not isinstance(registry, dict):
        return {}
    aliases = {str(vehicle_number): str(vehicle_number) for vehicle_number in registry}
    for primary_number, stored in registry.items():
        for related_number in stored.get("related_vehicle_numbers", []):
            related = str(related_number).strip()
            if related:
                aliases[related] = str(primary_number)
    return aliases


def vehicle_group_numbers(vehicle_number: str) -> set[str]:
    vehicle_number = vehicle_number.strip()
    if not vehicle_number:
        return set()
    registry = load_json_store(VEHICLES_FILE, {})
    if not isinstance(registry, dict):
        return {vehicle_number}
    alias_map = build_vehicle_alias_map(registry)
    primary = alias_map.get(vehicle_number, vehicle_number)
    stored = registry.get(primary, {})
    return {
        primary,
        *(
            str(value).strip()
            for value in stored.get("related_vehicle_numbers", [])
            if str(value).strip()
        ),
    }


def parse_card_suffixes(value: str) -> set[str]:
    suffixes: set[str] = set()
    for part in re.split(r"[,、\s/]+", value.strip()):
        digits = "".join(character for character in part if character.isdigit())
        if not digits:
            continue
        if len(digits) != 6:
            raise DashboardError(
                "ETCカード番号は下6桁をカンマ区切りで入力してください。"
            )
        suffixes.add(digits)
    return suffixes


def save_vehicle(
    vehicle_number: str,
    card_suffix: str = "",
    related_vehicle_numbers: list[str] | None = None,
    display_name: str | None = None,
    driver_name: str | None = None,
    memo: str | None = None,
    replace_card_suffixes: bool = False,
) -> dict[str, Any]:
    vehicle_number = vehicle_number.strip()
    if not vehicle_number:
        raise DashboardError("車両番号を入力してください。")
    parsed_suffixes = parse_card_suffixes(card_suffix)

    registry = load_json_store(VEHICLES_FILE, {})
    if not isinstance(registry, dict):
        registry = {}
    stored = registry.setdefault(
        vehicle_number,
        {
            "vehicle_number": vehicle_number,
            "card_suffixes": [],
            "latest_date": "",
            "display_name": "",
            "driver_name": "",
            "memo": "",
            "inspection": {},
        },
    )
    suffixes = {
        str(value) for value in stored.get("card_suffixes", []) if str(value)
    }
    if replace_card_suffixes:
        suffixes = parsed_suffixes
    else:
        suffixes.update(parsed_suffixes)
    stored["vehicle_number"] = vehicle_number
    stored["card_suffixes"] = sorted(suffixes)
    if related_vehicle_numbers is not None:
        related = sorted(
            {
                str(value).strip()
                for value in related_vehicle_numbers
                if str(value).strip() and str(value).strip() != vehicle_number
            }
        )
        numbers_claimed_by_this_vehicle = {vehicle_number, *related}
        for other_number, other_vehicle in registry.items():
            if other_number == vehicle_number:
                continue
            other_vehicle["related_vehicle_numbers"] = [
                value
                for value in other_vehicle.get("related_vehicle_numbers", [])
                if str(value) not in numbers_claimed_by_this_vehicle
            ]
        stored["related_vehicle_numbers"] = related
    else:
        stored.setdefault("related_vehicle_numbers", [])
    if display_name is not None:
        stored["display_name"] = display_name.strip()
    else:
        stored.setdefault("display_name", "")
    if driver_name is not None:
        stored["driver_name"] = driver_name.strip()
    else:
        stored.setdefault("driver_name", "")
    if memo is not None:
        stored["memo"] = memo.strip()
    else:
        stored.setdefault("memo", "")
    stored.setdefault("inspection", {})
    save_json_store(VEHICLES_FILE, registry)
    return stored


def save_vehicle_photo(
    vehicle_number: str,
    file_name: str,
    body: bytes,
) -> dict[str, Any]:
    vehicle = save_vehicle(vehicle_number)
    decoded_name = unquote(file_name).strip()
    safe_name = Path(decoded_name).name
    suffix = Path(safe_name).suffix.lower()
    if not safe_name or safe_name != decoded_name:
        raise DashboardError("画像ファイル名を確認できませんでした。")
    if suffix not in {".jpg", ".jpeg", ".png", ".webp"}:
        raise DashboardError("車両写真はJPG、PNG、WebPを選択してください。")
    if not body:
        raise DashboardError("選択した画像ファイルが空です。")
    if len(body) > MAX_PHOTO_SIZE:
        raise DashboardError("車両写真は10MB以下にしてください。")

    safe_vehicle = hashlib.sha256(vehicle_number.encode("utf-8")).hexdigest()[:20]
    VEHICLE_PHOTO_DIR.mkdir(parents=True, exist_ok=True)
    for old_path in VEHICLE_PHOTO_DIR.glob(f"{safe_vehicle}.*"):
        delete_path_from_cloud(old_path)
        old_path.unlink(missing_ok=True)
    photo_name = f"{safe_vehicle}{suffix}"
    photo_path = VEHICLE_PHOTO_DIR / photo_name
    photo_path.write_bytes(body)
    sync_path_to_cloud(photo_path)

    registry = load_json_store(VEHICLES_FILE, {})
    registry[vehicle_number]["photo_file"] = photo_name
    save_json_store(VEHICLES_FILE, registry)
    vehicle["photo_file"] = photo_name
    return vehicle


def resolve_vehicle_photo(vehicle_number: str) -> Path:
    registry = load_json_store(VEHICLES_FILE, {})
    stored = registry.get(vehicle_number, {}) if isinstance(registry, dict) else {}
    photo_name = Path(str(stored.get("photo_file", ""))).name
    if not photo_name:
        raise DashboardError("車両写真が登録されていません。")
    photo_path = (VEHICLE_PHOTO_DIR / photo_name).resolve()
    if not photo_path.is_relative_to(VEHICLE_PHOTO_DIR.resolve()) or not photo_path.is_file():
        raise DashboardError("車両写真が見つかりません。")
    return photo_path


def inspection_summary(stored: dict[str, Any]) -> dict[str, Any]:
    inspection = stored.get("inspection", {})
    if not isinstance(inspection, dict):
        inspection = {}
    pdf_file = Path(str(inspection.get("pdf_file", ""))).name
    return {
        "expiration_date": str(inspection.get("expiration_date", "")),
        "pdf_file": pdf_file,
        "pdf_url": (
            f"/api/vehicle-inspection?vehicle={quote(str(stored.get('vehicle_number', '')), safe='')}"
            if pdf_file
            else ""
        ),
        "uploaded_at": str(inspection.get("uploaded_at", "")),
        "source_name": str(inspection.get("source_name", "")),
        "detected_vehicle_number": str(inspection.get("detected_vehicle_number", "")),
        "extracted": bool(inspection.get("extracted", False)),
    }


def save_vehicle_inspection(
    vehicle_number: str,
    file_name: str,
    body: bytes,
    expiration_date: str = "",
) -> dict[str, Any]:
    vehicle = save_vehicle(vehicle_number)
    decoded_name = unquote(file_name).strip()
    safe_name = Path(decoded_name).name
    suffix = Path(safe_name).suffix.lower()
    if not safe_name or safe_name != decoded_name:
        raise DashboardError("車検PDFのファイル名を確認できませんでした。")
    if suffix != ".pdf":
        raise DashboardError("車検証はPDFファイルを選択してください。")
    if not body:
        raise DashboardError("選択した車検PDFが空です。")
    if len(body) > MAX_UPLOAD_SIZE:
        raise DashboardError("車検PDFは100MB以下にしてください。")

    safe_vehicle = hashlib.sha256(vehicle_number.encode("utf-8")).hexdigest()[:20]
    VEHICLE_INSPECTION_DIR.mkdir(parents=True, exist_ok=True)
    pdf_name = f"{safe_vehicle}.pdf"
    pdf_path = VEHICLE_INSPECTION_DIR / pdf_name
    temporary_pdf_path = VEHICLE_INSPECTION_DIR / f"{safe_vehicle}.uploading.pdf"
    temporary_pdf_path.write_bytes(body)

    extracted = extract_inspection_details_from_pdf(temporary_pdf_path)
    normalized_expiration = parse_japanese_date(expiration_date) if expiration_date else ""
    if not normalized_expiration:
        normalized_expiration = extracted.get("expiration_date", "")
    if not normalized_expiration:
        temporary_pdf_path.unlink(missing_ok=True)
        raise DashboardError(
            "車検満了日をPDFから自動取得できませんでした。満了日を入力してから保存してください。"
        )

    for old_path in VEHICLE_INSPECTION_DIR.glob(f"{safe_vehicle}.*"):
        if old_path != temporary_pdf_path:
            delete_path_from_cloud(old_path)
            old_path.unlink(missing_ok=True)
    temporary_pdf_path.replace(pdf_path)
    sync_path_to_cloud(pdf_path)

    registry = load_json_store(VEHICLES_FILE, {})
    if not isinstance(registry, dict):
        registry = {}
    stored = registry.setdefault(vehicle_number, vehicle)
    stored["inspection"] = {
        "expiration_date": normalized_expiration,
        "pdf_file": pdf_name,
        "uploaded_at": datetime.now().isoformat(timespec="seconds"),
        "source_name": safe_name,
        "detected_vehicle_number": extracted.get("detected_vehicle_number", ""),
        "extracted": bool(extracted.get("expiration_date")),
    }
    save_json_store(VEHICLES_FILE, registry)
    return stored


def update_vehicle_inspection_date(
    vehicle_number: str,
    expiration_date: str,
) -> dict[str, Any]:
    vehicle_number = vehicle_number.strip()
    normalized_expiration = parse_japanese_date(expiration_date)
    if not vehicle_number:
        raise DashboardError("車両番号を確認できませんでした。")
    if not normalized_expiration:
        raise DashboardError("車検満了日を入力してください。")

    registry = load_json_store(VEHICLES_FILE, {})
    if not isinstance(registry, dict):
        registry = {}
    stored = registry.get(vehicle_number)
    if not isinstance(stored, dict):
        raise DashboardError("車両が登録されていません。")
    inspection = stored.get("inspection", {})
    if not isinstance(inspection, dict) or not inspection.get("pdf_file"):
        raise DashboardError("車検PDFが登録されていません。先にPDFを保存してください。")

    inspection["expiration_date"] = normalized_expiration
    inspection["updated_at"] = datetime.now().isoformat(timespec="seconds")
    inspection["extracted"] = False
    stored["inspection"] = inspection
    save_json_store(VEHICLES_FILE, registry)
    return stored


def resolve_vehicle_inspection_pdf(vehicle_number: str) -> Path:
    registry = load_json_store(VEHICLES_FILE, {})
    stored = registry.get(vehicle_number, {}) if isinstance(registry, dict) else {}
    inspection = stored.get("inspection", {}) if isinstance(stored, dict) else {}
    pdf_name = Path(str(inspection.get("pdf_file", ""))).name
    if not pdf_name:
        raise DashboardError("車検PDFが登録されていません。")
    pdf_path = (VEHICLE_INSPECTION_DIR / pdf_name).resolve()
    if (
        not pdf_path.is_relative_to(VEHICLE_INSPECTION_DIR.resolve())
        or not pdf_path.is_file()
    ):
        raise DashboardError("車検PDFが見つかりません。")
    return pdf_path


def inspection_alerts(
    vehicles: list[dict[str, Any]],
    reference_date: date | None = None,
) -> list[dict[str, Any]]:
    today = reference_date or date.today()
    alerts: list[dict[str, Any]] = []
    for vehicle in vehicles:
        inspection = vehicle.get("inspection", {})
        if not isinstance(inspection, dict):
            continue
        expiration_date = str(inspection.get("expiration_date", ""))
        if not expiration_date:
            continue
        try:
            expiration = date.fromisoformat(expiration_date)
        except ValueError:
            continue
        days_remaining = (expiration - today).days
        if days_remaining <= 31:
            alerts.append(
                {
                    "vehicle_number": vehicle.get("vehicle_number", ""),
                    "display_name": vehicle.get("display_name", ""),
                    "expiration_date": expiration_date,
                    "days_remaining": days_remaining,
                    "pdf_url": inspection.get("pdf_url", ""),
                }
            )
    return sorted(alerts, key=lambda item: item["days_remaining"])


def parse_csv_file(path: Path) -> tuple[list[dict[str, Any]], str]:
    text, encoding = decode_csv(path)
    reader = csv.DictReader(text.splitlines())
    headers = set(reader.fieldnames or [])
    missing = [label for label in CSV_HEADERS.values() if label not in headers]
    if missing:
        raise DashboardError(
            f"{path.name} はETC明細CSVとして読み込めません。"
            f"不足項目: {', '.join(missing[:3])}"
        )

    records: list[dict[str, Any]] = []
    for source_row in reader:
        if not any((value or "").strip() for value in source_row.values()):
            continue

        record: dict[str, Any] = {}
        for key, header in CSV_HEADERS.items():
            value = (source_row.get(header) or "").strip()
            if key in MONEY_FIELDS:
                record[key] = parse_money(value)
            elif key in {"date_start", "date_end"}:
                record[key] = normalize_date(value)
            else:
                record[key] = value

        fingerprint_fields = [
            str(record[key])
            for key in CSV_HEADERS
            if key not in {"status"}
        ]
        record["id"] = hashlib.sha256(
            "\x1f".join(fingerprint_fields).encode("utf-8")
        ).hexdigest()[:16]
        record["card_number"] = mask_card_number(str(record["card_number"]))
        record["source_file"] = path.name
        records.append(record)

    return records, encoding


def scan_import_folder(import_folder: str) -> dict[str, Any]:
    if not import_folder:
        return {
            "records": [],
            "files": [],
            "errors": [],
            "message": "CSV取り込みフォルダーを設定してください。",
        }

    folder = Path(import_folder)
    if not folder.exists() or not folder.is_dir():
        return {
            "records": [],
            "files": [],
            "errors": ["設定された取り込みフォルダーが見つかりません。"],
            "message": "",
        }

    records_by_id: dict[str, dict[str, Any]] = {}
    files: list[dict[str, Any]] = []
    errors: list[str] = []

    for path in sorted(folder.rglob("*.csv")):
        try:
            records, encoding = parse_csv_file(path)
            files.append(
                {
                    "name": str(path.relative_to(folder)),
                    "rows": len(records),
                    "encoding": encoding,
                }
            )
            for record in records:
                records_by_id[record["id"]] = record
        except (OSError, DashboardError) as exc:
            errors.append(str(exc))

    records = sorted(
        records_by_id.values(),
        key=lambda row: (
            row["date_start"],
            row["time_start"],
            row["vehicle_number"],
        ),
        reverse=True,
    )

    message = ""
    if not files and not errors:
        message = "指定フォルダーにCSVファイルがありません。"

    return {
        "records": records,
        "files": files,
        "errors": errors,
        "message": message,
    }


def filter_records(
    records: list[dict[str, Any]],
    date_from: str = "",
    date_to: str = "",
    vehicle: str = "",
    status: str = "",
) -> list[dict[str, Any]]:
    result = []
    vehicle_numbers = vehicle_group_numbers(vehicle) if vehicle else set()
    for record in records:
        date_value = record["date_start"]
        if date_from and date_value < date_from:
            continue
        if date_to and date_value > date_to:
            continue
        if vehicle_numbers and record["vehicle_number"] not in vehicle_numbers:
            continue
        if status and record["status"] != status:
            continue
        result.append(record)
    return result


def summarize(records: list[dict[str, Any]]) -> dict[str, Any]:
    daily: dict[str, dict[str, int]] = {}
    for record in records:
        date_value = record["date_start"] or "日付不明"
        if date_value not in daily:
            daily[date_value] = {"amount": 0, "count": 0}
        daily[date_value]["amount"] += record["toll_fee"]
        daily[date_value]["count"] += 1

    dates = sorted(
        {record["date_start"] for record in records if record["date_start"]}
    )
    alias_map = build_vehicle_alias_map()
    return {
        "count": len(records),
        "amount": sum(record["toll_fee"] for record in records),
        "vehicles": len(
            {
                alias_map.get(record["vehicle_number"], record["vehicle_number"])
                for record in records
                if record["vehicle_number"]
            }
        ),
        "date_min": dates[0] if dates else "",
        "date_max": dates[-1] if dates else "",
        "daily": [
            {"date": date_value, **values}
            for date_value, values in sorted(daily.items())
        ],
    }


def summarize_vehicles(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    registry = load_json_store(VEHICLES_FILE, {})
    if not isinstance(registry, dict):
        registry = {}
    alias_map = build_vehicle_alias_map(registry)
    vehicles: dict[str, dict[str, Any]] = {}
    for record in records:
        actual_vehicle_number = record["vehicle_number"]
        if not actual_vehicle_number:
            continue
        vehicle_number = alias_map.get(actual_vehicle_number, actual_vehicle_number)
        if vehicle_number not in vehicles:
            vehicles[vehicle_number] = {
                "vehicle_number": vehicle_number,
                "display_name": "",
                "driver_name": "",
                "memo": "",
                "count": 0,
                "amount": 0,
                "latest_date": "",
                "card_suffixes": set(),
                "related_vehicle_numbers": set(),
                "inspection": {},
            }
        summary = vehicles[vehicle_number]
        summary["count"] += 1
        summary["amount"] += record["toll_fee"]
        suffix = card_number_suffix(record["card_number"])
        if suffix:
            summary["card_suffixes"].add(suffix)
        summary["latest_date"] = max(
            summary["latest_date"],
            record["date_start"],
        )

    for vehicle_number, stored in registry.items():
        primary_number = alias_map.get(vehicle_number, vehicle_number)
        if primary_number not in vehicles:
            vehicles[primary_number] = {
                "vehicle_number": primary_number,
                "display_name": "",
                "driver_name": "",
                "memo": "",
                "count": 0,
                "amount": 0,
                "latest_date": str(stored.get("latest_date", "")),
                "card_suffixes": set(),
                "related_vehicle_numbers": set(),
                "inspection": {},
            }
        vehicles[primary_number]["card_suffixes"].update(
            str(value) for value in stored.get("card_suffixes", []) if value
        )
        vehicles[primary_number]["latest_date"] = max(
            vehicles[primary_number]["latest_date"],
            str(stored.get("latest_date", "")),
        )
        if vehicle_number == primary_number:
            vehicles[primary_number]["display_name"] = str(
                stored.get("display_name", ""),
            )
            vehicles[primary_number]["driver_name"] = str(
                stored.get("driver_name", ""),
            )
            vehicles[primary_number]["memo"] = str(stored.get("memo", ""))
            vehicles[primary_number]["inspection"] = inspection_summary(
                {
                    **stored,
                    "vehicle_number": primary_number,
                }
            )
            vehicles[primary_number]["related_vehicle_numbers"].update(
                str(value)
                for value in stored.get("related_vehicle_numbers", [])
                if value
            )
            vehicles[primary_number]["photo_url"] = (
                f"/api/vehicle-photo?vehicle={quote(primary_number, safe='')}"
                if stored.get("photo_file")
                else ""
            )

    result = []
    for key in sorted(vehicles):
        item = vehicles[key]
        item["card_suffixes"] = sorted(item["card_suffixes"])
        item["related_vehicle_numbers"] = sorted(item["related_vehicle_numbers"])
        item.setdefault("photo_url", "")
        item.setdefault("inspection", {})
        result.append(item)
    return result


def register_vehicles(certificates: list[dict[str, Any]]) -> None:
    registry = load_json_store(VEHICLES_FILE, {})
    if not isinstance(registry, dict):
        registry = {}

    alias_map = build_vehicle_alias_map(registry)
    changed = False
    for certificate in certificates:
        csv_record = certificate.get("csv_record")
        if not certificate.get("matched") or not csv_record:
            continue
        actual_vehicle_number = str(csv_record.get("vehicle_number", "")).strip()
        if not actual_vehicle_number:
            continue
        vehicle_number = alias_map.get(actual_vehicle_number, actual_vehicle_number)

        stored = registry.setdefault(
            vehicle_number,
            {
                "vehicle_number": vehicle_number,
                "card_suffixes": [],
                "latest_date": "",
                "related_vehicle_numbers": [],
                "display_name": "",
                "driver_name": "",
                "memo": "",
                "inspection": {},
            },
        )
        stored.setdefault("display_name", "")
        stored.setdefault("driver_name", "")
        stored.setdefault("memo", "")
        suffixes = set(str(value) for value in stored.get("card_suffixes", []))
        suffix = card_number_suffix(str(csv_record.get("card_number", "")))
        if suffix:
            suffixes.add(suffix)
        stored["card_suffixes"] = sorted(suffixes)
        stored["latest_date"] = max(
            str(stored.get("latest_date", "")),
            str(certificate.get("date", "")),
        )
        changed = True

    if changed:
        save_json_store(VEHICLES_FILE, registry)


def store_certificates(
    certificates: list[dict[str, Any]],
    source_pdf: str,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    stored = load_json_store(CERTIFICATES_FILE, {})
    if not isinstance(stored, dict):
        stored = {}

    unique_current: dict[str, dict[str, Any]] = {}
    duplicate_dates: list[str] = []
    duplicate_count = 0
    seen_in_current: set[str] = set()

    for certificate in certificates:
        transaction_number = str(certificate.get("transaction_number", "")).strip()
        if not transaction_number:
            continue
        if transaction_number in stored or transaction_number in seen_in_current:
            duplicate_count += 1
            if certificate.get("date"):
                duplicate_dates.append(str(certificate["date"]))
        seen_in_current.add(transaction_number)

        saved_item = dict(certificate)
        saved_item["source_pdf"] = str(certificate.get("source_pdf") or source_pdf)
        stored[transaction_number] = saved_item
        unique_current[transaction_number] = certificate

    save_json_store(CERTIFICATES_FILE, stored)
    register_vehicles(list(unique_current.values()))
    duplicate_dates.sort()
    return list(unique_current.values()), {
        "count": duplicate_count,
        "date_min": duplicate_dates[0] if duplicate_dates else "",
        "date_max": duplicate_dates[-1] if duplicate_dates else "",
        "overwritten": duplicate_count > 0,
    }


def refresh_vehicle_latest_dates(
    certificates: dict[str, dict[str, Any]],
) -> None:
    registry = load_json_store(VEHICLES_FILE, {})
    if not isinstance(registry, dict):
        return

    for stored_vehicle in registry.values():
        stored_vehicle["latest_date"] = ""
    for certificate in certificates.values():
        csv_record = certificate.get("csv_record") or {}
        vehicle_number = str(csv_record.get("vehicle_number", "")).strip()
        if not vehicle_number or vehicle_number not in registry:
            continue
        registry[vehicle_number]["latest_date"] = max(
            str(registry[vehicle_number].get("latest_date", "")),
            str(certificate.get("date", "")),
        )
    save_json_store(VEHICLES_FILE, registry)


def delete_imported_pdf(relative_path: str) -> dict[str, Any]:
    settings = load_settings()
    pdf_path = resolve_pdf(settings["import_folder"], relative_path)
    managed_root = MANAGED_IMPORT_DIR.resolve()
    if not pdf_path.is_relative_to(managed_root):
        raise DashboardError(
            "安全のため、アプリのファイル選択画面から取り込んだPDFだけ削除できます。"
        )

    stored = load_json_store(CERTIFICATES_FILE, {})
    if not isinstance(stored, dict):
        stored = {}
    remaining = {
        transaction_number: certificate
        for transaction_number, certificate in stored.items()
        if str(certificate.get("source_pdf", "")) != relative_path
    }
    removed_records = len(stored) - len(remaining)

    try:
        delete_path_from_cloud(pdf_path)
        pdf_path.unlink()
    except OSError as exc:
        raise DashboardError("PDFファイルを削除できませんでした。") from exc
    save_json_store(CERTIFICATES_FILE, remaining)
    refresh_vehicle_latest_dates(remaining)
    return {
        "file": relative_path,
        "removed_records": removed_records,
    }


def data_freshness(
    records: list[dict[str, Any]],
    vehicle_summaries: list[dict[str, Any]] | None = None,
    reference_date: date | None = None,
) -> dict[str, Any]:
    today = reference_date or date.today()
    expected_date = today - timedelta(days=1)
    available_dates = []
    for record in records:
        if not record["date_start"]:
            continue
        try:
            available_dates.append(date.fromisoformat(record["date_start"]))
        except ValueError:
            continue
    latest_date = max(available_dates) if available_dates else None
    is_current = latest_date is not None and latest_date >= expected_date
    vehicle_statuses = []
    for vehicle in vehicle_summaries or []:
        latest_value = str(vehicle.get("latest_date", "")).strip()
        vehicle_latest = None
        if latest_value:
            try:
                vehicle_latest = date.fromisoformat(latest_value)
            except ValueError:
                vehicle_latest = None
        vehicle_statuses.append(
            {
                "vehicle_number": str(vehicle.get("vehicle_number", "")),
                "display_name": str(vehicle.get("display_name", "")),
                "latest_date": vehicle_latest.isoformat() if vehicle_latest else "",
                "is_current": vehicle_latest is not None
                and vehicle_latest >= expected_date,
            }
        )
    if vehicle_statuses:
        is_current = is_current and all(
            vehicle_status["is_current"] for vehicle_status in vehicle_statuses
        )
    return {
        "expected_date": expected_date.isoformat(),
        "latest_date": latest_date.isoformat() if latest_date else "",
        "is_current": is_current,
        "vehicles": vehicle_statuses,
    }


def list_pdfs(
    import_folder: str,
    allowed_paths: set[str] | None = None,
) -> list[dict[str, Any]]:
    if not import_folder:
        return []
    folder = Path(import_folder)
    if not folder.exists() or not folder.is_dir():
        return []

    result = []
    for path in sorted(folder.rglob("*.pdf")):
        try:
            relative_path = str(path.relative_to(folder)).replace("\\", "/")
            if allowed_paths is not None and relative_path not in allowed_paths:
                continue
            result.append(
                {
                    "name": path.name,
                    "relative_path": relative_path,
                    "size": path.stat().st_size,
                }
            )
        except OSError:
            continue
    return result


def grouped_pdf_paths_for_records(records: list[dict[str, Any]]) -> set[str]:
    return {
        str(record.get("source_pdf", "")).strip()
        for record in records
        if str(record.get("source_pdf", "")).strip()
    }


def organize_pdf_sources(
    import_folder: str,
    pdf_path: Path,
    original_relative_path: str,
    certificates: list[dict[str, Any]],
) -> str:
    root = Path(import_folder)
    alias_map = build_vehicle_alias_map()
    groups: dict[str, list[int]] = {}
    grouped_indexes: set[int] = set()

    for index, certificate in enumerate(certificates):
        csv_record = certificate.get("csv_record") or {}
        actual_vehicle_number = str(csv_record.get("vehicle_number", "")).strip()
        if not certificate.get("matched") or not actual_vehicle_number:
            continue
        vehicle_number = alias_map.get(actual_vehicle_number, actual_vehicle_number)
        groups.setdefault(vehicle_number, []).append(index)
        grouped_indexes.add(index)

    if not groups:
        for certificate in certificates:
            certificate["source_pdf"] = original_relative_path
        return original_relative_path

    unassigned_exists = len(grouped_indexes) < len(certificates)
    source_bytes = pdf_path.read_bytes()
    source_resolved = pdf_path.resolve()
    relative_by_vehicle: dict[str, str] = {}
    source_is_kept_as_group_pdf = False

    for vehicle_number, indexes in groups.items():
        group_certificates = [certificates[index] for index in indexes]
        target_name = vehicle_period_pdf_name(vehicle_number, group_certificates)
        target_path = (root / target_name).resolve()

        if len(groups) == 1 and not unassigned_exists:
            if target_path != source_resolved:
                original_cloud_path = pdf_path
                target_path.parent.mkdir(parents=True, exist_ok=True)
                pdf_path.replace(target_path)
                delete_path_from_cloud(original_cloud_path)
                sync_path_to_cloud(target_path)
                pdf_path = target_path
                source_resolved = target_path
        else:
            if target_path == source_resolved:
                source_is_kept_as_group_pdf = True
            else:
                target_path.parent.mkdir(parents=True, exist_ok=True)
                target_path.write_bytes(source_bytes)
                sync_path_to_cloud(target_path)

        relative_by_vehicle[vehicle_number] = relative_import_path(root, target_path)

    for index, certificate in enumerate(certificates):
        csv_record = certificate.get("csv_record") or {}
        actual_vehicle_number = str(csv_record.get("vehicle_number", "")).strip()
        vehicle_number = alias_map.get(actual_vehicle_number, actual_vehicle_number)
        certificate["source_pdf"] = relative_by_vehicle.get(
            vehicle_number,
            original_relative_path,
        )

    if len(groups) > 1 and not unassigned_exists and not source_is_kept_as_group_pdf:
        delete_path_from_cloud(pdf_path)
        pdf_path.unlink(missing_ok=True)

    first_vehicle = sorted(relative_by_vehicle)[0]
    return relative_by_vehicle[first_vehicle]


def preferred_pdf_for_vehicle(
    certificates: list[dict[str, Any]],
    vehicle_number: str,
    fallback: str,
) -> str:
    selected_group = vehicle_group_numbers(vehicle_number)
    if selected_group:
        for certificate in certificates:
            csv_record = certificate.get("csv_record") or {}
            actual_vehicle_number = str(csv_record.get("vehicle_number", "")).strip()
            if actual_vehicle_number in selected_group:
                source_pdf = str(certificate.get("source_pdf", "")).strip()
                if source_pdf:
                    return source_pdf

    for certificate in certificates:
        source_pdf = str(certificate.get("source_pdf", "")).strip()
        if source_pdf:
            return source_pdf
    return fallback


def resolve_pdf(import_folder: str, relative_path: str) -> Path:
    if not import_folder:
        raise DashboardError("取り込みフォルダーが設定されていません。")

    root = Path(import_folder).resolve()
    candidate = (root / unquote(relative_path)).resolve()
    if candidate.suffix.lower() != ".pdf":
        raise DashboardError("PDFファイルを指定してください。")
    if not candidate.is_relative_to(root):
        raise DashboardError("指定されたPDFを開けません。")
    if not candidate.is_file():
        raise DashboardError("PDFファイルが見つかりません。")
    return candidate


def normalize_match_value(value: str) -> str:
    return "".join(value.split()).replace("外", "")


def match_certificates(
    certificates: list[dict[str, Any]],
    records: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    used_record_ids: set[str] = set()
    result: list[dict[str, Any]] = []

    for certificate in certificates:
        best_record: dict[str, Any] | None = None
        best_score = -1
        for record in records:
            if record["id"] in used_record_ids:
                continue

            score = 0
            if certificate["date"] == record["date_start"]:
                score += 6
            if certificate["time"] == record["time_start"]:
                score += 3
            if certificate["fee"] == record["toll_fee"]:
                score += 3
            if (
                normalize_match_value(certificate["entry_ic"])
                == normalize_match_value(record["entry_ic"])
            ):
                score += 1
            if (
                normalize_match_value(certificate["exit_ic"])
                == normalize_match_value(record["exit_ic"])
            ):
                score += 1
            if certificate["vehicle_type"] == record["vehicle_type"]:
                score += 1

            if score > best_score:
                best_score = score
                best_record = record

        item = dict(certificate)
        if best_record is not None and best_score >= 10:
            used_record_ids.add(best_record["id"])
            item["matched"] = True
            item["match_score"] = best_score
            item["csv_record"] = {
                "id": best_record["id"],
                "vehicle_number": best_record["vehicle_number"],
                "card_number": best_record["card_number"],
                "status": best_record["status"],
                "source_file": best_record["source_file"],
            }
        else:
            item["matched"] = False
            item["match_score"] = max(best_score, 0)
            item["csv_record"] = None
        result.append(item)
    return result


def assign_certificates_to_vehicle(
    certificates: list[dict[str, Any]],
    vehicle_number: str,
) -> list[dict[str, Any]]:
    vehicle_number = vehicle_number.strip()
    if not vehicle_number:
        return certificates

    vehicle = save_vehicle(vehicle_number)
    suffixes = [
        str(value) for value in vehicle.get("card_suffixes", []) if str(value)
    ]
    masked_card = f"************{suffixes[0]}" if suffixes else ""
    result = []
    for certificate in certificates:
        item = dict(certificate)
        if not item.get("matched"):
            transaction_number = str(item.get("transaction_number", "")).strip()
            item["matched"] = True
            item["assigned_manually"] = True
            item["csv_record"] = {
                "id": pdf_record_id(transaction_number),
                "vehicle_number": vehicle_number,
                "card_number": masked_card,
                "status": "PDF登録",
                "source_file": str(item.get("source_pdf", "")),
            }
        result.append(item)
    return result


def merge_usage_records(csv_records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    records_by_id = {str(record["id"]): dict(record) for record in csv_records}
    stored = load_json_store(CERTIFICATES_FILE, {})
    if not isinstance(stored, dict):
        stored = {}

    for transaction_number, certificate in stored.items():
        csv_record = certificate.get("csv_record") or {}
        record_id = str(csv_record.get("id", "")).strip()
        if not record_id:
            continue
        if record_id in records_by_id:
            records_by_id[record_id]["transaction_number"] = transaction_number
            records_by_id[record_id]["certificate_available"] = True
            records_by_id[record_id]["source_pdf"] = str(
                certificate.get("source_pdf", "")
            )
            continue

        records_by_id[record_id] = {
            "id": record_id,
            "date_start": str(certificate.get("date", "")),
            "time_start": str(certificate.get("time", "")),
            "date_end": str(certificate.get("date", "")),
            "time_end": str(certificate.get("time", "")),
            "entry_ic": str(certificate.get("entry_ic", "")),
            "exit_ic": str(certificate.get("exit_ic", "")),
            "pre_discount_fee": int(certificate.get("fee", 0) or 0),
            "discount": 0,
            "toll_fee": int(certificate.get("fee", 0) or 0),
            "reduction_target_fee": 0,
            "postpaid_fee": int(certificate.get("fee", 0) or 0),
            "vehicle_type": str(certificate.get("vehicle_type", "")),
            "vehicle_number": str(csv_record.get("vehicle_number", "")),
            "card_number": str(csv_record.get("card_number", "")),
            "status": str(csv_record.get("status", "PDF登録")),
            "source_file": str(
                certificate.get("source_pdf")
                or csv_record.get("source_file", "PDF")
            ),
            "source_pdf": str(certificate.get("source_pdf", "")),
            "transaction_number": transaction_number,
            "certificate_available": True,
        }

    return sorted(
        records_by_id.values(),
        key=lambda row: (
            row["date_start"],
            row["time_start"],
            row["vehicle_number"],
        ),
        reverse=True,
    )


def analyze_pdf(
    relative_path: str,
    vehicle_number: str = "",
    require_csv_match: bool = False,
) -> dict[str, Any]:
    settings = load_settings()
    pdf_path = resolve_pdf(settings["import_folder"], relative_path)
    scan = scan_import_folder(settings["import_folder"])
    try:
        certificates = extract_pdf_certificates(pdf_path)
    except PdfExtractionError as exc:
        raise DashboardError(str(exc)) from exc

    matched = match_certificates(certificates, scan["records"])
    unmatched_count = sum(1 for item in matched if not item["matched"])
    if unmatched_count and not vehicle_number.strip() and not require_csv_match:
        raise DashboardError(
            "PDFだけで登録する場合は、先に左側の車両を選択してから解析してください。"
        )
    if not require_csv_match:
        matched = assign_certificates_to_vehicle(matched, vehicle_number)
    for item in matched:
        if item.get("csv_record") and not item["csv_record"].get("source_file"):
            item["csv_record"]["source_file"] = relative_path
    organized_default_pdf = organize_pdf_sources(
        settings["import_folder"],
        pdf_path,
        relative_path,
        matched,
    )
    response_pdf = preferred_pdf_for_vehicle(
        matched,
        vehicle_number,
        organized_default_pdf,
    )
    unique_matched, duplicate_summary = store_certificates(
        matched,
        source_pdf=response_pdf,
    )
    detected_vehicles = sorted(
        {
            str(item["csv_record"].get("vehicle_number", "")).strip()
            for item in unique_matched
            if item.get("matched") and item.get("csv_record")
            and str(item["csv_record"].get("vehicle_number", "")).strip()
        }
    )
    selected_group = vehicle_group_numbers(vehicle_number)
    different_vehicles = [
        value for value in detected_vehicles if value not in selected_group
    ]
    return {
        "file": response_pdf,
        "certificates": unique_matched,
        "summary": {
            "count": len(unique_matched),
            "matched": sum(1 for item in unique_matched if item["matched"]),
            "unmatched": sum(1 for item in unique_matched if not item["matched"]),
            "amount": sum(item["fee"] for item in unique_matched),
        },
        "duplicates": duplicate_summary,
        "detected_vehicles": detected_vehicles,
        "different_vehicles": different_vehicles,
        "csv_errors": scan["errors"],
    }


def crop_pdf_for_records(
    relative_path: str,
    record_ids: list[str],
) -> bytes:
    settings = load_settings()
    pdf_path = resolve_pdf(settings["import_folder"], relative_path)
    try:
        stored = load_json_store(CERTIFICATES_FILE, {})
        matched = [
            certificate
            for certificate in stored.values()
            if str(certificate.get("source_pdf", "")) == relative_path
        ] if isinstance(stored, dict) else []
        if not matched:
            scan = scan_import_folder(settings["import_folder"])
            certificates = extract_pdf_certificates(pdf_path)
            matched = match_certificates(certificates, scan["records"])
        by_record_id = {
            item["csv_record"]["id"]: item
            for item in matched
            if item["matched"] and item["csv_record"]
        }
        selected = [
            by_record_id[record_id]
            for record_id in record_ids
            if record_id in by_record_id
        ]
        return build_cropped_pdf(pdf_path, selected)
    except PdfExtractionError as exc:
        raise DashboardError(str(exc)) from exc


def submission_file_name(
    title: str,
    technician: str,
    date_from: str,
    date_to: str,
) -> str:
    date_part = "_".join(
        part
        for part in [
            compact_date_for_filename(date_from) if date_from else "",
            compact_date_for_filename(date_to) if date_to else "",
        ]
        if part
    ) or datetime.now().strftime("%Y%m%d")
    name_parts = [
        date_part,
        safe_submission_filename(title),
        safe_submission_filename(technician) if technician else "",
        "ETC提出",
    ]
    return safe_submission_filename("_".join(part for part in name_parts if part)) + ".pdf"


def normalize_submission_pdf_filename(value: str) -> str:
    raw_name = value.strip()
    if not raw_name:
        return ""
    file_name = safe_submission_filename(raw_name)
    if not file_name:
        return ""
    if not file_name.lower().endswith(".pdf"):
        file_name += ".pdf"
    stem = safe_submission_filename(Path(file_name).stem)[:120]
    return f"{stem or 'ETC提出用領収書'}.pdf"


def unique_submission_path(file_name: str) -> Path:
    SUBMISSIONS_DIR.mkdir(parents=True, exist_ok=True)
    base_name = Path(file_name).stem
    suffix = Path(file_name).suffix or ".pdf"
    candidate = SUBMISSIONS_DIR / f"{base_name}{suffix}"
    counter = 2
    while candidate.exists():
        candidate = SUBMISSIONS_DIR / f"{base_name}_{counter}{suffix}"
        counter += 1
    return candidate


def resolve_submission_pdf(relative_path: str) -> Path:
    safe_name = Path(unquote(relative_path)).name
    if not safe_name:
        raise DashboardError("提出PDFを指定してください。")
    candidate = (SUBMISSIONS_DIR / safe_name).resolve()
    if (
        candidate.suffix.lower() != ".pdf"
        or not candidate.is_relative_to(SUBMISSIONS_DIR.resolve())
        or not candidate.is_file()
    ):
        raise DashboardError("提出PDFが見つかりません。")
    return candidate


def unique_certificate_export_path(file_name: str) -> Path:
    CERTIFICATE_EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    normalized = normalize_submission_pdf_filename(file_name) or "ETC利用証明書.pdf"
    base_name = Path(normalized).stem
    suffix = Path(normalized).suffix or ".pdf"
    candidate = CERTIFICATE_EXPORTS_DIR / f"{base_name}{suffix}"
    counter = 2
    while candidate.exists():
        candidate = CERTIFICATE_EXPORTS_DIR / f"{base_name}_{counter}{suffix}"
        counter += 1
    return candidate


def resolve_certificate_export_pdf(relative_path: str) -> Path:
    safe_name = Path(unquote(relative_path)).name
    if not safe_name:
        raise DashboardError("保存した証明書PDFを指定してください。")
    candidate = (CERTIFICATE_EXPORTS_DIR / safe_name).resolve()
    if (
        candidate.suffix.lower() != ".pdf"
        or not candidate.is_relative_to(CERTIFICATE_EXPORTS_DIR.resolve())
        or not candidate.is_file()
    ):
        raise DashboardError("保存した証明書PDFが見つかりません。")
    return candidate


def create_certificate_export_pdf(payload: dict[str, Any]) -> dict[str, Any]:
    relative_path = str(payload.get("file", "")).strip()
    requested_file_name = normalize_submission_pdf_filename(
        str(payload.get("file_name", "")).strip()
    )
    raw_record_ids = payload.get("record_ids", [])
    if not relative_path:
        raise DashboardError("元PDFを指定してください。")
    if not isinstance(raw_record_ids, list):
        raise DashboardError("選択した明細を確認できませんでした。")
    record_ids = [str(value) for value in raw_record_ids[:200] if str(value).strip()]
    if not record_ids:
        raise DashboardError("保存する利用証明書を選択してください。")

    cropped_pdf = crop_pdf_for_records(relative_path, record_ids)
    output_path = unique_certificate_export_path(
        requested_file_name or "ETC利用証明書.pdf"
    )
    output_path.write_bytes(cropped_pdf)
    sync_path_to_cloud(output_path)
    saved_name = output_path.name
    return {
        "file": saved_name,
        "source_pdf": relative_path,
        "count": len(record_ids),
        "url": f"/api/certificate-export?file={quote(saved_name, safe='')}",
    }


def create_submission_pdf(payload: dict[str, Any]) -> dict[str, Any]:
    title = str(payload.get("title", "")).strip()
    technician = str(payload.get("technician", "")).strip()
    date_from = normalize_date(str(payload.get("date_from", "")).strip())
    date_to = normalize_date(str(payload.get("date_to", "")).strip())
    requested_file_name = normalize_submission_pdf_filename(
        str(payload.get("file_name", "")).strip()
    )
    raw_assignments = payload.get("assignments", [])
    if not isinstance(raw_assignments, list):
        raise DashboardError("提出する明細を確認できませんでした。")
    if not title:
        raise DashboardError("提出名を入力してください。")
    if not technician:
        raise DashboardError("技術員名を入力してください。")
    if not date_from or not date_to:
        raise DashboardError("提出期間を入力してください。")

    assignments: list[dict[str, str]] = []
    seen_record_ids: set[str] = set()
    for assignment in raw_assignments[:200]:
        if not isinstance(assignment, dict):
            continue
        record_id = str(assignment.get("record_id", "")).strip()
        work_number = str(assignment.get("work_number", "")).strip()
        if not record_id or not work_number or record_id in seen_record_ids:
            continue
        seen_record_ids.add(record_id)
        assignments.append({"record_id": record_id, "work_number": work_number})
    if not assignments:
        raise DashboardError("作業番号を割り当てたETC明細を選択してください。")

    settings = load_settings()
    scan = scan_import_folder(settings["import_folder"])
    all_records = merge_usage_records(scan["records"])
    records_by_id = {str(record["id"]): record for record in all_records}

    stored = load_json_store(CERTIFICATES_FILE, {})
    if not isinstance(stored, dict):
        stored = {}
    certificates_by_record_id = {}
    for certificate in stored.values():
        csv_record = certificate.get("csv_record") or {}
        record_id = str(csv_record.get("id", "")).strip()
        if record_id:
            certificates_by_record_id[record_id] = certificate

    items: list[dict[str, Any]] = []
    for assignment in assignments:
        record_id = assignment["record_id"]
        record = records_by_id.get(record_id)
        certificate = certificates_by_record_id.get(record_id)
        if not record or not certificate:
            raise DashboardError("提出PDFに使う利用証明書が見つかりません。")
        source_pdf = str(certificate.get("source_pdf", "")).strip()
        if not source_pdf:
            raise DashboardError("提出PDFに使う元PDFが登録されていません。")
        item = dict(certificate)
        item["source_path"] = resolve_pdf(settings["import_folder"], source_pdf)
        item["work_number"] = assignment["work_number"]
        item["fee"] = int(record.get("toll_fee", certificate.get("fee", 0)) or 0)
        items.append(item)

    try:
        pdf_body = build_submission_pdf(
            items,
            title=title,
            technician=technician,
            date_from=date_from,
            date_to=date_to,
        )
    except PdfExtractionError as exc:
        raise DashboardError(str(exc)) from exc
    output_path = unique_submission_path(
        requested_file_name or submission_file_name(title, technician, date_from, date_to)
    )
    output_path.write_bytes(pdf_body)
    sync_path_to_cloud(output_path)
    relative_path = output_path.name
    submissions = load_json_store(SUBMISSIONS_FILE, [])
    if not isinstance(submissions, list):
        submissions = []
    total_amount = sum(int(item.get("fee", 0) or 0) for item in items)
    saved = {
        "file": relative_path,
        "title": title,
        "technician": technician,
        "date_from": date_from,
        "date_to": date_to,
        "count": len(items),
        "amount": total_amount,
        "created_at": datetime.now().isoformat(timespec="seconds"),
    }
    submissions.append(saved)
    save_json_store(SUBMISSIONS_FILE, submissions[-200:])
    return {
        **saved,
        "url": f"/api/submission?file={quote(relative_path, safe='')}",
    }


def dashboard_payload(query: dict[str, list[str]]) -> dict[str, Any]:
    settings = load_settings()
    scan = scan_import_folder(settings["import_folder"])
    all_records = merge_usage_records(scan["records"])

    def first(name: str) -> str:
        return (query.get(name) or [""])[0].strip()

    filtered = filter_records(
        all_records,
        date_from=first("date_from"),
        date_to=first("date_to"),
        vehicle=first("vehicle"),
        status=first("status"),
    )
    filters_active = any(
        first(name) for name in ("date_from", "date_to", "vehicle", "status")
    )
    linked_pdf_paths = grouped_pdf_paths_for_records(filtered)
    allowed_pdf_paths = (
        linked_pdf_paths
        if linked_pdf_paths or filters_active
        else None
    )
    vehicle_summaries = summarize_vehicles(all_records)
    return {
        "settings": settings,
        "records": filtered,
        "summary": summarize(filtered),
        "available_vehicles": sorted(
            {
                vehicle["vehicle_number"]
                for vehicle in vehicle_summaries
                if vehicle["vehicle_number"]
            }
        ),
        "available_statuses": sorted(
            {row["status"] for row in all_records if row["status"]}
        ),
        "vehicle_summaries": vehicle_summaries,
        "inspection_alerts": inspection_alerts(vehicle_summaries),
        "freshness": data_freshness(all_records, vehicle_summaries),
        "files": scan["files"],
        "pdfs": list_pdfs(settings["import_folder"], allowed_pdf_paths),
        "errors": scan["errors"],
        "message": "" if all_records else scan["message"],
    }


class ETCRequestHandler(BaseHTTPRequestHandler):
    server_version = "ETCDashboard/0.1"
    current_user: dict[str, Any] = {"id": "unknown", "role": "user"}

    def log_message(self, format: str, *args: Any) -> None:
        print(f"[ETC] {self.address_string()} - {format % args}")

    def send_security_headers(self) -> None:
        self.send_header("X-Content-Type-Options", "nosniff")
        self.send_header("Referrer-Policy", "same-origin")
        self.send_header(
            "Strict-Transport-Security",
            "max-age=31536000; includeSubDomains",
        )
        self.send_header(
            "Content-Security-Policy",
            "upgrade-insecure-requests; block-all-mixed-content",
        )

    def ensure_authorized(self) -> bool:
        token = session_token_from_cookie(self.headers.get("Cookie", ""))
        user = user_from_session_token(token)
        if user is None:
            user = authenticated_user(self.headers.get("Authorization", ""))
        if user is not None:
            self.current_user = user
            return True

        body = json.dumps(
            {"error": "ログインIDまたはパスワードを確認してください。"},
            ensure_ascii=False,
        ).encode("utf-8")
        self.send_response(HTTPStatus.UNAUTHORIZED.value)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", "no-store")
        self.send_security_headers()
        self.end_headers()
        self.wfile.write(body)
        return False

    def require_admin(self) -> bool:
        if self.current_user.get("role") == "admin":
            return True
        self.send_json(
            {"error": "管理者だけが操作できます。"},
            status=HTTPStatus.FORBIDDEN,
        )
        return False

    def require_staff(self) -> bool:
        if self.current_user.get("role") != "contractor":
            return True
        self.send_json(
            {"error": "この画面は社内ユーザーだけが操作できます。"},
            status=HTTPStatus.FORBIDDEN,
        )
        return False

    def send_json(
        self,
        data: dict[str, Any],
        status: HTTPStatus = HTTPStatus.OK,
        headers: dict[str, str] | None = None,
    ) -> None:
        body = json.dumps(data, ensure_ascii=False).encode("utf-8")
        self.send_response(status.value)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", "no-store")
        self.send_security_headers()
        for name, value in (headers or {}).items():
            self.send_header(name, value)
        self.end_headers()
        self.wfile.write(body)

    def send_html(
        self,
        html: str,
        status: HTTPStatus = HTTPStatus.OK,
    ) -> None:
        body = html.encode("utf-8")
        self.send_response(status.value)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", "no-store")
        self.send_security_headers()
        self.end_headers()
        self.wfile.write(body)

    def send_bytes(
        self,
        body: bytes,
        content_type: str,
        status: HTTPStatus = HTTPStatus.OK,
        file_name: str = "selected-etc.pdf",
    ) -> None:
        self.send_response(status.value)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", "no-store")
        self.send_security_headers()
        self.send_header(
            "Content-Disposition",
            content_disposition_header("inline", file_name),
        )
        self.end_headers()
        self.wfile.write(body)

    def send_file(
        self,
        path: Path,
        content_type: str | None = None,
        file_name: str | None = None,
    ) -> None:
        try:
            body = path.read_bytes()
        except OSError:
            self.send_error(HTTPStatus.NOT_FOUND.value)
            return
        guessed_type = content_type or mimetypes.guess_type(path.name)[0]
        self.send_response(HTTPStatus.OK.value)
        self.send_header("Content-Type", guessed_type or "application/octet-stream")
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", "no-store")
        self.send_security_headers()
        if file_name:
            self.send_header(
                "Content-Disposition",
                content_disposition_header("inline", file_name),
            )
        self.end_headers()
        self.wfile.write(body)

    def public_base_url(self) -> str:
        host = self.headers.get("Host") or f"{self.server.server_address[0]}:{self.server.server_address[1]}"
        proto = self.headers.get("X-Forwarded-Proto", "http").split(",")[0].strip() or "http"
        return f"{proto}://{host}"

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path == "/api/mail/oauth/callback":
            query = parse_qs(parsed.query)
            state = (query.get("state") or [""])[0]
            code = (query.get("code") or [""])[0]
            error = (query.get("error_description") or query.get("error") or [""])[0]
            try:
                if error:
                    raise DashboardError(error)
                if not code:
                    raise DashboardError("Outlookから認証コードが返りませんでした。")
                consume_outlook_oauth_state(state)
                config = load_outlook_mail_config(
                    f"{self.public_base_url()}/api/mail/oauth/callback"
                )
                exchange_outlook_code(config, code)
                append_audit("connect_outlook", "system", config.mailbox)
                self.send_html(
                    "<!doctype html><meta charset='utf-8'>"
                    "<title>Outlook連携完了</title>"
                    "<body style='font-family:sans-serif;padding:32px'>"
                    "<h1>Outlook連携が完了しました</h1>"
                    "<p>このタブを閉じて、配送管理のメール取込画面へ戻ってください。</p>"
                    "</body>"
                )
            except DashboardError as exc:
                self.send_html(
                    "<!doctype html><meta charset='utf-8'>"
                    "<title>Outlook連携エラー</title>"
                    "<body style='font-family:sans-serif;padding:32px'>"
                    "<h1>Outlook連携に失敗しました</h1>"
                    f"<p>{html.escape(str(exc))}</p>"
                    "</body>",
                    status=HTTPStatus.BAD_REQUEST,
                )
            return

        if parsed.path in {"/", "/index.html"}:
            self.send_file(STATIC_DIR / "index.html", "text/html; charset=utf-8")
            return

        static_name = parsed.path.removeprefix("/")
        if static_name in {
            "styles.css",
            "app.js",
            "app-icon.svg",
            "app-icon-192.png",
            "app-icon-512.png",
            "spimaru.png",
            "manifest.webmanifest",
        }:
            content_type = (
                "application/manifest+json; charset=utf-8"
                if static_name == "manifest.webmanifest"
                else None
            )
            self.send_file(STATIC_DIR / static_name, content_type)
            return

        if parsed.path == "/api/session":
            if not self.ensure_authorized():
                return
            self.send_json({"user": self.current_user})
            return

        if not self.ensure_authorized():
            return

        if parsed.path == "/api/users":
            if not self.require_admin():
                return
            users = ensure_user_store()
            self.send_json(
                {"users": [public_user(user) for user in users.values()]}
            )
            return

        if parsed.path == "/api/audit-log":
            if not self.require_admin():
                return
            limit_value = (parse_qs(parsed.query).get("limit") or ["100"])[0]
            try:
                limit = max(1, min(300, int(limit_value)))
            except ValueError:
                limit = 100
            self.send_json({"entries": read_audit_log(limit)})
            return

        if parsed.path == "/api/system/database":
            if not self.require_admin():
                return
            query = parse_qs(parsed.query)
            try:
                database_config = load_database_config()
                payload = {
                    "database": redacted_database_status(database_config),
                }
                if (query.get("check") or ["0"])[0] in {"1", "true", "yes"}:
                    payload["check"] = check_database_connection(database_config)
                self.send_json(payload)
            except DatabaseConfigError as exc:
                self.send_json(
                    {"error": str(exc)},
                    status=HTTPStatus.BAD_REQUEST,
                )
            except OSError as exc:
                self.send_json(
                    {"error": f"データベース接続を確認できませんでした: {exc}"},
                    status=HTTPStatus.BAD_GATEWAY,
                )
            except Exception as exc:
                self.send_json(
                    {"error": f"データベース接続を確認できませんでした: {exc}"},
                    status=HTTPStatus.BAD_GATEWAY,
                )
            return

        if parsed.path == "/api/system/storage":
            if not self.require_admin():
                return
            query = parse_qs(parsed.query)
            try:
                self.send_json(
                    cloud_storage_status_payload(
                        (query.get("check") or ["0"])[0] in {"1", "true", "yes"}
                    )
                )
            except CloudStorageConfigError as exc:
                self.send_json(
                    {"error": str(exc)},
                    status=HTTPStatus.BAD_REQUEST,
                )
            except Exception as exc:
                self.send_json(
                    {"error": f"Google Cloud Storage connection check failed: {exc}"},
                    status=HTTPStatus.BAD_GATEWAY,
                )
            return

        if parsed.path == "/api/mail/settings":
            config = load_imap_mail_config()
            self.send_json(
                {
                    "imap": imap_redacted_status(config),
                    "auto_import": auto_import_status_payload(),
                }
            )
            return

        if parsed.path == "/api/mail/oauth/start":
            if not self.require_admin():
                return
            config = load_imap_mail_config()
            self.send_json({"imap": imap_redacted_status(config)})
            return

        if parsed.path == "/api/mail/messages":
            try:
                config = load_imap_mail_config()
                query = parse_qs(parsed.query)
                message_uid = (query.get("uid") or [""])[0].strip()
                if message_uid:
                    self.send_json(
                        {"message": imap_mailbox_message_detail(config, message_uid)}
                    )
                else:
                    try:
                        page = int((query.get("page") or ["1"])[0])
                        page_size = int((query.get("page_size") or ["20"])[0])
                    except ValueError:
                        raise DashboardError("ページ番号が正しくありません。")
                    self.send_json(imap_mailbox_messages(config, page=page, page_size=page_size))
            except DashboardError as exc:
                self.send_json({"error": str(exc)}, status=HTTPStatus.BAD_REQUEST)
            return

        if parsed.path == "/api/mail/imports":
            self.send_json(mail_imports_payload())
            return

        if parsed.path == "/api/mail/auto-import":
            self.send_json({"auto_import": auto_import_status_payload()})
            return

        if parsed.path == "/api/subcontractors":
            if not self.require_admin():
                return
            self.send_json(subcontractors_payload())
            return

        if parsed.path == "/api/return-shipments":
            if not self.require_staff():
                return
            self.send_json(return_shipments_payload())
            return

        if parsed.path == "/api/return-shipments/download":
            if not self.require_staff():
                return
            relative_path = (parse_qs(parsed.query).get("file") or [""])[0]
            try:
                xlsx_path = resolve_return_shipment_export(relative_path)
                self.send_file(
                    xlsx_path,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    xlsx_path.name,
                )
            except DashboardError as exc:
                self.send_json({"error": str(exc)}, status=HTTPStatus.NOT_FOUND)
            return

        if parsed.path == "/api/logistics/jobs":
            self.send_json(logistics_jobs_payload(parse_qs(parsed.query), self.current_user))
            return

        if parsed.path == "/api/inventory":
            try:
                payload = INVENTORY.snapshot()
                payload["user"] = self.current_user
                self.send_json(payload)
            except (InventoryError, DatabaseConfigError, OSError) as exc:
                self.send_json({"error": str(exc)}, status=HTTPStatus.BAD_REQUEST)
            return

        if parsed.path == "/api/dashboard":
            self.send_json(dashboard_payload(parse_qs(parsed.query)))
            return

        if parsed.path == "/api/pdf":
            settings = load_settings()
            relative_path = (parse_qs(parsed.query).get("file") or [""])[0]
            try:
                pdf_path = resolve_pdf(settings["import_folder"], relative_path)
                self.send_file(pdf_path, "application/pdf", pdf_path.name)
            except DashboardError as exc:
                self.send_json(
                    {"error": str(exc)},
                    status=HTTPStatus.BAD_REQUEST,
                )
            return

        if parsed.path == "/api/vehicle-photo":
            vehicle_number = (parse_qs(parsed.query).get("vehicle") or [""])[0]
            try:
                photo_path = resolve_vehicle_photo(vehicle_number)
                self.send_file(photo_path)
            except DashboardError as exc:
                self.send_json(
                    {"error": str(exc)},
                    status=HTTPStatus.NOT_FOUND,
                )
            return

        if parsed.path == "/api/user-avatar":
            user_id = (parse_qs(parsed.query).get("user") or [""])[0]
            try:
                avatar_path = resolve_user_avatar(user_id)
                self.send_file(avatar_path)
            except DashboardError as exc:
                self.send_json(
                    {"error": str(exc)},
                    status=HTTPStatus.NOT_FOUND,
                )
            return

        if parsed.path == "/api/vehicle-inspection":
            vehicle_number = (parse_qs(parsed.query).get("vehicle") or [""])[0]
            try:
                pdf_path = resolve_vehicle_inspection_pdf(vehicle_number)
                self.send_file(pdf_path, "application/pdf", pdf_path.name)
            except DashboardError as exc:
                self.send_json(
                    {"error": str(exc)},
                    status=HTTPStatus.NOT_FOUND,
                )
            return

        if parsed.path == "/api/submission":
            relative_path = (parse_qs(parsed.query).get("file") or [""])[0]
            try:
                pdf_path = resolve_submission_pdf(relative_path)
                self.send_file(pdf_path, "application/pdf", pdf_path.name)
            except DashboardError as exc:
                self.send_json(
                    {"error": str(exc)},
                    status=HTTPStatus.NOT_FOUND,
                )
            return

        if parsed.path == "/api/certificate-export":
            relative_path = (parse_qs(parsed.query).get("file") or [""])[0]
            try:
                pdf_path = resolve_certificate_export_pdf(relative_path)
                self.send_file(pdf_path, "application/pdf", pdf_path.name)
            except DashboardError as exc:
                self.send_json(
                    {"error": str(exc)},
                    status=HTTPStatus.NOT_FOUND,
                )
            return

        self.send_error(HTTPStatus.NOT_FOUND.value)

    def do_POST(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path == "/api/login":
            try:
                content_length = min(int(self.headers.get("Content-Length", "0")), 16_384)
                payload = json.loads(self.rfile.read(content_length).decode("utf-8"))
                user = authenticate_credentials(
                    str(payload.get("user_id", "")),
                    str(payload.get("password", "")),
                )
                if user is None:
                    self.send_json(
                        {"error": "ログインIDまたはパスワードを確認してください。"},
                        status=HTTPStatus.UNAUTHORIZED,
                    )
                    return
                token = create_session(user)
                append_audit("login", str(user.get("id", "")), "session")
                self.send_json(
                    {"user": user},
                    headers={"Set-Cookie": session_cookie_header(token)},
                )
            except (ValueError, json.JSONDecodeError):
                self.send_json(
                    {"error": "ログイン情報を読み取れませんでした。"},
                    status=HTTPStatus.BAD_REQUEST,
                )
            return

        if parsed.path == "/api/password-reset":
            try:
                content_length = min(int(self.headers.get("Content-Length", "0")), 16_384)
                payload = json.loads(self.rfile.read(content_length).decode("utf-8"))
                user = reset_user_password_with_key(
                    str(payload.get("user_id", "")),
                    str(payload.get("new_password", "")),
                    str(payload.get("reset_key", "")),
                )
                self.send_json({"user": user, "ok": True})
            except DashboardError as exc:
                self.send_json({"error": str(exc)}, status=HTTPStatus.BAD_REQUEST)
            except (ValueError, json.JSONDecodeError):
                self.send_json(
                    {"error": "パスワードリセット情報を読み取れませんでした。"},
                    status=HTTPStatus.BAD_REQUEST,
                )
            return

        if parsed.path == "/api/logout":
            token = session_token_from_cookie(self.headers.get("Cookie", ""))
            user = user_from_session_token(token)
            delete_session(token)
            append_audit("logout", str((user or {}).get("id", "")), "session")
            self.send_json(
                {"ok": True},
                headers={"Set-Cookie": expired_session_cookie_header()},
            )
            return

        if parsed.path == "/api/integrations/inventory/dispatch":
            integration_key = self.headers.get("X-Integration-Key", "")
            if not sagyou_integration_key_is_valid(integration_key):
                self.send_json(
                    {"error": "連携キーを確認できませんでした。"},
                    status=HTTPStatus.UNAUTHORIZED,
                )
                return
            try:
                content_length = min(int(self.headers.get("Content-Length", "0")), 16_384)
                payload = json.loads(self.rfile.read(content_length).decode("utf-8"))
                if not isinstance(payload, dict):
                    raise InventoryError("出庫情報を読み取れませんでした。")
                result = INVENTORY.dispatch_job(payload, actor="sagyou-app")
                append_audit(
                    "dispatch_inventory_from_sagyou",
                    "sagyou-app",
                    str(payload.get("job_id", "")),
                    {"reservation_id": result.get("reservation_id", "")},
                )
                self.send_json({"result": result})
            except (InventoryError, ValueError, json.JSONDecodeError) as exc:
                self.send_json({"error": str(exc)}, status=HTTPStatus.BAD_REQUEST)
            return

        if not self.ensure_authorized():
            return

        if parsed.path not in {
            "/api/settings",
            "/api/pdf/analyze",
            "/api/pdf/crop",
            "/api/pdf/delete",
            "/api/pdf/export",
            "/api/files",
            "/api/submissions",
            "/api/vehicles",
            "/api/user-avatar",
            "/api/vehicle-photo",
            "/api/vehicle-inspection",
            "/api/logistics/jobs",
            "/api/integrations/sagyou/sync",
            "/api/inventory/products",
            "/api/inventory/receive",
            "/api/inventory/return",
            "/api/inventory/dispatch",
            "/api/inventory/reservations/cancel",
            "/api/subcontractors",
            "/api/return-shipments/export",
            "/api/mail/settings",
            "/api/mail/import",
            "/api/mail/read-state",
            "/api/system/storage",
            "/api/users",
            "/api/users/password",
        }:
            self.send_error(HTTPStatus.NOT_FOUND.value)
            return

        if parsed.path.startswith("/api/inventory/"):
            try:
                content_length = min(int(self.headers.get("Content-Length", "0")), 32_768)
                raw = self.rfile.read(content_length) if content_length else b"{}"
                payload = json.loads(raw.decode("utf-8"))
                if not isinstance(payload, dict):
                    raise InventoryError("在庫情報を読み取れませんでした。")
                actor = str(self.current_user.get("id", ""))
                if parsed.path == "/api/inventory/products":
                    if not self.require_admin():
                        return
                    product_id = str(payload.get("id", "")).strip()
                    product = (
                        INVENTORY.update_product(product_id, payload, actor)
                        if product_id
                        else INVENTORY.create_product(payload, actor)
                    )
                    append_audit(
                        "update_inventory_product" if product_id else "create_inventory_product",
                        actor,
                        str(product.get("model", "")),
                        {"product_id": product.get("id", ""), "jan_code": product.get("jan_code", "")},
                    )
                    self.send_json({"product": product}, status=HTTPStatus.CREATED)
                    return
                if not self.require_staff():
                    return
                if parsed.path == "/api/inventory/receive":
                    movement = INVENTORY.add_stock(payload, actor, "receive")
                    action = "receive_inventory"
                    result: dict[str, Any] = {"movement": movement}
                elif parsed.path == "/api/inventory/return":
                    movement = INVENTORY.add_stock(payload, actor, "return")
                    action = "return_inventory"
                    result = {"movement": movement}
                elif parsed.path == "/api/inventory/dispatch":
                    dispatch = INVENTORY.dispatch_job(payload, actor)
                    action = "dispatch_inventory"
                    result = {"result": dispatch}
                else:
                    cancelled = INVENTORY.cancel_reservation(
                        str(payload.get("job_id", "")), actor
                    )
                    action = "cancel_inventory_reservation"
                    result = {"reservation": cancelled}
                append_audit(action, actor, str(payload.get("job_id") or payload.get("jan_code") or ""))
                self.send_json(result)
            except (InventoryError, DatabaseConfigError, OSError) as exc:
                self.send_json({"error": str(exc)}, status=HTTPStatus.BAD_REQUEST)
            except (ValueError, json.JSONDecodeError):
                self.send_json(
                    {"error": "在庫情報を読み取れませんでした。"},
                    status=HTTPStatus.BAD_REQUEST,
                )
            return

        if parsed.path == "/api/logistics/jobs":
            if not self.require_staff():
                return
            try:
                content_length = min(int(self.headers.get("Content-Length", "0")), 64_000)
                payload = json.loads(self.rfile.read(content_length).decode("utf-8"))
                if not isinstance(payload, dict):
                    raise DashboardError("案件情報を読み取れませんでした。")
                job = save_logistics_job(
                    payload,
                    actor=str(self.current_user.get("id", "")),
                )
                self.send_json({"job": job}, status=HTTPStatus.CREATED)
            except DashboardError as exc:
                self.send_json({"error": str(exc)}, status=HTTPStatus.BAD_REQUEST)
            except (ValueError, json.JSONDecodeError):
                self.send_json(
                    {"error": "案件情報を読み取れませんでした。"},
                    status=HTTPStatus.BAD_REQUEST,
                )
            return

        if parsed.path == "/api/subcontractors":
            if not self.require_admin():
                return
            try:
                content_length = min(int(self.headers.get("Content-Length", "0")), 64_000)
                payload = json.loads(self.rfile.read(content_length).decode("utf-8"))
                if not isinstance(payload, dict):
                    raise DashboardError("下請け業者情報を読み取れませんでした。")
                subcontractor = save_subcontractor(
                    payload,
                    actor=str(self.current_user.get("id", "")),
                )
                self.send_json({"subcontractor": subcontractor}, status=HTTPStatus.CREATED)
            except DashboardError as exc:
                self.send_json({"error": str(exc)}, status=HTTPStatus.BAD_REQUEST)
            except (ValueError, json.JSONDecodeError):
                self.send_json(
                    {"error": "下請け業者情報を読み取れませんでした。"},
                    status=HTTPStatus.BAD_REQUEST,
                )
            return

        if parsed.path == "/api/return-shipments/export":
            if not self.require_staff():
                return
            try:
                content_length = min(int(self.headers.get("Content-Length", "0")), 64_000)
                payload = json.loads(self.rfile.read(content_length).decode("utf-8"))
                if not isinstance(payload, dict):
                    raise DashboardError("配送表の発行条件を読み取れませんでした。")
                manifest = create_return_shipment_export(
                    payload,
                    actor=str(self.current_user.get("id", "")),
                )
                self.send_json({"manifest": manifest}, status=HTTPStatus.CREATED)
            except DashboardError as exc:
                self.send_json({"error": str(exc)}, status=HTTPStatus.BAD_REQUEST)
            except (ValueError, json.JSONDecodeError):
                self.send_json(
                    {"error": "配送表の発行条件を読み取れませんでした。"},
                    status=HTTPStatus.BAD_REQUEST,
                )
            return

        if parsed.path == "/api/mail/settings":
            if not self.require_admin():
                return
            try:
                content_length = min(int(self.headers.get("Content-Length", "0")), 32_000)
                payload = json.loads(self.rfile.read(content_length).decode("utf-8"))
                if not isinstance(payload, dict):
                    raise DashboardError("メール設定を読み取れませんでした。")
                imap = save_imap_mail_settings(
                    payload,
                    actor=str(self.current_user.get("id", "")),
                )
                self.send_json({"imap": imap, "auto_import": auto_import_status_payload()})
            except DashboardError as exc:
                self.send_json({"error": str(exc)}, status=HTTPStatus.BAD_REQUEST)
            except (ValueError, json.JSONDecodeError):
                self.send_json(
                    {"error": "メール設定を読み取れませんでした。"},
                    status=HTTPStatus.BAD_REQUEST,
                )
            return

        if parsed.path == "/api/integrations/sagyou/sync":
            if not self.require_staff():
                return
            try:
                content_length = min(int(self.headers.get("Content-Length", "0")), 16_384)
                payload: dict[str, Any] = {}
                if content_length:
                    raw_payload = json.loads(self.rfile.read(content_length).decode("utf-8"))
                    if isinstance(raw_payload, dict):
                        payload = raw_payload
                actor = str(self.current_user.get("id", ""))
                job_id = str(payload.get("job_id", "")).strip()
                work_order_number = str(payload.get("work_order_number", "")).strip()
                if job_id or work_order_number:
                    self.send_json(
                        {
                            "job": sync_mail_job_to_sagyou(
                                job_id,
                                actor=actor,
                                scheduled_date=str(payload.get("scheduled_date", "")),
                                work_order_number=work_order_number,
                            )
                        }
                    )
                else:
                    self.send_json(
                        sync_mail_jobs_to_sagyou(
                            actor=actor,
                            force=bool(payload.get("force", False)),
                        )
                    )
            except DashboardError as exc:
                self.send_json({"error": str(exc)}, status=HTTPStatus.BAD_REQUEST)
            except (ValueError, json.JSONDecodeError):
                self.send_json(
                    {"error": "再連携の指定を読み取れませんでした。"},
                    status=HTTPStatus.BAD_REQUEST,
                )
            return

        if parsed.path == "/api/mail/read-state":
            try:
                content_length = min(int(self.headers.get("Content-Length", "0")), 16_384)
                payload = json.loads(self.rfile.read(content_length).decode("utf-8"))
                if not isinstance(payload, dict):
                    raise DashboardError("既読状態の指定を読み取れませんでした。")
                result = set_imap_message_read_state(
                    load_imap_mail_config(),
                    str(payload.get("uid", "")).strip(),
                    unread=bool(payload.get("unread", False)),
                )
                self.send_json({"message": result})
            except DashboardError as exc:
                self.send_json({"error": str(exc)}, status=HTTPStatus.BAD_REQUEST)
            except (ValueError, json.JSONDecodeError):
                self.send_json(
                    {"error": "既読状態の指定を読み取れませんでした。"},
                    status=HTTPStatus.BAD_REQUEST,
                )
            return

        if parsed.path == "/api/mail/import":
            try:
                content_length = min(int(self.headers.get("Content-Length", "0")), 64_000)
                payload: dict[str, Any] = {}
                if content_length:
                    raw_payload = json.loads(self.rfile.read(content_length).decode("utf-8"))
                    if isinstance(raw_payload, dict):
                        payload = raw_payload
                selected_ids = payload.get("message_ids")
                selected_message_ids = None
                if isinstance(selected_ids, list):
                    selected_message_ids = {
                        str(item)
                        for item in selected_ids
                        if str(item).strip()
                    }
                result = import_imap_messages(
                    actor=str(self.current_user.get("id", "")),
                    selected_message_ids=selected_message_ids,
                )
                append_audit(
                    "import_imap_mail",
                    str(self.current_user.get("id", "")),
                    "info_order",
                    result.get("summary", {}),
                )
                self.send_json(result, status=HTTPStatus.CREATED)
            except DashboardError as exc:
                self.send_json({"error": str(exc)}, status=HTTPStatus.BAD_REQUEST)
            except (ValueError, json.JSONDecodeError):
                self.send_json(
                    {"error": "メール取込の選択情報を読み取れませんでした。"},
                    status=HTTPStatus.BAD_REQUEST,
                )
            return

        if parsed.path == "/api/files":
            try:
                content_length = int(self.headers.get("Content-Length", "0"))
                if content_length > MAX_UPLOAD_SIZE:
                    self.send_json(
                        {"error": "ファイルサイズは100MB以下にしてください。"},
                        status=HTTPStatus.REQUEST_ENTITY_TOO_LARGE,
                    )
                    return
                file_name = self.headers.get("X-File-Name", "")
                body = self.rfile.read(content_length)
                saved = save_uploaded_file(file_name, body)
                append_audit(
                    "upload_file",
                    str(self.current_user.get("id", "")),
                    str(saved.get("name", "")),
                    {"type": saved.get("type", "")},
                )
                self.send_json({"file": saved}, status=HTTPStatus.CREATED)
            except (OSError, ValueError, DashboardError) as exc:
                message = (
                    str(exc)
                    if isinstance(exc, DashboardError)
                    else "保存に失敗しました。"
                )
                self.send_json(
                    {"error": message},
                    status=HTTPStatus.BAD_REQUEST,
                )
            return

        if parsed.path == "/api/user-avatar":
            try:
                content_length = int(self.headers.get("Content-Length", "0"))
                if content_length > MAX_PHOTO_SIZE:
                    self.send_json(
                        {"error": "サムネイルは10MB以下にしてください。"},
                        status=HTTPStatus.REQUEST_ENTITY_TOO_LARGE,
                    )
                    return
                file_name = self.headers.get("X-File-Name", "")
                body = self.rfile.read(content_length)
                user = save_user_avatar(
                    str(self.current_user.get("id", "")),
                    file_name,
                    body,
                )
                self.current_user = user
                update_session_user(
                    session_token_from_cookie(self.headers.get("Cookie", "")),
                    user,
                )
                append_audit(
                    "save_user_avatar",
                    str(user.get("id", "")),
                    "profile",
                    {"file_name": unquote(file_name)},
                )
                self.send_json({"user": user}, status=HTTPStatus.CREATED)
            except (OSError, ValueError, DashboardError) as exc:
                self.send_json(
                    {"error": str(exc)},
                    status=HTTPStatus.BAD_REQUEST,
                )
            return

        if parsed.path == "/api/vehicle-photo":
            try:
                content_length = int(self.headers.get("Content-Length", "0"))
                if content_length > MAX_PHOTO_SIZE:
                    self.send_json(
                        {"error": "車両写真は10MB以下にしてください。"},
                        status=HTTPStatus.REQUEST_ENTITY_TOO_LARGE,
                    )
                    return
                vehicle_number = unquote(
                    self.headers.get("X-Vehicle-Number", "")
                ).strip()
                file_name = self.headers.get("X-File-Name", "")
                body = self.rfile.read(content_length)
                vehicle = save_vehicle_photo(vehicle_number, file_name, body)
                append_audit(
                    "save_vehicle_photo",
                    str(self.current_user.get("id", "")),
                    vehicle_number,
                    {"file_name": unquote(file_name)},
                )
                self.send_json({"vehicle": vehicle}, status=HTTPStatus.CREATED)
            except (OSError, ValueError, DashboardError) as exc:
                self.send_json(
                    {"error": str(exc)},
                    status=HTTPStatus.BAD_REQUEST,
                )
            return

        if parsed.path == "/api/vehicle-inspection":
            if not self.require_admin():
                return
            try:
                content_length = int(self.headers.get("Content-Length", "0"))
                if content_length > MAX_UPLOAD_SIZE:
                    self.send_json(
                        {"error": "車検PDFは100MB以下にしてください。"},
                        status=HTTPStatus.REQUEST_ENTITY_TOO_LARGE,
                    )
                    return
                content_type = self.headers.get("Content-Type", "")
                if content_type.startswith("application/json"):
                    payload = json.loads(
                        self.rfile.read(min(content_length, 16_384)).decode("utf-8")
                    )
                    vehicle_number = str(payload.get("vehicle_number", "")).strip()
                    expiration_date = str(payload.get("expiration_date", "")).strip()
                    vehicle = update_vehicle_inspection_date(
                        vehicle_number,
                        expiration_date,
                    )
                    append_audit(
                        "update_vehicle_inspection",
                        str(self.current_user.get("id", "")),
                        vehicle_number,
                        {
                            "expiration_date": vehicle.get("inspection", {}).get(
                                "expiration_date",
                                "",
                            ),
                        },
                    )
                    self.send_json({"vehicle": vehicle})
                    return
                vehicle_number = unquote(
                    self.headers.get("X-Vehicle-Number", "")
                ).strip()
                file_name = self.headers.get("X-File-Name", "")
                expiration_date = self.headers.get("X-Inspection-Expiration", "")
                body = self.rfile.read(content_length)
                vehicle = save_vehicle_inspection(
                    vehicle_number,
                    file_name,
                    body,
                    expiration_date,
                )
                append_audit(
                    "save_vehicle_inspection",
                    str(self.current_user.get("id", "")),
                    vehicle_number,
                    {
                        "file_name": unquote(file_name),
                        "expiration_date": vehicle.get("inspection", {}).get(
                            "expiration_date",
                            "",
                        ),
                    },
                )
                self.send_json({"vehicle": vehicle}, status=HTTPStatus.CREATED)
            except (OSError, ValueError, DashboardError) as exc:
                self.send_json(
                    {"error": str(exc)},
                    status=HTTPStatus.BAD_REQUEST,
                )
            return

        try:
            json_limit = (
                1_000_000
                if parsed.path in {"/api/submissions", "/api/pdf/export"}
                else 16_384
            )
            content_length = min(int(self.headers.get("Content-Length", "0")), json_limit)
            payload = json.loads(self.rfile.read(content_length).decode("utf-8"))
            if parsed.path == "/api/settings":
                settings = save_settings(str(payload.get("import_folder", "")).strip())
                append_audit(
                    "save_settings",
                    str(self.current_user.get("id", "")),
                    "settings",
                    {"import_folder": settings.get("import_folder", "")},
                )
                self.send_json({"settings": settings})
            elif parsed.path == "/api/vehicles":
                raw_related = payload.get("related_vehicle_numbers", [])
                if not isinstance(raw_related, list):
                    raise DashboardError("関連車番を確認できませんでした。")
                vehicle = save_vehicle(
                    str(payload.get("vehicle_number", "")),
                    str(payload.get("card_suffix", "")),
                    [str(value) for value in raw_related],
                    str(payload.get("display_name", "")),
                    str(payload.get("driver_name", "")),
                    str(payload.get("memo", "")),
                    replace_card_suffixes=bool(
                        payload.get("replace_card_suffixes", False)
                    ),
                )
                append_audit(
                    "save_vehicle",
                    str(self.current_user.get("id", "")),
                    str(vehicle.get("vehicle_number", "")),
                    {
                        "display_name": vehicle.get("display_name", ""),
                        "driver_name": vehicle.get("driver_name", ""),
                    },
                )
                self.send_json({"vehicle": vehicle})
            elif parsed.path == "/api/system/storage":
                if not self.require_admin():
                    return
                storage = save_cloud_storage_settings(
                    payload,
                    actor=str(self.current_user.get("id", "")),
                )
                self.send_json({"storage": storage})
            elif parsed.path == "/api/users":
                if not self.require_admin():
                    return
                user = create_user(
                    str(payload.get("user_id", "")),
                    str(payload.get("password", "")),
                    str(payload.get("role", "user")),
                    actor=str(self.current_user.get("id", "")),
                    contractor_code=str(payload.get("contractor_code", "")),
                    company_name=str(payload.get("company_name", "")),
                )
                self.send_json({"user": user}, status=HTTPStatus.CREATED)
            elif parsed.path == "/api/users/password":
                target_user_id = str(
                    payload.get("user_id") or self.current_user.get("id", "")
                )
                is_admin_override = (
                    self.current_user.get("role") == "admin"
                    and normalize_user_id(target_user_id)
                    != normalize_user_id(str(self.current_user.get("id", "")))
                )
                if (
                    normalize_user_id(target_user_id)
                    != normalize_user_id(str(self.current_user.get("id", "")))
                    and self.current_user.get("role") != "admin"
                ):
                    self.send_json(
                        {"error": "他のアカウントのパスワードは管理者だけが変更できます。"},
                        status=HTTPStatus.FORBIDDEN,
                    )
                    return
                user = change_user_password(
                    target_user_id,
                    str(payload.get("new_password", "")),
                    actor=str(self.current_user.get("id", "")),
                    current_password=str(payload.get("current_password", "")),
                    admin_override=is_admin_override,
                )
                self.send_json({"user": user})
            elif parsed.path == "/api/pdf/analyze":
                relative_path = str(payload.get("file", "")).strip()
                vehicle_number = str(payload.get("vehicle_number", "")).strip()
                require_csv_match = bool(payload.get("require_csv_match", False))
                result = analyze_pdf(
                    relative_path,
                    vehicle_number,
                    require_csv_match=require_csv_match,
                )
                append_audit(
                    "analyze_pdf",
                    str(self.current_user.get("id", "")),
                    relative_path,
                    {
                        "vehicle_number": vehicle_number,
                        "count": result.get("summary", {}).get("count", 0),
                    },
                )
                self.send_json(result)
            elif parsed.path == "/api/pdf/delete":
                if not self.require_admin():
                    return
                relative_path = str(payload.get("file", "")).strip()
                result = delete_imported_pdf(relative_path)
                append_audit(
                    "delete_pdf",
                    str(self.current_user.get("id", "")),
                    relative_path,
                    {"removed_records": result.get("removed_records", 0)},
                )
                self.send_json(result)
            elif parsed.path == "/api/submissions":
                result = create_submission_pdf(payload)
                append_audit(
                    "create_submission",
                    str(self.current_user.get("id", "")),
                    str(result.get("file", "")),
                    {
                        "title": result.get("title", ""),
                        "technician": result.get("technician", ""),
                        "count": result.get("count", 0),
                        "amount": result.get("amount", 0),
                    },
                )
                self.send_json(result, status=HTTPStatus.CREATED)
            elif parsed.path == "/api/pdf/export":
                result = create_certificate_export_pdf(payload)
                append_audit(
                    "export_certificates",
                    str(self.current_user.get("id", "")),
                    str(result.get("file", "")),
                    {
                        "source_pdf": result.get("source_pdf", ""),
                        "count": result.get("count", 0),
                    },
                )
                self.send_json(result, status=HTTPStatus.CREATED)
            else:
                relative_path = str(payload.get("file", "")).strip()
                file_name = normalize_submission_pdf_filename(
                    str(payload.get("file_name", "")).strip()
                ) or "selected-etc.pdf"
                raw_record_ids = payload.get("record_ids", [])
                if not isinstance(raw_record_ids, list):
                    raise DashboardError("選択した明細を確認できませんでした。")
                record_ids = [str(value) for value in raw_record_ids[:200]]
                cropped_pdf = crop_pdf_for_records(relative_path, record_ids)
                self.send_bytes(cropped_pdf, "application/pdf", file_name=file_name)
        except (ValueError, json.JSONDecodeError):
            self.send_json(
                {"error": "設定内容を読み取れませんでした。"},
                status=HTTPStatus.BAD_REQUEST,
            )
        except DashboardError as exc:
            self.send_json(
                {"error": str(exc)},
                status=HTTPStatus.BAD_REQUEST,
            )
        except CloudStorageConfigError as exc:
            self.send_json(
                {"error": str(exc)},
                status=HTTPStatus.BAD_REQUEST,
            )


class ExclusiveThreadingHTTPServer(ThreadingHTTPServer):
    # ThreadingHTTPServer enables SO_REUSEADDR. On Windows this can allow
    # multiple Python processes to listen on the same port, so requests may
    # reach an older process after a code update. Keep one authoritative API.
    allow_reuse_address = False


def main() -> None:
    parser = argparse.ArgumentParser(description="ETC利用管理ダッシュボード")
    parser.add_argument("--host", default=os.environ.get("ETC_HOST", "127.0.0.1"))
    parser.add_argument("--port", type=int, default=env_int("ETC_PORT", 8765))
    parser.add_argument(
        "--ssl-cert",
        default=os.environ.get("ETC_SSL_CERT", "").strip(),
        help="HTTPSで起動する場合の証明書ファイル",
    )
    parser.add_argument(
        "--ssl-key",
        default=os.environ.get("ETC_SSL_KEY", "").strip(),
        help="HTTPSで起動する場合の秘密鍵ファイル",
    )
    parser.add_argument(
        "--no-browser",
        action="store_true",
        help="起動時にブラウザーを開かない",
    )
    args = parser.parse_args()

    if bool(args.ssl_cert) != bool(args.ssl_key):
        parser.error("--ssl-cert と --ssl-key は両方指定してください。")

    if login_is_required() or auth_credentials() or USERS_FILE.exists():
        ensure_user_store()

    address = (args.host, args.port)
    server = ExclusiveThreadingHTTPServer(address, ETCRequestHandler)
    scheme = "https" if args.ssl_cert and args.ssl_key else "http"
    if args.ssl_cert and args.ssl_key:
        context = ssl.SSLContext(ssl.PROTOCOL_TLS_SERVER)
        context.load_cert_chain(args.ssl_cert, args.ssl_key)
        server.socket = context.wrap_socket(server.socket, server_side=True)

    url = f"{scheme}://{address[0]}:{address[1]}"
    print(f"Data folder: {APP_DATA_DIR}")
    print(
        "Login protection: "
        f"{'ON' if (login_is_required() or auth_credentials() or USERS_FILE.exists()) else 'OFF'}"
    )
    try:
        database_status = redacted_database_status(load_database_config())
        print(
            "Database backend: "
            f"{database_status['backend']} ({database_status['connection_mode']})"
        )
    except DatabaseConfigError as exc:
        print(f"Database backend: invalid ({exc})")
    try:
        storage_status = redacted_storage_status(load_cloud_storage_config())
        storage_detail = ""
        if storage_status.get("enabled"):
            storage_detail = (
                f" ({storage_status.get('bucket', '')}/"
                f"{storage_status.get('prefix', '')})"
            )
        print(f"File storage: {storage_status['backend']}{storage_detail}")
    except CloudStorageConfigError as exc:
        print(f"File storage: invalid ({exc})")
    print(f"ETC利用管理ダッシュボードを起動しました: {url}")
    print("終了するには Ctrl+C を押してください。")
    start_auto_import_worker()
    auto_status = auto_import_status_payload()
    if auto_status["enabled"]:
        print(
            "IMAP auto import: "
            f"ON ({auto_status['interval_seconds']} seconds)"
        )
    else:
        print("IMAP auto import: OFF")

    if not args.no_browser:
        threading.Timer(0.6, lambda: webbrowser.open(url)).start()

    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n終了します。")
    finally:
        stop_auto_import_worker()
        server.server_close()


if __name__ == "__main__":
    main()
